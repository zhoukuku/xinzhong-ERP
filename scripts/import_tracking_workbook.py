from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

APP_DIR = Path(__file__).resolve().parents[1]
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

import server


TARGET_FIELDS = [
    "projectName", "projectCompany", "projectNature", "projectContact", "projectPhone",
    "mainCompany", "mainContact", "mainPhone", "mainAddress", "relationGraph", "investor",
    "projectSituation", "projectProgress", "projectLocation", "recordCode", "remark",
    "projectSummary", "phoneFeedback",
]


def clean_cell(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, str) and value.startswith("="):
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"[\u00a0\u200b]+", " ", str(value)).strip()


def valid_record_code(value: str) -> bool:
    return bool(re.fullmatch(r"\d{4}-\d{6}-\d{2}-\d{2}-\d{6}", value.strip()))


def parse_section_date(value, year: int) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = clean_cell(value)
    if not text:
        return ""
    if re.fullmatch(r"\d{5}(?:\.0+)?", text):
        try:
            parsed = from_excel(float(text))
            return parsed.strftime("%Y-%m-%d")
        except (TypeError, ValueError, OverflowError):
            return ""
    matched = re.fullmatch(r"(\d{1,2})[./-](\d{1,2})", text)
    if not matched:
        return ""
    try:
        return date(year, int(matched.group(1)), int(matched.group(2))).isoformat()
    except ValueError:
        return ""


def project_category(nature: str, row: dict) -> str:
    text = nature.strip()
    if "储" in text:
        return "storage_charge"
    if "车棚" in text or "充电" in text:
        return "carport"
    if "国企" in text:
        return "roof_state"
    if "光伏" in text:
        return "roof_private"
    return server.infer_project_category(row)


def workbook_rows(path: Path, section_year: int | None = None) -> tuple[list[dict], list[dict], list[dict]]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    raw_sheet, target_sheet, met_sheet = workbook.worksheets[:3]
    section_year = section_year or datetime.now().year

    raw_rows = []
    for row_number in range(2, raw_sheet.max_row + 1):
        values = [clean_cell(raw_sheet.cell(row_number, column).value) for column in range(1, 4)]
        if values[0] or values[1] or values[2]:
            raw_rows.append({
                "projectName": values[0], "projectCompany": values[1],
                "recordCode": values[2], "rowNumber": row_number,
            })

    target_rows = []
    section_date = ""
    for row_number in range(4, target_sheet.max_row + 1):
        raw_values = [target_sheet.cell(row_number, column).value for column in range(1, 19)]
        values = [clean_cell(value) for value in raw_values]
        if not any(values):
            continue
        nonempty = [value for value in raw_values if clean_cell(value)]
        marker_date = parse_section_date(nonempty[0], section_year) if len(nonempty) == 1 else ""
        if marker_date:
            section_date = marker_date
            continue
        row = dict(zip(TARGET_FIELDS, values))
        row["category"] = project_category(row.pop("projectNature", ""), row)
        row["recordDate"] = section_date
        row["rowNumber"] = row_number
        target_rows.append(row)

    met_rows = []
    for row_number in range(4, met_sheet.max_row + 1):
        values = [clean_cell(met_sheet.cell(row_number, column).value) for column in range(1, 11)]
        if not any(values):
            continue
        met_rows.append({
            "projectName": values[0], "projectCompany": values[1],
            "projectContact": values[2], "projectPhone": values[3],
            "mainCompany": values[4], "projectSituation": values[5],
            "projectProgress": values[6], "projectLocation": values[7],
            "projectSummary": values[8], "phoneFeedback": values[9],
            "rowNumber": row_number,
        })
    return raw_rows, target_rows, met_rows


def stable_source_id(prefix: str, project_name: str, company: str, record_code: str = "") -> str:
    payload = "|".join((prefix, project_name.strip(), company.strip(), record_code.strip()))
    return f"excel-{hashlib.sha1(payload.encode('utf-8')).hexdigest()[:24]}"


def fetch_existing_leads(config: dict) -> list[dict]:
    return server.json_rows(
        "SELECT JSON_OBJECT('id',CAST(`id` AS CHAR),'projectName',`company_name`,"
        "'projectCompany',`project_unit`,'recordCode',`credit_code`) FROM `sales_leads` ORDER BY `id`",
        config,
    )


def build_indexes(rows: list[dict]) -> tuple[dict, dict]:
    by_code = {}
    by_pair = {}
    for row in rows:
        code = clean_cell(row.get("recordCode"))
        if valid_record_code(code):
            by_code.setdefault(code, row)
        pair = (clean_cell(row.get("projectName")), clean_cell(row.get("projectCompany")))
        if pair != ("", ""):
            by_pair.setdefault(pair, row)
    return by_code, by_pair


def run_statements(config: dict, statements: list[str], max_length: int = 12000) -> None:
    batch = []
    length = 0
    for statement in statements:
        if batch and length + len(statement) > max_length:
            server.run_mysql("START TRANSACTION;" + "".join(batch) + "COMMIT;", config)
            batch, length = [], 0
        batch.append(statement)
        length += len(statement)
    if batch:
        server.run_mysql("START TRANSACTION;" + "".join(batch) + "COMMIT;", config)


def plan_import(path: Path, batch_date: str) -> dict:
    config = server.load_config()
    field_map = config["query"]["field_map"]
    section_year = int(batch_date[:4]) if re.fullmatch(r"\d{4}-\d{2}-\d{2}", batch_date) else datetime.now().year
    raw_rows, target_rows, met_rows = workbook_rows(path, section_year)
    met_by_pair = {
        (row.get("projectName", ""), row.get("projectCompany", "")): row
        for row in met_rows
    }
    for row in target_rows:
        met = met_by_pair.get((row.get("projectName", ""), row.get("projectCompany", "")), {})
        for key, value in met.items():
            if key in TARGET_FIELDS and value:
                row[key] = value
    existing = fetch_existing_leads(config)
    by_code, by_pair = build_indexes(existing)

    updates = []
    inserts = []
    matched_ids = set()
    imported_index = {}
    skipped_graph_formulas = 0
    seen_new_keys = set()

    for row in target_rows:
        code = row.get("recordCode", "")
        pair = (row.get("projectName", ""), row.get("projectCompany", ""))
        match = by_code.get(code) if valid_record_code(code) else None
        match = match or by_pair.get(pair)
        mapped = {key: value for key, value in row.items() if key in field_map and value}
        if not row.get("relationGraph"):
            skipped_graph_formulas += 1
        if match:
            assignments = []
            for key, value in mapped.items():
                if key == "recordCode" and not valid_record_code(value) and valid_record_code(match.get("recordCode", "")):
                    continue
                assignments.append(f"{server.quote_identifier(field_map[key])}={server.sql_literal(value)}")
            assignments.extend([
                f"`category`={server.sql_literal(row['category'])}",
                "`review_status`=CASE WHEN `review_status`='' THEN '待人工审核' ELSE `review_status` END",
            ])
            updates.append(
                f"UPDATE `sales_leads` SET {','.join(assignments)} WHERE `id`={int(match['id'])} LIMIT 1;"
            )
            matched_ids.add(str(match["id"]))
            imported_index[pair] = str(match["id"])
            if valid_record_code(code):
                imported_index[(code, "")] = str(match["id"])
            continue

        unique_key = code if valid_record_code(code) else "|".join(pair)
        if unique_key in seen_new_keys:
            continue
        seen_new_keys.add(unique_key)
        source_id = stable_source_id("target", pair[0], pair[1], code)
        insert_values = {
            **mapped,
            "source": "Excel整理表",
            "sourceId": source_id,
            "recordDate": row.get("recordDate") or batch_date,
        }
        columns = [field_map[key] for key in insert_values if field_map.get(key)]
        values = [insert_values[key] for key in insert_values if field_map.get(key)]
        columns.extend(["category", "tungee_status", "review_status", "queued"])
        values.extend([row["category"], "已取信息" if row.get("projectPhone") or row.get("mainPhone") else "待查探迹", "待人工审核", 0])
        inserts.append(
            f"INSERT INTO `sales_leads` ({','.join(server.quote_identifier(column) for column in columns)}) "
            f"VALUES ({','.join(server.sql_literal(value) for value in values)});"
        )
        imported_index[pair] = source_id
        if valid_record_code(code):
            imported_index[(code, "")] = source_id

    met_updates = []
    for row in met_rows:
        pair = (row.get("projectName", ""), row.get("projectCompany", ""))
        match = by_pair.get(pair)
        if not match:
            continue
        assignments = [
            f"{server.quote_identifier(field_map[key])}={server.sql_literal(value)}"
            for key, value in row.items() if key in field_map and value
        ]
        if assignments:
            met_updates.append(
                f"UPDATE `sales_leads` SET {','.join(assignments)} WHERE `id`={int(match['id'])} LIMIT 1;"
            )

    existing_raw = server.json_rows(
        "SELECT JSON_OBJECT('projectName',`project_name`,'projectCompany',`project_unit`,"
        "'recordCode',`project_code`) FROM `filing_projects`",
        config,
    )
    raw_codes, raw_pairs = build_indexes(existing_raw)
    raw_inserts = []
    seen_raw = set()
    for row in raw_rows:
        code = row.get("recordCode", "")
        pair = (row.get("projectName", ""), row.get("projectCompany", ""))
        key = code if valid_record_code(code) else "|".join(pair)
        if key in seen_raw or (valid_record_code(code) and code in raw_codes) or pair in raw_pairs:
            continue
        seen_raw.add(key)
        source_id = stable_source_id("raw", pair[0], pair[1], code)
        raw_inserts.append(
            "INSERT INTO `filing_projects` (`source`,`source_project_id`,`project_code`,`project_name`,"
            "`project_unit`,`record_date`,`keyword_text`) VALUES ("
            f"'Excel整理表',{server.sql_literal(source_id)},{server.sql_literal(code)},"
            f"{server.sql_literal(pair[0])},{server.sql_literal(pair[1])},{server.sql_literal(batch_date)},"
            f"{server.sql_literal(pair[0] + ' ' + pair[1])});"
        )

    return {
        "config": config,
        "rawRows": len(raw_rows),
        "targetRows": len(target_rows),
        "metRows": len(met_rows),
        "leadUpdates": updates,
        "leadInserts": inserts,
        "metUpdates": met_updates,
        "rawInserts": raw_inserts,
        "skippedGraphFormulaRows": skipped_graph_formulas,
    }


def apply_import(plan: dict) -> dict:
    config = plan["config"]
    suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
    sales_backup = f"sales_leads_backup_{suffix}_xlsx"
    filing_backup = f"filing_projects_backup_{suffix}_xlsx"
    server.run_mysql(
        f"CREATE TABLE {server.quote_identifier(sales_backup)} LIKE `sales_leads`;"
        f"INSERT INTO {server.quote_identifier(sales_backup)} SELECT * FROM `sales_leads`;"
        f"CREATE TABLE {server.quote_identifier(filing_backup)} LIKE `filing_projects`;"
        f"INSERT INTO {server.quote_identifier(filing_backup)} SELECT * FROM `filing_projects`;",
        config,
    )
    statements = plan["leadUpdates"] + plan["leadInserts"] + plan["metUpdates"] + plan["rawInserts"]
    run_statements(config, statements)
    return {
        "salesBackup": sales_backup,
        "filingBackup": filing_backup,
        "leadUpdates": len(plan["leadUpdates"]),
        "leadInserts": len(plan["leadInserts"]),
        "metUpdates": len(plan["metUpdates"]),
        "rawInserts": len(plan["rawInserts"]),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Import the curated customer tracking workbook into Xinzhong ERP.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--batch-date", default="", help="Only use when every ungrouped row has one confirmed filing date.")
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    plan = plan_import(args.workbook, args.batch_date)
    summary = {
        "rawRows": plan["rawRows"],
        "targetRows": plan["targetRows"],
        "metRows": plan["metRows"],
        "leadUpdates": len(plan["leadUpdates"]),
        "leadInserts": len(plan["leadInserts"]),
        "metUpdates": len(plan["metUpdates"]),
        "rawInserts": len(plan["rawInserts"]),
        "skippedGraphFormulaRows": plan["skippedGraphFormulaRows"],
        "mode": "apply" if args.apply else "dry-run",
    }
    if args.apply:
        summary.update(apply_import(plan))
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
