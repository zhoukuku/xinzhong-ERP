from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

import import_tracking_workbook


class WorkbookDateTests(unittest.TestCase):
    def test_date_heading_is_not_a_project_and_applies_to_following_rows(self):
        workbook = Workbook()
        raw = workbook.active
        raw.title = "raw"
        raw.append(["name", "company", "code"])
        target = workbook.create_sheet("target")
        target.append(["title"])
        target.append([])
        target.append([])
        target.append([7.6])
        target.append(["测试光伏项目", "测试公司"])
        target.append([46212])
        target.append(["测试充电站", "另一公司"])
        met = workbook.create_sheet("met")

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "tracking.xlsx"
            workbook.save(path)
            _, rows, _ = import_tracking_workbook.workbook_rows(path, 2026)

        self.assertEqual([row["projectName"] for row in rows], ["测试光伏项目", "测试充电站"])
        self.assertEqual([row["recordDate"] for row in rows], ["2026-07-06", "2026-07-09"])

    def test_invalid_single_cell_row_remains_a_project(self):
        workbook = Workbook()
        raw = workbook.active
        raw.append(["name", "company", "code"])
        target = workbook.create_sheet("target")
        target.append([])
        target.append([])
        target.append([])
        target.append(["正常项目名称"])
        workbook.create_sheet("met")

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "tracking.xlsx"
            workbook.save(path)
            _, rows, _ = import_tracking_workbook.workbook_rows(path, 2026)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["projectName"], "正常项目名称")
        self.assertEqual(rows[0]["recordDate"], "")


if __name__ == "__main__":
    unittest.main()
