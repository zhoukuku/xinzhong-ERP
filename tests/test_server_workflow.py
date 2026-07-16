from __future__ import annotations

import json
import sys
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import server

PROJECT_CODE = ROOT.parent
if str(PROJECT_CODE) not in sys.path:
    sys.path.insert(0, str(PROJECT_CODE))
import enrich_tungee


class DummyThread:
    instances = []

    def __init__(self, target=None, args=(), daemon=None):
        self.target = target
        self.args = args
        self.daemon = daemon
        self.started = False
        self.__class__.instances.append(self)

    def start(self):
        self.started = True


class WorkflowTests(unittest.TestCase):
    def test_tungee_unlock_list_is_redirected_to_enterprise_search(self):
        class SearchInput:
            def is_displayed(self):
                return True

            def get_attribute(self, name):
                return "请输入企业名、人名或品牌名" if name == "placeholder" else ""

        class Driver:
            current_url = enrich_tungee.TUNGEE_UNLOCK_LIST
            window_handles = []

            def __init__(self):
                self.visited = []

            def get(self, url):
                self.visited.append(url)
                self.current_url = url

            def execute_script(self, script):
                return "企业查询"

            def find_elements(self, by, selector):
                return [SearchInput()] if selector == "input" else []

        driver = Driver()
        with patch.object(enrich_tungee.time, "sleep", return_value=None):
            entered = enrich_tungee.enter_enterprise_query(driver, timeout_sec=1)
        self.assertTrue(entered)
        self.assertEqual(driver.visited, [enrich_tungee.TUNGEE_SEARCH])

    def test_tungee_unlock_action_accepts_current_spaced_labels(self):
        self.assertTrue(enrich_tungee.is_unlock_action_text("解 锁"))
        self.assertTrue(enrich_tungee.is_unlock_action_text("解锁企业并查看"))
        self.assertFalse(enrich_tungee.is_unlock_action_text("解锁列表"))

    def test_tungee_company_name_matching_normalizes_parentheses(self):
        ascii_name = "晖跃新能源(深圳)有限公司"
        chinese_name = "晖跃新能源（深圳）有限公司"
        self.assertEqual(
            enrich_tungee.normalize_company_name_for_match(ascii_name),
            enrich_tungee.normalize_company_name_for_match(chinese_name),
        )
        self.assertEqual(enrich_tungee.company_search_candidates(ascii_name)[:2], [ascii_name, chinese_name])

    def test_tungee_fuzzy_company_match_prefers_same_city_and_brand(self):
        target = "河源市睿轩实业有限公司"
        same_city = "河源市睿轩新能源科技有限公司"
        other_city = "上海睿轩实业有限公司"

        self.assertGreaterEqual(enrich_tungee.company_name_match_score(target, same_city), 0.78)
        self.assertEqual(enrich_tungee.company_name_match_score(target, other_city), 0.0)

    def test_identical_filing_requests_share_one_running_job(self):
        server.FILING_JOBS.clear()
        started = []

        class FilingThread(DummyThread):
            def start(self):
                self.started = True
                started.append(self.args[0])

        with patch.object(server.threading, "Thread", FilingThread), patch.object(server, "persist_filing_job"):
            first = server.create_filing_job("2026-07-13", "光伏 充电桩")
            second = server.create_filing_job("2026-07-13", "充电桩 光伏")

        self.assertEqual(first["id"], second["id"])
        self.assertTrue(second["shared"])
        self.assertEqual(len(started), 1)
        server.FILING_JOBS.clear()

    def test_filing_record_maps_five_source_fields(self):
        row = server.record_to_filing_row(
            {
                "credit_code": "2607-test",
                "company_name": "不应作为项目名",
                "project_unit": "测试项目单位",
                "project_desc": "测试光伏项目",
                "project_status": "备案",
                "record_date": "2026-07-08",
            }
        )
        self.assertEqual(row["projectCode"], "2607-test")
        self.assertEqual(row["projectName"], "测试光伏项目")
        self.assertEqual(row["projectUnit"], "测试项目单位")
        self.assertEqual(row["projectType"], "备案")
        self.assertEqual(row["recordDate"], "2026-07-08")

    def test_raw_filing_fields_are_normalized_before_intake(self):
        row = server.project_to_lead_row({
            "projectName": "测试光伏项目", "projectUnit": "测试备案公司有限公司",
            "projectCode": "2607-test", "projectType": "备案", "recordDate": "2026-07-08",
        })
        self.assertEqual(row["projectCompany"], "测试备案公司有限公司")
        self.assertEqual(row["recordCode"], "2607-test")
        self.assertEqual(row["sourceId"], "2607-test")
        self.assertEqual(row["projectProgress"], "备案")

    def test_doubao_mapping_never_writes_manual_fields(self):
        result = {
            "公司性质": "民企",
            "最终主体公司": "深圳恒盛能源有限公司",
            "股权穿透": {"最终控股方": "恒盛投资集团", "实际控制人": "张三"},
            "关键对接人": [{"姓名": "李四", "职务": "法人", "电话": "13800138000"}],
            "项目情况": "100kW 屋顶光伏",
            "项目进展": "已备案",
            "项目所在地": "深圳市宝安区",
        }
        update = server.doubao_result_to_project(result, "1", [])
        self.assertNotIn("projectSummary", update)
        self.assertNotIn("phoneFeedback", update)
        self.assertEqual(update["mainCompany"], "深圳恒盛能源有限公司")
        self.assertEqual(update["mainContact"], "张三")
        self.assertEqual(update["mainPhone"], "13800138000")
        self.assertNotIn("projectContact", update)
        self.assertNotIn("projectPhone", update)
        self.assertEqual(update["projectLocation"], "深圳市宝安区")
        self.assertNotIn("projectAddress", update)

    def test_project_location_is_inferred_from_filing_title(self):
        title = "深圳市宝安区沙井街道金沙一路金沙工业园40kWp光伏发电项目"
        self.assertEqual(server.infer_project_location(title), "深圳市宝安区沙井街道金沙一路金沙工业园")

    def test_project_title_is_never_a_tungee_enterprise_target(self):
        self.assertFalse(server.looks_like_enterprise_name("华南城西苑地下停车场7KW交流充电站工程"))
        self.assertFalse(server.looks_like_enterprise_name("金沙工业园40kWp光伏发电项目"))
        self.assertTrue(server.looks_like_enterprise_name("深圳市高卓新能源有限公司"))

    def test_legacy_subject_prefixes_are_removed_before_tungee(self):
        self.assertEqual(server.normalize_enterprise_target("由深圳市佳霖新能源科技有限公司"), "深圳市佳霖新能源科技有限公司")
        self.assertEqual(server.normalize_enterprise_target("一级股东（深圳市丰赫粤科技有限公司"), "深圳市丰赫粤科技有限公司")
        self.assertEqual(server.normalize_enterprise_target("实际业主均为深圳市曦耀新能源有限公司"), "深圳市曦耀新能源有限公司")
        self.assertEqual(server.normalize_enterprise_target("晖跃新能源(深圳)有限公司"), "晖跃新能源（深圳）有限公司")
        self.assertEqual(
            server.enterprise_name_key("晖跃新能源(深圳)有限公司"),
            server.enterprise_name_key("晖跃新能源（深圳）有限公司"),
        )

    def test_tungee_selenium_stacktrace_is_not_exposed_to_employees(self):
        error = "Message: stale element reference: stale element not found\nStacktrace:\nchromedriver!foo"
        self.assertEqual(server.friendly_tungee_error(error), "探迹页面刷新导致元素失效（技术异常，可重试）")
        self.assertNotIn("Stacktrace", server.friendly_tungee_error(error))

    def test_successful_parenthesis_variant_archives_older_equivalent_failure(self):
        statements = []
        with (
            patch.object(server, "json_rows", return_value=[
                {"id": "431", "companyName": "晖跃新能源(深圳)有限公司"},
                {"id": "432", "companyName": "其他新能源有限公司"},
            ]),
            patch.object(server, "run_mysql", side_effect=lambda sql, config=None: statements.append(sql) or ""),
        ):
            archived = server.archive_equivalent_failed_tasks(603, "晖跃新能源（深圳）有限公司", {})

        self.assertEqual(archived, 1)
        self.assertIn("IN (431)", statements[0])
        self.assertIn("历史失败已归档", statements[0])

    def test_old_intake_recovers_project_company_from_filing_cache(self):
        project = {"sourceId": "2607-test", "projectName": "测试项目", "projectCompany": ""}
        with patch.object(server, "json_rows", return_value=[{"projectUnit": "测试备案公司有限公司", "projectName": "测试项目"}]):
            hydrated = server.hydrate_intake_project(project, {})
        self.assertEqual(hydrated["projectCompany"], "测试备案公司有限公司")

    def test_lead_with_new_record_code_falls_back_to_name_and_company_match(self):
        config = {
            "query": {
                "table": "sales_leads",
                "field_map": {
                    "projectName": "company_name", "projectCompany": "project_unit",
                    "recordCode": "credit_code", "recordDate": "record_date",
                    "source": "source", "sourceId": "source_project_id",
                },
            },
        }
        statements = []
        matches = iter([[], [{
            "dbId": "7", "source": "caitoubiao", "sourceId": "old", "projectName": "测试光伏项目",
        }]])
        with (
            patch.object(server, "load_config", return_value=config),
            patch.object(server, "ensure_sales_workflow_columns"),
            patch.object(server, "json_rows", side_effect=lambda *args, **kwargs: next(matches)),
            patch.object(server, "run_mysql", side_effect=lambda sql, config=None: statements.append(sql) or ""),
        ):
            result = server.save_leads([{
                "projectName": "测试光伏项目", "projectCompany": "测试公司有限公司",
                "recordCode": "2607-440300-04-01-000001", "recordDate": "2026-07-08",
                "source": "深圳投资项目公示", "sourceId": "2607-440300-04-01-000001",
            }], auto_investigate=False)

        self.assertEqual(result["linked"], 1)
        self.assertEqual(result["inserted"], 0)
        self.assertTrue(any("`credit_code`=COALESCE" in sql for sql in statements))

    def test_doubao_stage_writes_verified_company_to_tungee_target(self):
        result = {"建议探迹查询主体公司": "东莞市凤顺新能源投资有限公司"}
        db_rows = [{"intakeId": "8", "companyName": "华南城西苑地下停车场7KW交流充电站工程", "existing": "{}"}]
        statements = []
        with (
            patch.object(server, "load_config", return_value={}),
            patch.object(server, "json_rows", return_value=db_rows),
            patch.object(server, "run_mysql", side_effect=lambda sql, config=None: statements.append(sql) or ""),
            patch.object(server, "start_auto_dispatch"),
        ):
            server.finish_doubao_stage("3", result)
        self.assertIn("`company_name`='东莞市凤顺新能源投资有限公司'", statements[0])

    def test_doubao_stage_falls_back_to_authoritative_filing_company(self):
        result = {"最终主体公司": "[未确认]", "建议探迹查询主体公司": "[未确认]"}
        db_rows = [{
            "intakeId": "8",
            "companyName": "珑门名苑地下停车场充电站",
            "projectJson": json.dumps({
                "projectName": "珑门名苑地下停车场充电站",
                "projectCompany": "深圳雁茗企业管理中心（有限合伙）",
            }, ensure_ascii=False),
            "existing": "{}",
        }]
        statements = []
        with (
            patch.object(server, "load_config", return_value={}),
            patch.object(server, "json_rows", return_value=db_rows),
            patch.object(server, "run_mysql", side_effect=lambda sql, config=None: statements.append(sql) or ""),
            patch.object(server, "start_auto_dispatch"),
        ):
            server.finish_doubao_stage("3", result)

        self.assertIn("`company_name`='深圳雁茗企业管理中心（有限合伙）'", statements[0])
        self.assertIn('"tungee_target_source": "filing_company_fallback"', statements[0])
        self.assertNotIn("珑门名苑地下停车场充电站'", statements[0])

    def test_tungee_shareholder_parser_finds_controlling_company(self):
        text = (
            "股东信息1股权穿透图\n序号\n股东\n持股比例\n"
            "1\n凤顺\n东莞市凤顺新能源投资有限公司\n控股股东\n100%\n100万人民币\n"
            "主要人员2\n许培忠\n监事"
        )
        self.assertEqual(
            enrich_tungee.parse_shareholders(text),
            [{"name": "东莞市凤顺新能源投资有限公司", "ratio": 100.0, "type": "company", "controlling": True}],
        )

    def test_tungee_shareholder_parser_finds_controlling_person(self):
        text = (
            "股东信息1股权穿透图\n序号\n股东\n持股比例\n"
            "1\n李\n李岩\n现任职 1\n控股股东\n100%\t100万人民币\n"
            "主要人员1\n李岩\n董事"
        )
        self.assertEqual(
            enrich_tungee.parse_shareholders(text),
            [{"name": "李岩", "ratio": 100.0, "type": "person", "controlling": True}],
        )

    def test_tungee_contact_ranking_prefers_shareholder_matched_second_hot_contact(self):
        candidates = [
            {"phone": "13800138000", "name": "王**", "tags": "推荐", "hot_level": 3, "index": 0},
            {"phone": "13900139000", "name": "李**", "tags": "来自年报", "hot_level": 2, "index": 1},
        ]
        selected = enrich_tungee.choose_contact_candidate(candidates, ["李岩"])
        self.assertEqual(selected["phone"], "13900139000")
        self.assertEqual(selected["match_type"], "股东或法人姓氏匹配")

    def test_tungee_contact_ranking_rejects_suspected_empty_or_bookkeeping_phone(self):
        candidates = [
            {"phone": "13800138000", "name": "李**", "tags": "疑似空号 疑似代理记账", "hot_level": 3, "index": 0},
            {"phone": "13900139000", "name": "王**", "tags": "近期收录", "hot_level": 1, "index": 1},
        ]
        self.assertEqual(enrich_tungee.choose_contact_candidate(candidates, ["李岩"])["phone"], "13900139000")

    def test_tungee_shareholder_parser_does_not_treat_equity_pledge_as_person(self):
        text = (
            "股东信息1\n序号\n股东\n持股比例\n1\n张文举\n控股股东\n80%\n"
            "2\n股权出质\n20%\n主要人员1"
        )
        self.assertEqual(
            enrich_tungee.parse_shareholders(text),
            [{"name": "张文举", "ratio": 80.0, "type": "person", "controlling": True}],
        )

    def test_tungee_mapping_does_not_overwrite_filing_number(self):
        update = server.tungee_result_to_project(
            {
                "project_company_address": "深圳市项目单位地址",
                "project_company_credit_code": "91440300TESTCREDIT",
                "main_company": "东莞市控股公司",
                "main_address": "东莞市控股公司地址",
            }
        )
        self.assertNotIn("recordCode", update)
        self.assertEqual(update["projectAddress"], "深圳市项目单位地址")
        self.assertEqual(update["mainAddress"], "东莞市控股公司地址")

    def test_tungee_mapping_writes_project_company_contact_pair(self):
        update = server.tungee_result_to_project(
            {"project_company_contact": "刘**", "project_company_phone": "13800138000"}
        )
        self.assertEqual(update["projectContact"], "刘**")
        self.assertEqual(update["projectPhone"], "13800138000")

    def test_filing_company_fallback_is_visible_in_employee_equity_text(self):
        result = server.annotate_tungee_target_source(
            {"relation_graph": "备案项目单位：测试公司", "remark": "探迹股权链"},
            {"tungee_target_source": "filing_company_fallback"},
        )

        self.assertIn("备案公司作为探迹股权穿透起点", result["relation_graph"])
        self.assertIn("不是已确认的最终主体", result["remark"])

    def test_tungee_equity_chain_is_employee_readable_text(self):
        class Driver:
            current_url = ""

            def get(self, url):
                self.current_url = url

        filing = {
            "company_name": "深圳市项目公司有限公司",
            "address": "深圳市项目单位地址",
            "shareholders": [{"name": "东莞市控股公司有限公司", "ratio": 100.0, "type": "company"}],
        }
        parent = {
            "company_name": "东莞市控股公司有限公司",
            "contact_person": "张三",
            "phone": "13800138000",
        }
        with (
            patch.object(enrich_tungee, "search_company", side_effect=["filing-url", "parent-url"]),
            patch.object(enrich_tungee, "extract_detail", side_effect=[filing, parent]),
        ):
            result = enrich_tungee.investigate_company_chain(Driver(), "深圳市项目公司有限公司")
        self.assertIn("一级股东：东莞市控股公司有限公司（持股 100%）", result["relation_graph"])
        self.assertIn("实际控制关系需人工复核", result["relation_graph"])

    def test_strong_category_rules(self):
        self.assertEqual(server.infer_project_category({"projectName": "园区光储充一体化项目"}), "storage_charge")
        self.assertEqual(server.infer_project_category({"projectName": "停车场超充站"}), "carport")
        self.assertEqual(server.infer_project_category({"projectName": "华能屋顶光伏项目"}), "roof_state")

    def test_filing_keywords_match_power_unit_variants(self):
        rows = [
            {"projectName": "厂房 499.84Kwp 分布式光伏发电项目", "projectUnit": "测试公司"},
            {"projectName": "屋顶 33 千瓦时 储能项目", "projectUnit": "测试公司"},
        ]
        self.assertEqual(len(server.filter_filing_rows(rows, ["kwp"])), 1)
        self.assertEqual(len(server.filter_filing_rows(rows, ["kwh"])), 1)

    def test_filing_legacy_cache_fallback_initializes_all_rows(self):
        config = {
            "mysql": {"database": "xinzhong_erp"},
            "query": {
                "table": "sales_leads",
                "field_map": {
                    "recordCode": "record_code",
                    "projectName": "project_name",
                    "projectCompany": "project_company",
                    "projectProgress": "project_progress",
                    "recordDate": "record_date",
                    "source": "source",
                    "sourceId": "source_id",
                },
            },
        }
        legacy_row = {
            "projectCode": "2607-test",
            "projectName": "测试光伏项目",
            "projectUnit": "测试新能源有限公司",
            "projectType": "备案",
            "recordDate": "2026-07-09",
            "source": "",
            "sourceId": "",
        }
        empty_cache = {"rows": [], "allRows": [], "availableDates": [], "hasDate": False, "totalCount": 0}
        with (
            patch.object(server, "fetch_filing_cache", return_value=empty_cache),
            patch.object(server, "crawl_filings_live", side_effect=RuntimeError("官网暂不可用")),
            patch.object(server, "load_config", return_value=config),
            patch.object(server, "run_mysql", return_value=json.dumps(legacy_row, ensure_ascii=False)),
            patch.object(server, "fetch_available_dates", return_value=["2026-07-09"]),
        ):
            payload = server.fetch_filings("2026-07-09", "光伏")

        self.assertEqual(payload["readMode"], "local_cache")
        self.assertEqual(payload["totalCount"], 1)
        self.assertEqual(payload["count"], 1)
        self.assertEqual(payload["allRows"][0]["projectUnit"], "测试新能源有限公司")

    def test_filing_failure_with_empty_local_cache_is_not_reported_as_zero_success(self):
        config = {
            "mysql": {"database": "xinzhong_erp"},
            "query": {
                "table": "sales_leads",
                "field_map": {
                    "recordCode": "record_code", "projectName": "project_name",
                    "projectCompany": "project_company", "projectProgress": "project_progress",
                    "recordDate": "record_date", "source": "source", "sourceId": "source_id",
                },
            },
        }
        empty_cache = {"rows": [], "allRows": [], "availableDates": [], "hasDate": False, "totalCount": 0}
        with (
            patch.object(server, "fetch_filing_cache", return_value=empty_cache),
            patch.object(server, "crawl_filings_live", side_effect=RuntimeError("Chrome failed to start")),
            patch.object(server, "load_config", return_value=config),
            patch.object(server, "run_mysql", return_value=""),
            patch.object(server, "fetch_available_dates", return_value=[]),
        ):
            with self.assertRaisesRegex(server.ApiError, "本地没有 2026-07-06 的可用数据"):
                server.fetch_filings("2026-07-06", "")

    def test_recover_filing_jobs_marks_interrupted_jobs_failed(self):
        statements = []
        with (
            patch.object(server, "load_config", return_value={}),
            patch.object(server, "ensure_filing_jobs_table"),
            patch.object(server, "run_mysql", side_effect=lambda sql, config=None: statements.append(sql) or ""),
        ):
            server.recover_filing_jobs()
        self.assertIn("WHERE `status` IN ('queued','running')", statements[0])
        self.assertIn("`status`='failed'", statements[0])

    def test_filing_live_rows_survive_cache_write_failure(self):
        live_rows = [{
            "projectCode": "2607-live",
            "projectName": "测试屋顶光伏项目",
            "projectUnit": "实时新能源有限公司",
            "projectType": "备案",
            "recordDate": "2026-07-09",
        }]
        empty_cache = {"rows": [], "allRows": [], "availableDates": [], "hasDate": False, "totalCount": 0}
        with (
            patch.object(server, "fetch_filing_cache", return_value=empty_cache),
            patch.object(server, "crawl_filings_live", return_value=(live_rows, ["采集成功"])),
            patch.object(server, "store_filing_rows", side_effect=server.ApiError("缓存写入失败")),
        ):
            payload = server.fetch_filings("2026-07-09", "光伏")

        self.assertEqual(payload["readMode"], "live_crawler")
        self.assertEqual(payload["totalCount"], 1)
        self.assertEqual(payload["count"], 1)
        self.assertEqual(payload["cacheWriteError"], "缓存写入失败")

    def test_filing_cache_deduplicates_excel_copy_and_prefers_official_row(self):
        rows = [
            {
                "projectCode": "2607-duplicate",
                "projectName": "重复光伏项目",
                "projectUnit": "测试新能源有限公司",
                "source": "Excel整理表",
            },
            {
                "projectCode": "2607-duplicate",
                "projectName": "重复光伏项目",
                "projectUnit": "测试新能源有限公司",
                "source": server.FILING_SOURCE_NAME,
                "sourceUrl": server.FILING_SOURCE_URL,
            },
        ]

        result = server.deduplicate_filing_rows(rows)

        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["source"], server.FILING_SOURCE_NAME)
        self.assertEqual(result[0]["sourceUrl"], server.FILING_SOURCE_URL)

    def test_filing_cache_write_batches_large_result_sets(self):
        rows = [{
            "projectCode": f"2607-{index:04d}",
            "projectName": f"测试光伏项目{index}" + "长名称" * 80,
            "projectUnit": "测试新能源有限公司",
            "projectType": "备案",
            "recordDate": "2026-07-09",
            "sourceId": f"source-{index}",
        } for index in range(90)]
        statements = []
        with (
            patch.object(server, "load_config", return_value={}),
            patch.object(server, "ensure_filing_cache_table"),
            patch.object(server, "run_mysql", side_effect=lambda sql, config=None: statements.append(sql) or ""),
        ):
            server.store_filing_rows(rows, "")

        self.assertGreater(len(statements), 1)
        self.assertTrue(all(len(sql) <= 13000 for sql in statements))
        self.assertEqual(sum(sql.count("深圳投资项目公示") for sql in statements), 90)

    def test_tungee_assignment_query_uses_any_active_idle_seat(self):
        statements = []
        responses = iter(["1", "7", ""])

        def fake_mysql(sql, config=None):
            statements.append(sql)
            return next(responses)

        with (
            patch.object(server, "load_config", return_value={}),
            patch.object(server, "ensure_investigation_tables"),
            patch.object(server, "run_mysql", side_effect=fake_mysql),
            patch.object(server.threading, "Thread", DummyThread),
            patch.object(server, "fetch_investigation_center", return_value={"seats": [], "tasks": []}),
        ):
            server.assign_next_investigation_task()
        self.assertIn("`status`='active' AND `state`='idle'", statements[0])
        self.assertNotIn(f"`id`={server.PRIMARY_TUNGEE_SEAT_ID}", statements[0])

    def test_new_tungee_seat_stays_disabled_until_login_is_verified(self):
        statements = []
        with (
            patch.object(server, "load_config", return_value={}),
            patch.object(server, "ensure_investigation_tables"),
            patch.object(server, "run_mysql", side_effect=lambda sql, config=None: statements.append(sql) or ""),
            patch.object(server, "fetch_investigation_center", return_value={"seats": [], "tasks": []}),
        ):
            server.create_tungee_seat({"name": "探迹席位测试", "profileKey": "seat-test"})

        self.assertIn("'disabled', 'offline'", statements[0])

    def test_disabled_tungee_seat_cannot_be_enabled_without_login_verification(self):
        statements = []
        with (
            patch.object(server, "load_config", return_value={}),
            patch.object(server, "ensure_investigation_tables"),
            patch.object(server, "json_rows", return_value=[{"status": "disabled", "state": "offline"}]),
            patch.object(server, "run_mysql", side_effect=lambda sql, config=None: statements.append(sql) or ""),
        ):
            with self.assertRaisesRegex(server.ApiError, "登录并启用"):
                server.toggle_tungee_seat({"seatId": "2"})

        self.assertEqual(statements, [])

    def test_investigation_ui_scopes_queue_filters_from_project_categories(self):
        app_source = (ROOT / "app.js").read_text(encoding="utf-8")
        html_source = (ROOT / "index.html").read_text(encoding="utf-8")

        self.assertIn('document.querySelectorAll("[data-category]")', app_source)
        self.assertIn('data-queue-filter="active"', html_source)
        self.assertIn('data-queue-filter="failed"', html_source)
        self.assertIn('let queueFilter = "active"', app_source)
        self.assertIn("当前任务 #", app_source)

    def test_existing_lead_refresh_queries_confirmed_main_company(self):
        statements = []
        config = {
            "query": {
                "table": "sales_leads",
                "field_map": {
                    "projectName": "company_name",
                    "projectCompany": "project_unit",
                    "mainCompany": "main_company",
                },
            },
        }
        with (
            patch.object(server, "load_config", return_value=config),
            patch.object(server, "ensure_investigation_tables"),
            patch.object(server, "json_rows", return_value=[{
                "projectName": "测试光伏项目",
                "companyName": "备案壳公司有限公司",
                "mainCompany": "实际投资主体有限公司",
            }]),
            patch.object(server, "run_mysql", side_effect=lambda sql, config=None: statements.append(sql) or ""),
            patch.object(server, "fetch_investigation_center", return_value={"seats": [], "tasks": []}),
            patch.object(server, "start_auto_dispatch"),
        ):
            result = server.create_investigation_task({"leadId": "7", "platforms": "探迹"})

        self.assertFalse(result["duplicate"])
        insert_sql = next(sql for sql in statements if "INSERT INTO `investigation_tasks`" in sql)
        self.assertIn("实际投资主体有限公司", insert_sql)
        self.assertNotIn("备案壳公司有限公司", insert_sql)

    def test_refresh_can_prioritize_project_company_for_missing_project_contact(self):
        statements = []
        config = {
            "query": {
                "table": "sales_leads",
                "field_map": {
                    "projectName": "company_name",
                    "projectCompany": "project_unit",
                    "mainCompany": "main_company",
                },
            },
        }
        with (
            patch.object(server, "load_config", return_value=config),
            patch.object(server, "ensure_investigation_tables"),
            patch.object(server, "json_rows", return_value=[{
                "projectName": "测试光伏项目",
                "companyName": "备案项目公司有限公司",
                "mainCompany": "实际投资主体有限公司",
            }]),
            patch.object(server, "run_mysql", side_effect=lambda sql, config=None: statements.append(sql) or ""),
            patch.object(server, "fetch_investigation_center", return_value={"seats": [], "tasks": []}),
            patch.object(server, "start_auto_dispatch"),
        ):
            server.create_investigation_task({
                "leadId": "7", "platforms": "探迹", "targetCompany": "备案项目公司有限公司",
            })

        insert_sql = next(sql for sql in statements if "INSERT INTO `investigation_tasks`" in sql)
        self.assertIn("备案项目公司有限公司", insert_sql)

    def test_shared_company_backfill_only_fills_blank_fields(self):
        rows = [
            {"id": "1", "projectCompany": "项目公司", "projectContact": "", "projectPhone": "", "projectAddress": "", "mainCompany": "主体公司", "mainContact": "张三", "mainPhone": "13800138000", "mainAddress": "深圳地址"},
            {"id": "2", "projectCompany": "项目公司", "projectContact": "李四", "projectPhone": "13900139000", "projectAddress": "项目地址", "mainCompany": "主体公司", "mainContact": "", "mainPhone": "", "mainAddress": ""},
        ]
        statements = []
        with (
            patch.object(server, "load_config", return_value={}),
            patch.object(server, "json_rows", return_value=rows),
            patch.object(server, "run_mysql", side_effect=lambda sql, config=None: statements.append(sql) or ""),
        ):
            result = server.backfill_shared_company_data()

        self.assertEqual(result["updatedRows"], 2)
        self.assertGreaterEqual(result["filledFields"], 6)
        row_two_sql = next(sql for sql in statements if "WHERE `id`=2" in sql)
        self.assertIn("`main_phone`='13800138000'", row_two_sql)
        self.assertNotIn("`contact_person`", row_two_sql)

    def test_auto_dispatch_drains_available_capacity(self):
        calls = []

        def fake_assign():
            calls.append("assign")
            if len(calls) >= 3:
                raise server.ApiError("done")
            return {"ok": True}

        with (
            patch.object(server, "load_config", return_value={}),
            patch.object(server, "run_mysql", return_value="1"),
            patch.object(server, "assign_next_investigation_task", side_effect=fake_assign),
        ):
            server.auto_dispatch_available_tasks()
        self.assertEqual(len(calls), 3)

    def test_tungee_done_starts_research_worker(self):
        DummyThread.instances.clear()
        with (
            patch.object(server, "load_config", return_value={}),
            patch.object(server, "run_mysql", side_effect=["tungee_done", "0", ""]),
            patch.object(server.threading, "Thread", DummyThread),
        ):
            started = server.start_research_task("7")
        self.assertTrue(started)
        self.assertEqual(len(DummyThread.instances), 1)
        self.assertTrue(DummyThread.instances[0].started)
        self.assertEqual(DummyThread.instances[0].args, ("7",))

    def test_research_worker_does_not_start_when_another_is_active(self):
        DummyThread.instances.clear()
        with (
            patch.object(server, "load_config", return_value={}),
            patch.object(server, "run_mysql", side_effect=["tungee_done", "1"]),
            patch.object(server.threading, "Thread", DummyThread),
        ):
            started = server.start_research_task("8")
        self.assertFalse(started)
        self.assertEqual(DummyThread.instances, [])

    def test_auto_dispatch_resumes_research_states_after_restart(self):
        statements = []

        def fake_mysql(sql, config=None):
            statements.append(sql)
            if "COUNT(*)" in sql:
                return "0"
            if "status` IN ('doubao_queued','tungee_done','research_failed')" in sql:
                return "17"
            return ""

        with (
            patch.object(server, "load_config", return_value={}),
            patch.object(server, "run_mysql", side_effect=fake_mysql),
            patch.object(server, "start_research_task", return_value=True) as start_research,
            patch.object(server, "assign_next_investigation_task", side_effect=server.ApiError("done")),
        ):
            server.auto_dispatch_available_tasks()

        start_research.assert_called_once_with("17")

    def test_research_worker_dispatches_next_task_after_browser_closes(self):
        events = []

        class Driver:
            def quit(self):
                events.append("quit")

        with patch.object(server, "start_auto_dispatch", side_effect=lambda: events.append("dispatch")):
            server.release_research_driver_and_dispatch(Driver())

        self.assertEqual(events[-2:], ["quit", "dispatch"])

    def test_excel_export_uses_sales_result_layout(self):
        from openpyxl import load_workbook

        project = {
            "recordDate": "2026-07-13",
            "recordCode": "2607-test",
            "projectCompany": "测试项目单位",
            "projectName": "测试光伏项目",
            "mainCompany": "测试主体公司",
            "mainContact": "张三",
            "mainPhone": "13800138000",
            "projectAddress": "深圳市测试路1号",
            "projectSituation": "已备案",
            "category": "carport",
        }
        with (
            patch.object(server, "fetch_projects", return_value={"projects": [project]}),
            patch.object(server, "fetch_original_customers_for_export", return_value=[{
                "projectName": "原始备案项目", "projectCompany": "原始备案公司", "recordCode": "2607-raw",
            }]),
        ):
            workbook = load_workbook(BytesIO(server.build_export_workbook()))

        self.assertEqual(workbook.sheetnames, ["原始客户", "目标客户", "已见面客户", "业绩进度表"])
        self.assertEqual([cell.value for cell in workbook["原始客户"][1]], ["项目信息", "备案公司", "项目备案编号"])
        sheet = workbook["目标客户"]
        self.assertEqual([cell.value for cell in sheet[1]], server.SALES_RESULT_LABELS)
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual([cell.value for cell in sheet[2]], [
            "测试光伏项目", "测试项目单位", "光伏车棚", None, None, "测试主体公司",
            "张三", "13800138000", None, None, None, "已备案", None, None, "2607-test", None, None, None,
        ])

    def test_excel_export_filters_selected_date(self):
        from openpyxl import load_workbook

        projects = [
            {"recordDate": "2026-07-13", "recordCode": "new", "projectName": "当天充电站", "category": "carport"},
            {"recordDate": "2026-07-12", "recordCode": "old", "projectName": "历史充电站", "category": "carport"},
        ]
        with (
            patch.object(server, "fetch_projects", return_value={"projects": projects}),
            patch.object(server, "fetch_original_customers_for_export", return_value=[]),
        ):
            workbook = load_workbook(BytesIO(server.build_export_workbook("2026-07-13")))

        sheet = workbook["目标客户"]
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet["A2"].value, "当天充电站")
        self.assertEqual(sheet["O2"].value, "new")

    def test_excel_export_filters_undated_records(self):
        from openpyxl import load_workbook

        projects = [
            {"recordDate": "", "recordCode": "undated", "projectName": "未标日期项目", "category": "carport"},
            {"recordDate": "2026-07-08", "recordCode": "dated", "projectName": "已标日期项目", "category": "carport"},
        ]
        with (
            patch.object(server, "fetch_projects", return_value={"projects": projects}),
            patch.object(server, "fetch_original_customers_for_export", return_value=[]),
        ):
            workbook = load_workbook(BytesIO(server.build_export_workbook("__undated__")))

        sheet = workbook["目标客户"]
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet["A2"].value, "未标日期项目")
        self.assertEqual(sheet["O2"].value, "undated")


if __name__ == "__main__":
    unittest.main()
