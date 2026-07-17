from __future__ import annotations

import json
import os
import queue
import re
import subprocess
import sys
import threading
import time
import urllib.request
import uuid
import hmac
import secrets
import unicodedata
from datetime import date, timedelta
from io import BytesIO
from hashlib import sha1
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, urlencode, urlparse

BASE_DIR = Path(__file__).resolve().parent
PROJECT_CODE_DIR = BASE_DIR.parent
MODULE_CODE_DIR = BASE_DIR if (BASE_DIR / "enrich_tungee.py").exists() else PROJECT_CODE_DIR
CONFIG_PATH = BASE_DIR / "config.local.json"
FILING_SOURCE_NAME = "深圳投资项目公示"
FILING_SOURCE_URL = "https://wsbs.sz.gov.cn/investment/pubInformation/index"
PRIMARY_TUNGEE_SEAT_ID = 1
AUTO_FILING_KEYWORDS = "光伏 分布式光伏 屋顶 车棚 充电桩 充电 超充 光储充 储能 新能源 kW kWh"
PUBLIC_STATIC_PATHS = {"/", "/index.html", "/styles.css", "/app.js", "/caitoubiao_seed.js"}
CRAWLER_LOCK = threading.Lock()
JOB_LOCK = threading.Lock()
TASK_ASSIGN_LOCK = threading.Lock()
DOUBAO_LOCK = threading.Lock()
AUTO_DISPATCH_LOCK = threading.Lock()
AUTO_DISPATCH_INTERVAL_SECONDS = 15
FILING_JOBS: dict[str, dict] = {}
SESSION_LOCK = threading.Lock()
SESSIONS: dict[str, dict] = {}

EXPORT_FIELDS = [
    "projectName",
    "projectCompany",
    "projectContact",
    "projectPhone",
    "projectAddress",
    "mainCompany",
    "mainContact",
    "mainPhone",
    "mainAddress",
    "relationGraph",
    "investor",
    "projectSituation",
    "projectProgress",
    "projectLocation",
    "recordCode",
    "remark",
    "projectSummary",
    "phoneFeedback",
]

META_FIELDS = ["source", "sourceId", "recordDate"]
EXPORT_LABELS = [
    "项目信息", "项目单位", "项目公司联系人", "手机号", "地址",
    "控股公司", "控股公司联系人", "手机号", "控股公司地址", "关系图谱",
    "投资人", "项目情况", "项目进展", "项目所在地", "项目备案编号",
    "备注", "项目总结", "电话反馈",
]
SALES_RESULT_FIELDS = [
    "projectName", "projectCompany", "projectNature", "projectContact", "projectPhone",
    "mainCompany", "mainContact", "mainPhone", "mainAddress", "relationGraph", "investor",
    "projectSituation", "projectProgress", "projectLocation", "recordCode", "remark", "projectSummary", "phoneFeedback",
]
SALES_RESULT_LABELS = [
    "项目信息", "项目单位", "项目性质", "项目公司联系人", "手机号", "控股公司",
    "控股公司联系人", "手机号", "控股公司地址", "关系图谱", "投资人", "项目情况",
    "项目进展", "项目所在地", "项目备案编号", "备注", "项目总结", "电话反馈",
]
EXPORT_SHEETS = {
    "roof_state": "屋顶光伏（国企）",
    "storage_charge": "光储充",
    "carport": "光伏车棚",
    "roof_private": "屋顶光伏（民企）",
}

FILING_FIELDS = [
    ("projectCode", "recordCode"),
    ("projectName", "projectName"),
    ("projectUnit", "projectCompany"),
    ("projectType", "projectProgress"),
    ("recordDate", "recordDate"),
    ("source", "source"),
    ("sourceId", "sourceId"),
]

LEAD_INSERT_KEYS = EXPORT_FIELDS + META_FIELDS + ["category"]
WORKFLOW_COLUMNS = {
    "category": "category",
    "tungeeStatus": "tungee_status",
    "reviewStatus": "review_status",
    "queued": "queued",
}


class ApiError(Exception):
    pass


def now_text() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")


def configured_access_code() -> str:
    return os.environ.get("XINZHONG_ERP_ACCESS_CODE", "").strip()


def configured_users() -> list[dict]:
    """Load internal accounts without exposing public self-registration."""
    try:
        users = load_config().get("auth", {}).get("users", [])
    except Exception:
        users = []
    return [item for item in users if isinstance(item, dict) and item.get("username") and item.get("password")]


def session_token_from_headers(headers) -> str:
    cookie = headers.get("Cookie") or ""
    for part in cookie.split(";"):
        name, separator, value = part.strip().partition("=")
        if separator and name == "xinzhong_session":
            return value
    return ""


def session_user(headers) -> dict | None:
    if not configured_access_code() and not configured_users():
        return {"username": "local-admin", "name": "本机管理员", "role": "admin", "department": "all"}
    token = session_token_from_headers(headers)
    if not token:
        return None
    now = time.time()
    with SESSION_LOCK:
        session = SESSIONS.get(token) or {}
        expires_at = session.get("expiresAt", 0)
        if expires_at <= now:
            SESSIONS.pop(token, None)
            return None
        session["expiresAt"] = now + 12 * 60 * 60
        SESSIONS[token] = session
        return dict(session.get("user") or {})


def session_is_valid(headers) -> bool:
    return session_user(headers) is not None


def user_can_access(user: dict, area: str) -> bool:
    role = str((user or {}).get("role") or "")
    department = str((user or {}).get("department") or "")
    if role == "admin":
        return True
    if area == "sales":
        return department == "sales"
    if area == "business":
        return department == "business"
    if area == "hr":
        return department == "hr"
    return False


def load_config() -> dict:
    if not CONFIG_PATH.exists():
        raise ApiError("缺少 config.local.json。请复制 config.example.json 后填入 Navicat 同一套 MySQL 连接信息。")
    with CONFIG_PATH.open("r", encoding="utf-8") as f:
        return json.load(f)


def quote_identifier(identifier: str) -> str:
    text = str(identifier).strip()
    if not text:
        raise ApiError("MySQL 配置里有空字段名。")
    if "`" in text or ";" in text or "\x00" in text:
        raise ApiError(f"字段名不安全：{text}")
    return ".".join(f"`{part}`" for part in text.split("."))


def value_expr(field_map: dict, key: str) -> str:
    column = field_map.get(key)
    if not column:
        return "''"
    if isinstance(column, list):
        parts = [f"NULLIF(CAST({quote_identifier(item)} AS CHAR), '')" for item in column if item]
        if not parts:
            return "''"
        return f"IFNULL(COALESCE({', '.join(parts)}), '')"
    return f"IFNULL(CAST({quote_identifier(column)} AS CHAR), '')"


def sql_literal(value: str) -> str:
    return "'" + str(value).replace("\\", "\\\\").replace("'", "''") + "'"


def split_keywords(text: str) -> list[str]:
    return [normalize_search_text(item) for item in re.split(r"[\s,，、]+", text.strip()) if item.strip()]


def normalize_search_text(value: object) -> str:
    """Normalize keyword and filing text so kW/kWh/千瓦 variants match."""
    text = str(value or "").lower().strip()
    text = text.replace("ｋ", "k").replace("ｗ", "w").replace("ｈ", "h")
    text = text.replace("千瓦时", "kwh").replace("千瓦時", "kwh").replace("千瓦", "kw")
    return re.sub(r"[\s_\-_/·•]+", "", text)


def filing_search_text(row: dict) -> str:
    # Business relevance is determined by the project itself. A company whose
    # name contains "新能源" may also file unrelated construction projects.
    return normalize_search_text(row.get("projectName", ""))


def is_relevant_filing_project(project: dict, keywords: list[str] | None = None) -> bool:
    project_name = project.get("projectName") or project.get("company_name") or ""
    normalized_name = normalize_search_text(project_name)
    active_keywords = keywords or split_keywords(AUTO_FILING_KEYWORDS)
    return bool(normalized_name) and any(keyword in normalized_name for keyword in active_keywords)


def phone_field_quality(value: object) -> dict:
    text = str(value or "").strip()
    candidates = list(dict.fromkeys(re.findall(r"(?<!\d)1[3-9]\d{9}(?!\d)", text)))
    if not text:
        return {"status": "missing", "label": "缺失", "phones": []}
    if len(candidates) > 1:
        return {"status": "multiple", "label": "多个号码，需确认联系人", "phones": candidates}
    if len(candidates) == 1 and re.fullmatch(r"1[3-9]\d{9}", text):
        return {"status": "formatted", "label": "单号码格式有效，仍需通话核验", "phones": candidates}
    if len(candidates) == 1:
        return {"status": "noted", "label": "号码夹带备注，需人工整理", "phones": candidates}
    return {"status": "invalid", "label": "未识别到有效手机号", "phones": []}


def assess_project_accuracy(project: dict) -> dict:
    issues = []
    project_phone = phone_field_quality(project.get("projectPhone"))
    main_phone = phone_field_quality(project.get("mainPhone"))
    for label, quality in (("项目公司手机号", project_phone), ("控股公司手机号", main_phone)):
        if quality["status"] != "formatted":
            issues.append(f"{label}：{quality['label']}")

    main_company = str(project.get("mainCompany") or "").strip()
    project_company = str(project.get("projectCompany") or "").strip()
    relation_graph = str(project.get("relationGraph") or "").strip()
    combined_evidence = f"{relation_graph} {project.get('remark') or ''}"
    if not main_company:
        issues.append("控股公司尚未确认")
    elif enterprise_name_key(main_company) == enterprise_name_key(project_company):
        issues.append("控股公司与备案公司相同，需确认是否为企业自投或无上级控股企业")
    if not relation_graph:
        issues.append("缺少股权穿透依据")
    elif any(marker in combined_evidence for marker in ("备案公司作为探迹股权穿透起点", "模糊命中", "不是已确认的最终主体")):
        issues.append("控股公司使用兜底或模糊匹配，必须人工复核")
    if not str(project.get("projectLocation") or "").strip():
        issues.append("项目所在地缺失")
    return {
        "status": "需人工核验" if issues else "资料格式完整",
        "issues": issues,
        "projectPhone": project_phone,
        "mainPhone": main_phone,
    }


def record_to_filing_row(record: dict) -> dict:
    project_name = record.get("project_desc") or record.get("company_name", "")
    project_unit = record.get("project_unit") or record.get("company_name", "")
    return {
        "projectCode": record.get("credit_code", ""),
        "projectName": project_name,
        "projectUnit": project_unit,
        "projectType": record.get("project_status", ""),
        "recordDate": record.get("record_date", ""),
        "source": FILING_SOURCE_NAME,
        "sourceId": record.get("credit_code") or project_name,
        "sourceUrl": FILING_SOURCE_URL,
    }


def crawl_filings_live(date_text: str, keywords: list[str]) -> tuple[list[dict], list[str]]:
    if not CRAWLER_LOCK.acquire(blocking=False):
        raise ApiError("已有备案采集任务正在运行，请稍后再查。")

    try:
        if str(MODULE_CODE_DIR) not in sys.path:
            sys.path.insert(0, str(MODULE_CODE_DIR))
        from main import SZInvestCrawler

        class ServerFilingCrawler(SZInvestCrawler):
            def _build_driver(self):
                from selenium import webdriver as selenium_webdriver
                from selenium.webdriver import ChromeOptions as SeleniumChromeOptions

                options = SeleniumChromeOptions()
                if self.chrome_path:
                    options.binary_location = self.chrome_path
                options.add_argument("--headless=new")
                options.add_argument("--disable-gpu")
                options.add_argument("--no-sandbox")
                options.add_argument("--disable-dev-shm-usage")
                options.add_argument("--disable-extensions")
                options.add_argument("--no-first-run")
                options.add_argument("--window-size=1440,1000")
                last_error = None
                for attempt in range(2):
                    try:
                        # Selenium creates an isolated temporary profile. The filing site needs no login state.
                        self.driver = selenium_webdriver.Chrome(options=options)
                        break
                    except Exception as exc:
                        last_error = exc
                        self.log(f"备案 Chrome 启动失败，正在重试 {attempt + 1}/2：{str(exc)[:200]}")
                        time.sleep(2)
                if self.driver is None:
                    raise last_error or ApiError("备案 Chrome 启动失败。")
                self.driver.set_page_load_timeout(30)
                self.driver.set_script_timeout(20)
                self.log("服务器备案专用无头 Chrome 已启动")
                return self.driver

        logs: queue.Queue[str] = queue.Queue()
        crawl_stop = threading.Event()
        crawler = ServerFilingCrawler(
            log_queue=logs,
            interval=1,
            debug_port=9218,
            headless=True,
            filter_keywords=None,
            stop_event=crawl_stop,
        )
        result_box = {}

        def crawl_target():
            try:
                result_box["records"] = crawler.crawl_by_date(date_text, date_text)
            except Exception as exc:
                result_box["error"] = exc

        crawl_thread = threading.Thread(target=crawl_target, daemon=True)
        crawl_thread.start()
        timeout_seconds = max(90, int(os.environ.get("FILING_CRAWLER_TIMEOUT_SECONDS", "240")))
        crawl_thread.join(timeout_seconds)
        if crawl_thread.is_alive():
            crawl_stop.set()
            if crawler.driver:
                try:
                    crawler.driver.quit()
                except Exception:
                    pass
            crawl_thread.join(10)
            raise ApiError(f"采投标网站响应超过 {timeout_seconds} 秒，已保留已有数据库数据，请稍后重试。")
        if result_box.get("error"):
            raise result_box["error"]
        records = result_box.get("records", [])
        crawler_logs = []
        while not logs.empty():
            crawler_logs.append(logs.get_nowait())
        fatal_logs = [line for line in crawler_logs if "爬取出错" in line or "no such window" in line.lower()]
        if fatal_logs:
            raise ApiError(fatal_logs[-1][:1000])
        rows = [record_to_filing_row(record) for record in records]
        if keywords:
            rows = [
                row
                for row in rows
                if any(
                    keyword in filing_search_text(row)
                    for keyword in keywords
                )
            ]
        return rows, crawler_logs[-20:]
    finally:
        CRAWLER_LOCK.release()


def run_mysql(sql: str, config: dict) -> str:
    mysql = config.get("mysql", {})
    mysql_exe = mysql.get("mysql_exe") or "mysql"
    args = [
        mysql_exe,
        "--default-character-set=utf8mb4",
        "--batch",
        "--raw",
        "--skip-column-names",
        "-h",
        str(mysql.get("host", "127.0.0.1")),
        "-P",
        str(mysql.get("port", 3306)),
        "-u",
        str(mysql.get("user", "")),
    ]
    database = mysql.get("database")
    if database:
        args.extend(["-D", str(database)])
    args.extend(["-e", sql])

    env = os.environ.copy()
    password = mysql.get("password")
    if password:
        env["MYSQL_PWD"] = str(password)

    try:
        result = subprocess.run(
            args,
            cwd=BASE_DIR,
            env=env,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=30,
        )
    except FileNotFoundError as exc:
        if getattr(exc, "winerror", None) == 206:
            raise ApiError("MySQL 写入命令过长，请分批提交数据。") from exc
        raise ApiError(f"找不到 mysql.exe：{mysql_exe}") from exc
    except subprocess.TimeoutExpired as exc:
        raise ApiError("MySQL 查询超时。") from exc

    if result.returncode != 0:
        message = result.stderr.strip() or result.stdout.strip() or "MySQL 查询失败"
        raise ApiError(message.replace(str(password or ""), "***"))
    return result.stdout


def ensure_filing_cache_table(config: dict) -> None:
    sql = """
CREATE TABLE IF NOT EXISTS `filing_projects` (
  `id` bigint unsigned NOT NULL AUTO_INCREMENT,
  `source` varchar(64) NOT NULL DEFAULT '深圳投资项目公示',
  `source_project_id` varchar(128) NOT NULL DEFAULT '',
  `project_code` varchar(128) NOT NULL DEFAULT '',
  `project_name` varchar(512) NOT NULL DEFAULT '',
  `project_unit` varchar(512) NOT NULL DEFAULT '',
  `project_type` varchar(255) NOT NULL DEFAULT '',
  `record_date` varchar(64) NOT NULL DEFAULT '',
  `source_url` varchar(512) NOT NULL DEFAULT '',
  `keyword_text` varchar(512) NOT NULL DEFAULT '',
  `raw_json` json DEFAULT NULL,
  `created_at` timestamp NOT NULL DEFAULT CURRENT_TIMESTAMP,
  `updated_at` timestamp NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
  PRIMARY KEY (`id`),
  UNIQUE KEY `uniq_source_project` (`source`,`source_project_id`),
  KEY `idx_record_date` (`record_date`),
  KEY `idx_project_unit` (`project_unit`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci
"""
    run_mysql(sql, config)


def ensure_filing_jobs_table(config: dict) -> None:
    run_mysql(
        """
CREATE TABLE IF NOT EXISTS `filing_jobs` (
  `id` varchar(32) NOT NULL,
  `record_date` varchar(10) NOT NULL DEFAULT '',
  `keyword_text` varchar(1000) NOT NULL DEFAULT '',
  `status` varchar(32) NOT NULL DEFAULT 'queued',
  `requested_by` varchar(128) NOT NULL DEFAULT '',
  `message` varchar(1000) NOT NULL DEFAULT '',
  `read_mode` varchar(64) NOT NULL DEFAULT '',
  `result_count` int unsigned NOT NULL DEFAULT 0,
  `result_json` longtext,
  `error_text` text,
  `started_at` datetime DEFAULT NULL,
  `finished_at` datetime DEFAULT NULL,
  `created_at` datetime NOT NULL,
  `updated_at` datetime NOT NULL,
  PRIMARY KEY (`id`),
  KEY `idx_filing_jobs_date` (`record_date`,`created_at`),
  KEY `idx_filing_jobs_status` (`status`,`created_at`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci
""",
        config,
    )


def recover_filing_jobs() -> None:
    """A queued/running filing job cannot survive an ERP process restart."""
    config = load_config()
    ensure_filing_jobs_table(config)
    run_mysql(
        "UPDATE `filing_jobs` SET `status`='failed',`message`='服务重启，原采集任务已中断，请重新查询',"
        "`error_text`='ERP服务重启导致备案采集任务中断',`finished_at`=NOW(),`updated_at`=NOW() "
        "WHERE `status` IN ('queued','running')",
        config,
    )


def persist_filing_job(job: dict) -> None:
    """Best-effort persistence; a crawler job must still run if history storage fails."""
    try:
        config = load_config()
        ensure_filing_jobs_table(config)
        rows_json = json.dumps(job.get("rows", []), ensure_ascii=False)
        started = sql_literal(job["startedAt"]) if job.get("startedAt") else "NULL"
        finished = sql_literal(job["finishedAt"]) if job.get("finishedAt") else "NULL"
        sql = (
            "INSERT INTO `filing_jobs` (`id`,`record_date`,`keyword_text`,`status`,`requested_by`,`message`,`read_mode`,"
            "`result_count`,`result_json`,`error_text`,`started_at`,`finished_at`,`created_at`,`updated_at`) VALUES ("
            f"{sql_literal(job.get('id',''))},{sql_literal(job.get('date',''))},{sql_literal(job.get('keywordText',''))},"
            f"{sql_literal(job.get('status',''))},{sql_literal(job.get('requestedBy',''))},{sql_literal(job.get('message',''))},"
            f"{sql_literal(job.get('readMode',''))},{int(job.get('count') or 0)},{sql_literal(rows_json)},"
            f"{sql_literal(job.get('error',''))},{started},{finished},{sql_literal(job.get('createdAt') or now_text())},"
            f"{sql_literal(job.get('updatedAt') or now_text())}) ON DUPLICATE KEY UPDATE "
            "`status`=VALUES(`status`),`message`=VALUES(`message`),`read_mode`=VALUES(`read_mode`),"
            "`result_count`=VALUES(`result_count`),`result_json`=VALUES(`result_json`),`error_text`=VALUES(`error_text`),"
            "`started_at`=VALUES(`started_at`),`finished_at`=VALUES(`finished_at`),`updated_at`=VALUES(`updated_at`)"
        )
        run_mysql(sql, config)
    except Exception:
        pass


def filing_job_from_db(job_id: str) -> dict | None:
    try:
        config = load_config()
        ensure_filing_jobs_table(config)
        rows = json_rows(
            "SELECT JSON_OBJECT('id',`id`,'date',`record_date`,'keywordText',`keyword_text`,'status',`status`,"
            "'requestedBy',`requested_by`,'message',`message`,'readMode',`read_mode`,'count',`result_count`,"
            "'rows',IFNULL(`result_json`,'[]'),'error',IFNULL(`error_text`,''),'startedAt',IFNULL(CAST(`started_at` AS CHAR),''),"
            "'finishedAt',IFNULL(CAST(`finished_at` AS CHAR),''),'createdAt',CAST(`created_at` AS CHAR),"
            "'updatedAt',CAST(`updated_at` AS CHAR)) FROM `filing_jobs` "
            f"WHERE `id`={sql_literal(job_id)} LIMIT 1",
            config,
        )
        if not rows:
            return None
        job = rows[0]
        if isinstance(job.get("rows"), str):
            job["rows"] = json.loads(job["rows"] or "[]")
        job["sourceLabel"] = FILING_SOURCE_NAME
        job["sourceUrl"] = FILING_SOURCE_URL
        return job
    except Exception:
        return None


def recent_filing_jobs(limit: int = 20) -> dict:
    try:
        config = load_config()
        ensure_filing_jobs_table(config)
        rows = json_rows(
            "SELECT JSON_OBJECT('id',`id`,'date',`record_date`,'keywordText',`keyword_text`,'status',`status`,"
            "'requestedBy',`requested_by`,'message',`message`,'readMode',`read_mode`,'count',`result_count`,"
            "'error',IFNULL(`error_text`,''),'startedAt',IFNULL(CAST(`started_at` AS CHAR),''),"
            "'finishedAt',IFNULL(CAST(`finished_at` AS CHAR),''),'createdAt',CAST(`created_at` AS CHAR),"
            "'updatedAt',CAST(`updated_at` AS CHAR)) FROM `filing_jobs` ORDER BY `created_at` DESC "
            f"LIMIT {max(1, min(int(limit), 100))}",
            config,
        )
    except Exception:
        with JOB_LOCK:
            rows = sorted((public_job(item) for item in FILING_JOBS.values()), key=lambda item: item.get("createdAt", ""), reverse=True)[:limit]
    return {"ok": True, "jobs": rows}


def stable_source_id(row: dict) -> str:
    explicit = str(row.get("sourceId") or row.get("projectCode") or "").strip()
    if explicit:
        return explicit[:128]
    basis = "|".join(
        str(row.get(key, "")).strip()
        for key in ("projectName", "projectUnit", "recordDate")
    )
    return sha1(basis.encode("utf-8")).hexdigest()[:24]


def store_filing_rows(rows: list[dict], keyword_text: str) -> None:
    if not rows:
        return
    config = load_config()
    ensure_filing_cache_table(config)
    values = []
    for row in rows:
        source_id = stable_source_id(row)
        raw_json = json.dumps(row, ensure_ascii=False)
        values.append(
            "("
            + ", ".join(
                [
                    sql_literal(FILING_SOURCE_NAME),
                    sql_literal(source_id),
                    sql_literal(row.get("projectCode", "")),
                    sql_literal(row.get("projectName", "")),
                    sql_literal(row.get("projectUnit", "")),
                    sql_literal(row.get("projectType", "")),
                    sql_literal(row.get("recordDate", "")),
                    sql_literal(row.get("sourceUrl", FILING_SOURCE_URL)),
                    sql_literal(keyword_text),
                    sql_literal(raw_json),
                ]
            )
            + ")"
        )
    insert_prefix = (
        "INSERT INTO `filing_projects` "
        "(`source`,`source_project_id`,`project_code`,`project_name`,`project_unit`,`project_type`,`record_date`,`source_url`,`keyword_text`,`raw_json`) VALUES "
    )
    update_suffix = (
        " ON DUPLICATE KEY UPDATE "
        "`project_code`=VALUES(`project_code`), "
        "`project_name`=VALUES(`project_name`), "
        "`project_unit`=VALUES(`project_unit`), "
        "`project_type`=VALUES(`project_type`), "
        "`record_date`=VALUES(`record_date`), "
        "`source_url`=VALUES(`source_url`), "
        "`keyword_text`=VALUES(`keyword_text`), "
        "`raw_json`=VALUES(`raw_json`)"
    )
    batch = []
    batch_length = len(insert_prefix) + len(update_suffix)
    max_command_length = 12000
    for value in values:
        added_length = len(value) + (2 if batch else 0)
        if batch and batch_length + added_length > max_command_length:
            run_mysql(insert_prefix + ", ".join(batch) + update_suffix, config)
            batch = []
            batch_length = len(insert_prefix) + len(update_suffix)
        batch.append(value)
        batch_length += len(value) + (2 if len(batch) > 1 else 0)
    if batch:
        run_mysql(insert_prefix + ", ".join(batch) + update_suffix, config)


def fetch_filing_cache(date_text: str, keywords: list[str]) -> dict:
    config = load_config()
    ensure_filing_cache_table(config)
    sql = (
        "SELECT JSON_OBJECT("
        "'projectCode', IFNULL(CAST(`project_code` AS CHAR), ''), "
        "'projectName', IFNULL(CAST(`project_name` AS CHAR), ''), "
        "'projectUnit', IFNULL(CAST(`project_unit` AS CHAR), ''), "
        "'projectType', IFNULL(CAST(`project_type` AS CHAR), ''), "
        "'recordDate', IFNULL(CAST(`record_date` AS CHAR), ''), "
        "'source', IFNULL(CAST(`source` AS CHAR), ''), "
        "'sourceId', IFNULL(CAST(`source_project_id` AS CHAR), ''), "
        "'sourceUrl', IFNULL(CAST(`source_url` AS CHAR), '')"
        ") FROM `filing_projects` "
        f"WHERE LEFT(CAST(`record_date` AS CHAR), 10) = {sql_literal(date_text)} "
        "ORDER BY `id` DESC LIMIT 5000"
    )
    all_rows = []
    for line in run_mysql(sql, config).splitlines():
        line = line.strip()
        if not line:
            continue
        item = json.loads(line)
        all_rows.append(item)
    all_rows = deduplicate_filing_rows(all_rows)
    rows = filter_filing_rows(all_rows, keywords)
    dates = [
        line.strip()
        for line in run_mysql(
            "SELECT DISTINCT LEFT(CAST(`record_date` AS CHAR), 10) FROM `filing_projects` "
            "WHERE `record_date` <> '' ORDER BY 1 DESC LIMIT 5",
            config,
        ).splitlines()
        if line.strip()
    ]
    count_rows = run_mysql(
        "SELECT COUNT(*) FROM `filing_projects` "
        f"WHERE LEFT(CAST(`record_date` AS CHAR), 10) = {sql_literal(date_text)}",
        config,
    ).strip()
    return {
        "rows": rows,
        "allRows": all_rows,
        "availableDates": dates,
        "hasDate": int(count_rows or 0) > 0,
        "totalCount": len(all_rows),
    }


def lead_value(project: dict, key: str) -> str:
    aliases = {
        "projectCompany": "projectUnit",
        "recordCode": "projectCode",
        "projectProgress": "projectType",
    }
    value = project.get(key, "")
    if (value is None or value == "") and aliases.get(key):
        value = project.get(aliases[key], "")
    if value is None:
        return ""
    return str(value)


def ensure_sales_workflow_columns(config: dict, table: str) -> None:
    existing = {
        line.split("\t", 1)[0]
        for line in run_mysql(f"SHOW COLUMNS FROM {quote_identifier(table)}", config).splitlines()
        if line.strip()
    }
    definitions = {
        "tungee_status": "varchar(32) NOT NULL DEFAULT ''",
        "review_status": "varchar(32) NOT NULL DEFAULT ''",
        "queued": "tinyint(1) NOT NULL DEFAULT 0",
    }
    missing = [name for name in definitions if name not in existing]
    if not missing:
        return
    clauses = [f"ADD COLUMN {quote_identifier(name)} {definitions[name]}" for name in missing]
    run_mysql(f"ALTER TABLE {quote_identifier(table)} {', '.join(clauses)}", config)


def project_to_lead_row(project: dict) -> dict:
    project_name = lead_value(project, "projectName")
    project_company = lead_value(project, "projectCompany")
    record_date = lead_value(project, "recordDate")
    source = lead_value(project, "source") or FILING_SOURCE_NAME
    source_id = lead_value(project, "sourceId") or lead_value(project, "recordCode")
    if not source_id:
        source_id = sha1(f"{project_name}|{project_company}|{record_date}".encode("utf-8")).hexdigest()[:24]
    source_id = source_id[:64]
    row = {key: lead_value(project, key) for key in LEAD_INSERT_KEYS}
    row["source"] = source
    row["sourceId"] = source_id
    row["projectName"] = project_name
    row["projectCompany"] = project_company
    row["recordDate"] = record_date
    row["category"] = lead_value(project, "category")
    if not row.get("projectLocation"):
        row["projectLocation"] = infer_project_location(project_name)
    return row


def infer_project_location(project_name: str) -> str:
    """Extract an obvious project site from a filing title."""
    text = re.sub(r"\s+", "", str(project_name or "")).strip()
    match = re.match(
        r"((?:广东省)?深圳市[^，,；;]{4,100}?)(?=\d+(?:\.\d+)?(?:kWp|kW|MW|千瓦|兆瓦))",
        text,
        re.IGNORECASE,
    )
    return match.group(1) if match else ""


def save_leads(projects: list[dict], auto_investigate: bool = True) -> dict:
    if not projects:
        return {"ok": True, "count": 0, "saved": []}

    config = load_config()
    query = config.get("query", {})
    table = query.get("table")
    field_map = query.get("field_map", {})
    if not table:
        raise ApiError("config.local.json 缺少 query.table。")
    ensure_sales_workflow_columns(config, table)

    key_to_column = {key: field_map.get(key) for key in EXPORT_FIELDS + META_FIELDS}
    key_to_column.update(WORKFLOW_COLUMNS)
    insert_keys = [key for key in LEAD_INSERT_KEYS if key_to_column.get(key)]
    columns = [quote_identifier(key_to_column[key]) for key in insert_keys]
    values_sql = []
    saved = []
    linked = 0
    source_column = field_map.get("sourceId")
    for project in projects:
        row = project_to_lead_row(project)
        record_column = field_map.get("recordCode")
        name_column = field_map.get("projectName")
        company_column = field_map.get("projectCompany")
        match_candidates = []
        if row.get("recordCode") and record_column:
            match_candidates.append(f"{quote_identifier(record_column)}={sql_literal(row['recordCode'])}")
        if row.get("projectName") and row.get("projectCompany") and name_column and company_column:
            match_candidates.append(
                f"{quote_identifier(name_column)}={sql_literal(row['projectName'])} AND "
                f"{quote_identifier(company_column)}={sql_literal(row['projectCompany'])}"
            )
        existing = []
        for match_sql in match_candidates:
            existing = json_rows(
                "SELECT JSON_OBJECT('dbId',CAST(`id` AS CHAR),'source',`source`,"
                "'sourceId',`source_project_id`,'projectName',`company_name`) "
                f"FROM {quote_identifier(table)} WHERE {match_sql} ORDER BY `id` LIMIT 1",
                config,
            )
            if existing:
                break
        if not existing and row.get("sourceId") and source_column:
            existing = json_rows(
                "SELECT JSON_OBJECT('dbId',CAST(`id` AS CHAR),'source',`source`,"
                "'sourceId',`source_project_id`,'projectName',`company_name`) "
                f"FROM {quote_identifier(table)} WHERE {quote_identifier(source_column)}={sql_literal(row['sourceId'])} "
                "ORDER BY `id` LIMIT 1",
                config,
            )
        if existing:
            existing_row = existing[0]
            supplemental = []
            for key in ("recordDate", "projectProgress", "projectLocation", "recordCode", "category"):
                column = key_to_column.get(key)
                value = row.get(key, "")
                if column and value:
                    quoted = quote_identifier(column)
                    supplemental.append(f"{quoted}=COALESCE(NULLIF({quoted},''),{sql_literal(value)})")
            if supplemental:
                run_mysql(
                    f"UPDATE {quote_identifier(table)} SET {', '.join(supplemental)} "
                    f"WHERE `id`={int(existing_row['dbId'])} LIMIT 1",
                    config,
                )
            existing_row["duplicate"] = True
            saved.append(existing_row)
            linked += 1
            continue
        saved.append(
            {
                "source": row["source"],
                "sourceId": row["sourceId"],
                "projectName": row["projectName"],
                "duplicate": False,
            }
        )
        values_sql.append("(" + ", ".join(sql_literal(row.get(key, "")) for key in insert_keys) + ")")

    update_columns = [
        "projectName",
        "projectCompany",
        "projectProgress",
        "projectLocation",
        "recordCode",
        "remark",
        "recordDate",
        "category",
    ]
    updates = []
    for key in update_columns:
        column = key_to_column.get(key)
        if column:
            quoted = quote_identifier(column)
            updates.append(f"{quoted}=VALUES({quoted})")
    if values_sql:
        sql = (
            f"INSERT INTO {quote_identifier(table)} ({', '.join(columns)}) "
            f"VALUES {', '.join(values_sql)} "
            f"ON DUPLICATE KEY UPDATE {', '.join(updates)}"
        )
        run_mysql(sql, config)
    source_column = field_map.get("sourceId")
    if source_column:
        for item in saved:
            if item.get("dbId") or not item.get("sourceId"):
                continue
            matched = json_rows(
                "SELECT JSON_OBJECT('dbId',CAST(`id` AS CHAR)) "
                f"FROM {quote_identifier(table)} WHERE {quote_identifier(source_column)}="
                f"{sql_literal(item['sourceId'])} ORDER BY `id` DESC LIMIT 1",
                config,
            )
            if matched:
                item["dbId"] = matched[0]["dbId"]

    investigation_queued = 0
    investigation_duplicates = 0
    if auto_investigate:
        for item in saved:
            if not item.get("dbId"):
                continue
            task_result = create_investigation_task(
                {
                    "leadId": item["dbId"],
                    "platforms": "探迹,豆包,百度",
                    "requestedBy": "正式线索入库",
                }
            )
            if task_result.get("duplicate"):
                investigation_duplicates += 1
            else:
                investigation_queued += 1
    return {
        "ok": True,
        "count": len(projects),
        "inserted": len(values_sql),
        "linked": linked,
        "saved": saved,
        "investigationQueued": investigation_queued,
        "investigationDuplicates": investigation_duplicates,
    }


def submit_intake_projects(projects: list[dict]) -> dict:
    """Store raw filing projects in intake and start research before formal lead creation."""
    if not projects:
        return {"ok": True, "count": 0, "queued": 0, "duplicates": 0, "saved": []}
    config = load_config()
    ensure_investigation_tables(config)
    queued = 0
    duplicates = 0
    saved = []
    skipped = []
    for project in projects:
        if str(project.get("source") or "").strip() == FILING_SOURCE_NAME and not is_relevant_filing_project(project):
            skipped.append({
                "sourceId": project.get("sourceId") or project.get("recordCode") or "",
                "projectName": project.get("projectName") or "",
                "reason": "项目名称未命中新能源业务关键词",
            })
            continue
        row = project_to_lead_row(project)
        source_id = row["sourceId"]
        existing = json_rows(
            "SELECT JSON_OBJECT('intakeId',CAST(`id` AS CHAR),'taskId',IFNULL(CAST(`task_id` AS CHAR),''),'status', `status`) "
            f"FROM `investigation_intake` WHERE `source_id`={sql_literal(source_id)} LIMIT 1",
            config,
        )
        if existing:
            existing_task_id = existing[0].get("taskId", "")
            if existing[0].get("status") in {"failed", "research_failed"} and existing_task_id:
                run_mysql(
                    "UPDATE `investigation_tasks` SET `status`='doubao_queued', `assigned_seat_id`=NULL, "
                    "`started_at`=NULL, `finished_at`=NULL, `error_text`='' "
                    f"WHERE `id`={int(existing_task_id)} LIMIT 1; "
                    f"UPDATE `investigation_intake` SET `status`='doubao_queued', `error_text`='' WHERE `id`={int(existing[0]['intakeId'])} LIMIT 1",
                    config,
                )
                saved.append({"intakeId": existing[0]["intakeId"], "taskId": existing_task_id, "sourceId": source_id, "duplicate": False})
                queued += 1
            else:
                saved.append({"intakeId": existing[0]["intakeId"], "taskId": existing_task_id, "sourceId": source_id, "duplicate": True})
                duplicates += 1
            continue
        raw = json.dumps(row, ensure_ascii=False)
        run_mysql(
            "INSERT INTO `investigation_intake` (`source_id`,`company_name`,`project_name`,`project_json`) VALUES ("
            f"{sql_literal(source_id)},{sql_literal(row['projectCompany'])},{sql_literal(row['projectName'])},{sql_literal(raw)})",
            config,
        )
        intake_id = run_mysql(
            f"SELECT `id` FROM `investigation_intake` WHERE `source_id`={sql_literal(source_id)} ORDER BY `id` DESC LIMIT 1",
            config,
        ).strip()
        task_id = run_mysql(
            "INSERT INTO `investigation_tasks` (`lead_id`,`company_name`,`project_name`,`platforms`,`requested_by`) VALUES ("
            f"NULL,{sql_literal(row['projectCompany'])},{sql_literal(row['projectName'])},'探迹,豆包,百度','备案项目入库'); "
            "SELECT LAST_INSERT_ID()",
            config,
        ).splitlines()[-1].strip()
        run_mysql(
            f"UPDATE `investigation_tasks` SET `status`='doubao_queued' WHERE `id`={int(task_id)} LIMIT 1; "
            f"UPDATE `investigation_intake` SET `status`='doubao_queued' WHERE `id`={int(intake_id)} LIMIT 1",
            config,
        )
        run_mysql(
            f"UPDATE `investigation_intake` SET `task_id`={int(task_id)} WHERE `id`={int(intake_id)} LIMIT 1",
            config,
        )
        saved.append({"intakeId": intake_id, "taskId": task_id, "sourceId": source_id, "duplicate": False})
        queued += 1
    start_auto_dispatch()
    return {
        "ok": True,
        "count": len(projects),
        "accepted": len(projects) - len(skipped),
        "queued": queued,
        "duplicates": duplicates,
        "skipped": len(skipped),
        "skippedProjects": skipped,
        "saved": saved,
    }


def update_lead(project: dict) -> dict:
    config = load_config()
    query = config.get("query", {})
    table = query.get("table")
    field_map = query.get("field_map", {})
    if not table:
        raise ApiError("config.local.json 缺少 query.table。")
    ensure_sales_workflow_columns(config, table)

    db_id = str(project.get("dbId") or "").strip()
    if not db_id.isdigit():
        raise ApiError("缺少有效的数据库项目 ID，刷新数据库后再保存。")

    key_to_column = {key: field_map.get(key) for key in EXPORT_FIELDS + META_FIELDS}
    key_to_column.update(WORKFLOW_COLUMNS)
    assignments = []
    for key, column in key_to_column.items():
        if not column or key not in project:
            continue
        if key == "queued":
            value = "1" if bool(project.get(key)) else "0"
        else:
            value = sql_literal(lead_value(project, key))
        assignments.append(f"{quote_identifier(column)}={value}")
    if not assignments:
        raise ApiError("没有可保存的项目字段。")

    sql = (
        f"UPDATE {quote_identifier(table)} SET {', '.join(assignments)} "
        f"WHERE `id`={int(db_id)} LIMIT 1"
    )
    run_mysql(sql, config)
    return {"ok": True, "dbId": db_id, "savedAt": now_text()}


def ensure_investigation_tables(config: dict) -> None:
    run_mysql(
        """
CREATE TABLE IF NOT EXISTS `tungee_seats` (
  `id` bigint unsigned NOT NULL AUTO_INCREMENT,
  `name` varchar(128) NOT NULL,
  `profile_key` varchar(128) NOT NULL DEFAULT '',
  `status` varchar(32) NOT NULL DEFAULT 'active',
  `state` varchar(32) NOT NULL DEFAULT 'idle',
  `today_count` int unsigned NOT NULL DEFAULT 0,
  `last_used_at` datetime DEFAULT NULL,
  `last_error` varchar(1000) NOT NULL DEFAULT '',
  `created_at` timestamp NOT NULL DEFAULT CURRENT_TIMESTAMP,
  `updated_at` timestamp NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
  PRIMARY KEY (`id`),
  UNIQUE KEY `uniq_seat_name` (`name`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci
""",
        config,
    )
    run_mysql(
        """
CREATE TABLE IF NOT EXISTS `investigation_tasks` (
  `id` bigint unsigned NOT NULL AUTO_INCREMENT,
  `lead_id` bigint unsigned NOT NULL,
  `company_name` varchar(512) NOT NULL DEFAULT '',
  `project_name` varchar(512) NOT NULL DEFAULT '',
  `platforms` varchar(255) NOT NULL DEFAULT '探迹,豆包,百度',
  `status` varchar(32) NOT NULL DEFAULT 'queued',
  `priority` int NOT NULL DEFAULT 100,
  `assigned_seat_id` bigint unsigned DEFAULT NULL,
  `requested_by` varchar(128) NOT NULL DEFAULT '',
  `result_json` json DEFAULT NULL,
  `error_text` text,
  `started_at` datetime DEFAULT NULL,
  `finished_at` datetime DEFAULT NULL,
  `created_at` timestamp NOT NULL DEFAULT CURRENT_TIMESTAMP,
  `updated_at` timestamp NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
  PRIMARY KEY (`id`),
  KEY `idx_task_status_priority` (`status`,`priority`,`id`),
  KEY `idx_task_lead` (`lead_id`),
  KEY `idx_task_seat` (`assigned_seat_id`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci
""",
        config,
    )
    run_mysql(
        """
CREATE TABLE IF NOT EXISTS `investigation_intake` (
  `id` bigint unsigned NOT NULL AUTO_INCREMENT,
  `source_id` varchar(128) NOT NULL DEFAULT '',
  `company_name` varchar(512) NOT NULL DEFAULT '',
  `project_name` varchar(512) NOT NULL DEFAULT '',
  `project_json` json NOT NULL,
  `task_id` bigint unsigned DEFAULT NULL,
  `status` varchar(32) NOT NULL DEFAULT 'queued',
  `error_text` text,
  `created_at` timestamp NOT NULL DEFAULT CURRENT_TIMESTAMP,
  `updated_at` timestamp NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
  PRIMARY KEY (`id`),
  UNIQUE KEY `uniq_intake_source` (`source_id`),
  KEY `idx_intake_status` (`status`,`id`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci
""",
        config,
    )
    run_mysql(
        "ALTER TABLE `investigation_tasks` MODIFY COLUMN `lead_id` bigint unsigned DEFAULT NULL",
        config,
    )
    count = run_mysql("SELECT COUNT(*) FROM `tungee_seats`", config).strip()
    if count == "0":
        run_mysql(
            "INSERT INTO `tungee_seats` (`name`,`profile_key`) VALUES "
            "('探迹席位 A','seat-a'),('探迹席位 B','seat-b')",
            config,
        )


def json_rows(sql: str, config: dict) -> list[dict]:
    rows = []
    for line in run_mysql(sql, config).splitlines():
        line = line.strip()
        if line:
            rows.append(json.loads(line))
    return rows


def fetch_tungee_seats(config: dict | None = None) -> list[dict]:
    config = config or load_config()
    return json_rows(
        "SELECT JSON_OBJECT("
        "'id', CAST(`id` AS CHAR), 'name', `name`, 'profileKey', `profile_key`, "
        "'status', `status`, 'state', `state`, 'todayCount', `today_count`, "
        "'lastUsedAt', IFNULL(CAST(`last_used_at` AS CHAR), ''), 'lastError', `last_error`) "
        "FROM `tungee_seats` ORDER BY `id`",
        config,
    )


def fetch_investigation_center() -> dict:
    config = load_config()
    ensure_investigation_tables(config)
    seats = fetch_tungee_seats(config)
    tasks = json_rows(
        "SELECT JSON_OBJECT("
        "'id', CAST(t.`id` AS CHAR), 'leadId', IFNULL(CAST(t.`lead_id` AS CHAR), ''), "
        "'intakeId', IFNULL(CAST(i.`id` AS CHAR), ''), "
        "'companyName', t.`company_name`, 'projectName', t.`project_name`, "
        "'platforms', t.`platforms`, 'status', t.`status`, 'priority', t.`priority`, "
        "'assignedSeatId', IFNULL(CAST(t.`assigned_seat_id` AS CHAR), ''), "
        "'assignedSeatName', IFNULL(s.`name`, ''), 'requestedBy', t.`requested_by`, "
        "'sourceId', IFNULL(i.`source_id`, ''), "
        "'recordDate', IFNULL(JSON_UNQUOTE(JSON_EXTRACT(i.`project_json`, '$.recordDate')), ''), "
        "'error', IFNULL(t.`error_text`, ''), 'createdAt', CAST(t.`created_at` AS CHAR), "
        "'startedAt', IFNULL(CAST(t.`started_at` AS CHAR), ''), "
        "'finishedAt', IFNULL(CAST(t.`finished_at` AS CHAR), '')) "
        "FROM `investigation_tasks` t LEFT JOIN `tungee_seats` s ON s.`id`=t.`assigned_seat_id` "
        "LEFT JOIN `investigation_intake` i ON i.`task_id`=t.`id` "
        "ORDER BY FIELD(t.`status`,'doubao_running','running','research_running','doubao_queued','tungee_queued','queued','tungee_done','research_failed','failed','completed','superseded','cancelled'), "
        "t.`priority`, t.`id` DESC LIMIT 200",
        config,
    )
    for task in tasks:
        task["error"] = friendly_tungee_error(task.get("error", ""))
    return {"ok": True, "seats": seats, "tasks": tasks}


def create_investigation_task(body: dict, auto_dispatch: bool = True) -> dict:
    config = load_config()
    ensure_investigation_tables(config)
    query = config.get("query", {})
    table = query.get("table")
    field_map = query.get("field_map", {})
    lead_id = str(body.get("leadId") or "").strip()
    if not lead_id.isdigit():
        raise ApiError("请选择已入库的项目后再加入调查队列。")

    project_name_column = field_map.get("projectName")
    company_name_column = field_map.get("projectCompany")
    main_company_column = field_map.get("mainCompany")
    if not table or not project_name_column or not company_name_column:
        raise ApiError("项目字段映射不完整。")
    rows = json_rows(
        "SELECT JSON_OBJECT("
        f"'projectName', IFNULL(CAST({quote_identifier(project_name_column)} AS CHAR), ''), "
        f"'companyName', IFNULL(CAST({quote_identifier(company_name_column)} AS CHAR), ''), "
        f"'mainCompany', {value_expr({'mainCompany': main_company_column}, 'mainCompany')}) "
        f"FROM {quote_identifier(table)} WHERE `id`={int(lead_id)} LIMIT 1",
        config,
    )
    if not rows:
        raise ApiError("项目不存在或已被删除。")
    allow_refresh = bool(body.get("allowRefresh"))
    protected_statuses = "'queued','doubao_queued','tungee_queued','running','tungee_done','research_running','doubao_running','research_failed'"
    if not allow_refresh:
        protected_statuses += ",'completed'"
    existing = run_mysql(
        "SELECT `id` FROM `investigation_tasks` "
        f"WHERE `lead_id`={int(lead_id)} AND `status` IN "
        f"({protected_statuses}) "
        "ORDER BY `id` DESC LIMIT 1",
        config,
    ).strip()
    if existing:
        payload = {"ok": True, "duplicate": True, "taskId": existing}
        return payload if body.get("skipCenter") else {**payload, **fetch_investigation_center()}

    platforms = str(body.get("platforms") or "探迹,豆包,百度")[:255]
    requested_by = str(body.get("requestedBy") or "")[:128]
    priority = int(body.get("priority") or 100)
    project = rows[0]
    main_company = normalize_enterprise_target(project.get("mainCompany", ""))
    project_company = normalize_enterprise_target(project.get("companyName", ""))
    requested_target = normalize_enterprise_target(body.get("targetCompany", ""))
    target_company = requested_target if looks_like_enterprise_name(requested_target) else (
        main_company if looks_like_enterprise_name(main_company) else project_company
    )
    if not looks_like_enterprise_name(target_company):
        raise ApiError("该线索没有可用于探迹查询的有效企业全称。")
    run_mysql(
        "INSERT INTO `investigation_tasks` "
        "(`lead_id`,`company_name`,`project_name`,`platforms`,`priority`,`requested_by`) VALUES ("
        f"{int(lead_id)}, {sql_literal(target_company)}, {sql_literal(project.get('projectName', ''))}, "
        f"{sql_literal(platforms)}, {priority}, {sql_literal(requested_by)})",
        config,
    )
    run_mysql(f"UPDATE {quote_identifier(table)} SET `queued`=1 WHERE `id`={int(lead_id)} LIMIT 1", config)
    if auto_dispatch:
        start_auto_dispatch()
    payload = {"ok": True, "duplicate": False}
    return payload if body.get("skipCenter") else {**payload, **fetch_investigation_center()}


def backfill_shared_company_data() -> dict:
    """Reuse verified company facts across projects without overwriting populated fields."""
    config = load_config()
    rows = json_rows(
        "SELECT JSON_OBJECT('id',CAST(`id` AS CHAR),'projectCompany',`project_unit`,'projectContact',`contact_person`,"
        "'projectPhone',`phone`,'projectAddress',`address`,'relationGraph',`relation_graph`,'investor',`investor`,"
        "'mainCompany',`main_company`,'mainContact',`main_contact`,"
        "'mainPhone',`main_phone`,'mainAddress',`main_address`) FROM `sales_leads` ORDER BY `updated_at` DESC,`id` DESC",
        config,
    )
    main_facts: dict[str, dict[str, str]] = {}
    project_facts: dict[str, dict[str, str]] = {}
    for row in rows:
        main_company = valid_public_value(row.get("mainCompany"))
        if main_company:
            facts = main_facts.setdefault(main_company, {})
            for key in ("mainContact", "mainPhone", "mainAddress"):
                value = valid_public_value(row.get(key))
                if value and not facts.get(key):
                    facts[key] = value
        project_company = valid_public_value(row.get("projectCompany"))
        if project_company:
            facts = project_facts.setdefault(project_company, {})
            for key in ("projectContact", "projectPhone", "projectAddress", "relationGraph", "investor"):
                value = valid_public_value(row.get(key))
                if value and not facts.get(key):
                    facts[key] = value

    columns = {
        "mainContact": "main_contact", "mainPhone": "main_phone", "mainAddress": "main_address",
        "projectContact": "contact_person", "projectPhone": "phone", "projectAddress": "address",
        "relationGraph": "relation_graph", "investor": "investor",
    }
    updated_rows = 0
    filled_fields = 0
    for row in rows:
        assignments = []
        for facts in (
            main_facts.get(valid_public_value(row.get("mainCompany")), {}),
            project_facts.get(valid_public_value(row.get("projectCompany")), {}),
        ):
            for key, value in facts.items():
                if not valid_public_value(row.get(key)):
                    assignments.append(f"{quote_identifier(columns[key])}={sql_literal(value)}")
                    row[key] = value
        if assignments:
            run_mysql(f"UPDATE `sales_leads` SET {', '.join(assignments)} WHERE `id`={int(row['id'])} LIMIT 1", config)
            updated_rows += 1
            filled_fields += len(assignments)
    return {"updatedRows": updated_rows, "filledFields": filled_fields}


def queue_incomplete_company_refreshes(limit: int = 60, auto_dispatch: bool = True) -> dict:
    """Create at most one v2 Tungee contact refresh per company across all record dates."""
    config = load_config()
    task_history = json_rows(
        "SELECT JSON_OBJECT('companyName',`company_name`,'status',`status`) FROM `investigation_tasks` "
        "WHERE `requested_by`='缺失信息补全v2'",
        config,
    )
    attempts: dict[str, int] = {}
    completed = set()
    active = set()
    for task in task_history:
        company = enterprise_name_key(task.get("companyName"))
        status = str(task.get("status") or "")
        if not company:
            continue
        attempts[company] = attempts.get(company, 0) + 1
        if status == "completed":
            completed.add(company)
        if status in {"queued", "tungee_queued", "running", "tungee_done", "research_running"}:
            active.add(company)
    leads = json_rows(
        "SELECT JSON_OBJECT('leadId',CAST(`id` AS CHAR),'projectCompany',`project_unit`,'projectContact',`contact_person`,"
        "'projectPhone',`phone`,'relationGraph',`relation_graph`,'projectSituation',`project_desc`,"
        "'projectProgress',`project_status`,'projectLocation',`location`,'mainCompany',`main_company`) "
        "FROM `sales_leads` WHERE `phone`='' OR `contact_person`='' OR `main_phone`='' OR `main_contact`='' "
        "OR `main_address`='' OR COALESCE(`relation_graph`,'')='' OR COALESCE(`project_desc`,'')='' "
        "OR COALESCE(`project_status`,'')='' OR COALESCE(`location`,'')='' "
        "ORDER BY `record_date` DESC,`id` DESC LIMIT 500",
        config,
    )
    queued = 0
    targets = set(active) | set(completed)
    for lead in leads:
        main_company = normalize_enterprise_target(lead.get("mainCompany"))
        project_company = normalize_enterprise_target(lead.get("projectCompany"))
        needs_project_company = (
            not valid_public_value(lead.get("projectPhone"))
            or not valid_public_value(lead.get("projectContact"))
            or not valid_public_value(lead.get("relationGraph"))
        )
        needs_research = (
            not valid_public_value(lead.get("projectSituation"))
            or not valid_public_value(lead.get("projectProgress"))
            or not valid_public_value(lead.get("projectLocation"))
        )
        target = project_company if needs_project_company and looks_like_enterprise_name(project_company) else main_company
        if not looks_like_enterprise_name(target):
            target = project_company
        target_key = enterprise_name_key(target)
        if not looks_like_enterprise_name(target) or target_key in targets or attempts.get(target_key, 0) >= 3:
            continue
        result = create_investigation_task(
            {
                "leadId": lead["leadId"],
                "platforms": "探迹,豆包,百度" if needs_research else "探迹",
                "priority": 80,
                "requestedBy": "缺失信息补全v2",
                "allowRefresh": True,
                "targetCompany": target,
                "skipCenter": True,
            },
            auto_dispatch=False,
        )
        targets.add(target_key)
        attempts[target_key] = attempts.get(target_key, 0) + 1
        if not result.get("duplicate"):
            queued += 1
        if queued >= limit:
            break
    if auto_dispatch and queued:
        start_auto_dispatch()
    return {"queued": queued, "attemptedCompanies": len(attempts), "completedCompanies": len(completed)}


def normalize_queued_investigation_targets() -> dict:
    """Repair narrative company names already persisted in the pending queue."""
    config = load_config()
    rows = json_rows(
        "SELECT JSON_OBJECT('id',CAST(`id` AS CHAR),'companyName',`company_name`) "
        "FROM `investigation_tasks` WHERE `status` IN ('queued','tungee_queued')",
        config,
    )
    updated = 0
    merged = 0
    seen = set()
    for row in rows:
        original = valid_public_value(row.get("companyName"))
        cleaned = normalize_enterprise_target(original)
        if not looks_like_enterprise_name(cleaned):
            continue
        cleaned_key = enterprise_name_key(cleaned)
        if cleaned_key in seen:
            run_mysql(
                "UPDATE `investigation_tasks` SET `status`='cancelled',"
                f"`error_text`={sql_literal('同一控股公司重复补全，已自动合并')},`finished_at`=NOW() "
                f"WHERE `id`={int(row['id'])} LIMIT 1",
                config,
            )
            merged += 1
            continue
        seen.add(cleaned_key)
        if cleaned != original:
            run_mysql(
                f"UPDATE `investigation_tasks` SET `company_name`={sql_literal(cleaned)} "
                f"WHERE `id`={int(row['id'])} LIMIT 1",
                config,
            )
            updated += 1
    return {"checked": len(rows), "updated": updated, "merged": merged}


def create_tungee_seat(body: dict) -> dict:
    config = load_config()
    ensure_investigation_tables(config)
    name = str(body.get("name") or "").strip()
    if not name:
        raise ApiError("席位名称不能为空。")
    profile_key = str(body.get("profileKey") or f"seat-{uuid.uuid4().hex[:8]}")[:128]
    run_mysql(
        "INSERT INTO `tungee_seats` (`name`,`profile_key`,`status`,`state`) VALUES "
        f"({sql_literal(name[:128])}, {sql_literal(profile_key)}, 'disabled', 'offline')",
        config,
    )
    return {"ok": True, "seats": fetch_tungee_seats(config)}


def toggle_tungee_seat(body: dict) -> dict:
    config = load_config()
    ensure_investigation_tables(config)
    seat_id = str(body.get("seatId") or "").strip()
    if not seat_id.isdigit():
        raise ApiError("席位 ID 无效。")
    rows = json_rows(
        "SELECT JSON_OBJECT('status',`status`,'state',`state`) FROM `tungee_seats` "
        f"WHERE `id`={int(seat_id)} LIMIT 1",
        config,
    )
    if not rows:
        raise ApiError("席位不存在。")
    if rows[0].get("state") == "busy":
        raise ApiError("席位正在执行任务，不能停用。")
    if rows[0].get("status") != "active":
        raise ApiError("停用席位必须通过“登录并启用”验证会话后启用。")
    run_mysql(
        "UPDATE `tungee_seats` SET `status`='disabled', `state`='offline' "
        f"WHERE `id`={int(seat_id)} LIMIT 1",
        config,
    )
    return {"ok": True, "seats": fetch_tungee_seats(config)}


def seat_runtime_config(seat_id: str, profile_key: str) -> tuple[int, Path]:
    numeric_id = int(seat_id)
    port = 9222 + (numeric_id - 1) * 2
    safe_key = re.sub(r"[^a-zA-Z0-9_-]+", "-", profile_key or f"seat-{numeric_id}").strip("-")
    if numeric_id == 1 and safe_key == "seat-a":
        profile_dir = PROJECT_CODE_DIR / "browser_debug_profile"
    else:
        profile_dir = PROJECT_CODE_DIR.parent / "data" / "browser_profiles" / safe_key
    return port, profile_dir


def open_debug_browser_url(port: int, url: str) -> None:
    endpoint = f"http://127.0.0.1:{port}/json/new?{quote(url, safe='')}"
    request = urllib.request.Request(endpoint, method="PUT")
    with urllib.request.urlopen(request, timeout=5):
        pass


def open_tungee_seat_login(body: dict) -> dict:
    config = load_config()
    ensure_investigation_tables(config)
    seat_id = str(body.get("seatId") or "").strip()
    if not seat_id.isdigit():
        raise ApiError("席位 ID 无效。")
    rows = json_rows(
        "SELECT JSON_OBJECT('id',CAST(`id` AS CHAR),'profileKey',`profile_key`,'status',`status`) "
        f"FROM `tungee_seats` WHERE `id`={int(seat_id)} LIMIT 1",
        config,
    )
    if not rows:
        raise ApiError("席位不存在。")
    if str(MODULE_CODE_DIR) not in sys.path:
        sys.path.insert(0, str(MODULE_CODE_DIR))
    from enrich_tungee import (
        TUNGEE_SEARCH,
        TUNGEE_SIGN_IN,
        attach_driver,
        chrome_running,
        enter_bidding_product,
        enter_enterprise_query,
        is_logged_in,
        launch_chrome,
        select_account_login,
    )

    port, profile_dir = seat_runtime_config(seat_id, rows[0].get("profileKey", ""))
    if not chrome_running(port):
        profile_dir.mkdir(parents=True, exist_ok=True)
        launch_chrome(port, profile_dir)
    driver = None
    message = "探迹登录窗口已在服务器打开，请完成一次登录。"
    try:
        driver = attach_driver(port, profile_dir)
        if is_logged_in(driver):
            if not enter_bidding_product(driver) or not enter_enterprise_query(driver):
                driver.get(TUNGEE_SEARCH)
            run_mysql(
                f"UPDATE `tungee_seats` SET `status`='active',`state`='idle',`last_error`='' WHERE `id`={int(seat_id)} LIMIT 1",
                config,
            )
            message = "探迹已登录，已直接打开企业查询页面。"
        else:
            driver.get(TUNGEE_SIGN_IN)
            time.sleep(1)
            select_account_login(driver)
    finally:
        if driver:
            driver.service.stop()
    return {
        "ok": True,
        "seatId": seat_id,
        "port": port,
        "message": message,
        "seats": fetch_tungee_seats(config),
    }


def tungee_result_to_project(result: dict) -> dict:
    mapping = {
        "project_company_address": "projectAddress",
        "project_company_contact": "projectContact",
        "project_company_phone": "projectPhone",
        "address": "projectAddress",
        "main_company": "mainCompany",
        "main_contact": "mainContact",
        "main_phone": "mainPhone",
        "main_address": "mainAddress",
        "investor": "investor",
        "relation_graph": "relationGraph",
        "remark": "remark",
    }
    return {target: result.get(source) for source, target in mapping.items() if result.get(source)}


def annotate_tungee_target_source(result: dict, task_result: dict) -> dict:
    """Expose filing-company fallback to employees instead of hiding it in task JSON."""
    if not isinstance(result, dict) or task_result.get("tungee_target_source") != "filing_company_fallback":
        return result
    note = "主体识别说明：豆包未确认最终主体，本次仅以备案公司作为探迹股权穿透起点；该公司不是已确认的最终主体。"
    enriched = dict(result)
    relation = str(enriched.get("relation_graph") or "").strip()
    remark = str(enriched.get("remark") or "").strip()
    enriched["relation_graph"] = f"{note}\n{relation}" if relation else note
    enriched["remark"] = f"{note}；{remark}" if remark else note
    return enriched


def suggested_tungee_company(result: dict, fallback: str = "") -> str:
    """Pick the company name that Doubao identified for the next Tungee lookup."""
    if not isinstance(result, dict):
        return fallback
    normalized = {str(key).strip(): value for key, value in result.items()}
    equity = normalized.get("股权穿透") if isinstance(normalized.get("股权穿透"), dict) else {}
    candidates = [
        normalized.get("最终控股公司"),
        normalized.get("建议探迹查询控股公司"),
        normalized.get("最终主体公司"),
        normalized.get("建议探迹查询主体公司"),
        normalized.get("主体公司"),
        normalized.get("实际投资方"),
        normalized.get("真正出资方"),
        normalized.get("最终实际控股方"),
        normalized.get("投资方"),
        normalized.get("投资主体"),
        normalized.get("业主单位"),
        normalized.get("实际业主单位"),
        normalized.get("项目投资主体"),
        equity.get("主体公司"),
        equity.get("实际业主单位"),
        equity.get("最终实际控股方"),
        equity.get("实际投资方"),
    ]
    for value in candidates:
        text = valid_public_value(value)
        is_placeholder = any(marker in text for marker in ("某某", "示例", "测试", "公司全称"))
        if text and not is_placeholder and text not in {"[未确认]", "[未公开]"} and len(text) <= 120:
            return text
    return fallback


def normalized_result_value(result: dict, *keys: str) -> str:
    """Read Doubao keys defensively because web responses may add whitespace."""
    normalized = {str(key).strip(): value for key, value in result.items()} if isinstance(result, dict) else {}
    for key in keys:
        value = valid_public_value(normalized.get(key))
        if value:
            return value
    return ""


def finish_tungee_task(task_id: str, seat_id: str, result: dict) -> None:
    config = load_config()
    query = config.get("query", {})
    table = query.get("table")
    rows = json_rows(
        "SELECT JSON_OBJECT('leadId',IFNULL(CAST(t.`lead_id` AS CHAR),''),'intakeId',IFNULL(CAST(i.`id` AS CHAR),''),'platforms',t.`platforms`,"
        "'companyName',t.`company_name`,'existing',IFNULL(CAST(t.`result_json` AS CHAR),'{}')) "
        f"FROM `investigation_tasks` t LEFT JOIN `investigation_intake` i ON i.`task_id`=t.`id` WHERE t.`id`={int(task_id)} LIMIT 1",
        config,
    )
    if not rows:
        raise ApiError("调查任务不存在。")
    lead_id = rows[0]["leadId"]
    intake_id = rows[0].get("intakeId", "")
    platforms = rows[0].get("platforms", "")
    if intake_id:
        try:
            combined = json.loads(rows[0].get("existing") or "{}")
        except Exception:
            combined = {}
        if not isinstance(combined, dict):
            combined = {}
        combined["tungee"] = result
        run_mysql(
            "UPDATE `investigation_tasks` SET `result_json`="
            f"{sql_literal(json.dumps(combined, ensure_ascii=False))}, `finished_at`=NOW(), `error_text`='' WHERE `id`={int(task_id)} LIMIT 1; "
            f"UPDATE `investigation_intake` SET `status`='tungee_done', `error_text`='' WHERE `id`={int(intake_id)} LIMIT 1; "
            f"UPDATE `tungee_seats` SET `state`='idle', `last_error`='' WHERE `id`={int(seat_id)} LIMIT 1",
            config,
        )
        doubao_result = combined.get("doubao") if isinstance(combined.get("doubao"), dict) else {}
        finish_research_task(task_id, doubao_result, combined.get("baidu", []))
        archive_equivalent_failed_tasks(task_id, rows[0].get("companyName", ""), config)
        start_auto_dispatch()
        return
    needs_research = "豆包" in platforms or "百度" in platforms
    if lead_id:
        project_update = {"dbId": lead_id, "tungeeStatus": "已取信息", "reviewStatus": "待人工审核", "queued": needs_research}
        project_update.update(tungee_result_to_project(result))
        update_lead(project_update)
        try:
            backfill_shared_company_data()
        except Exception:
            pass
    run_mysql(
        f"UPDATE `investigation_tasks` SET `status`={sql_literal('tungee_done' if needs_research else 'completed')}, "
        f"`result_json`={sql_literal(json.dumps({'tungee': result}, ensure_ascii=False))}, `finished_at`=NOW(), `error_text`='' "
        f"WHERE `id`={int(task_id)} LIMIT 1; " +
        (f"UPDATE `investigation_intake` SET `status`={sql_literal('tungee_done')}, `error_text`='' WHERE `id`={int(rows[0].get('intakeId'))} LIMIT 1; " if rows[0].get("intakeId") and needs_research else "") +
        "UPDATE `tungee_seats` SET `state`='idle', `last_error`='' "
        f"WHERE `id`={int(seat_id)} LIMIT 1",
        config,
    )
    if needs_research:
        start_research_task(task_id)
    archive_equivalent_failed_tasks(task_id, rows[0].get("companyName", ""), config)
    start_auto_dispatch()


def fail_tungee_task(task_id: str, seat_id: str, error: str) -> None:
    config = load_config()
    raw_error = str(error)[:4000]
    message = friendly_tungee_error(raw_error)
    task_rows = json_rows(
        "SELECT JSON_OBJECT('leadId',IFNULL(CAST(t.`lead_id` AS CHAR),''),'intakeId',IFNULL(CAST(i.`id` AS CHAR),''),"
        "'existing',IFNULL(CAST(t.`result_json` AS CHAR),'{}')) "
        "FROM `investigation_tasks` t LEFT JOIN `investigation_intake` i ON i.`task_id`=t.`id` "
        f"WHERE t.`id`={int(task_id)} LIMIT 1",
        config,
    )
    diagnostic = {}
    if task_rows:
        try:
            diagnostic = json.loads(task_rows[0].get("existing") or "{}")
        except Exception:
            diagnostic = {}
    if not isinstance(diagnostic, dict):
        diagnostic = {}
    diagnostic["diagnostic"] = {
        "summary": message,
        "raw_error": raw_error,
        "failed_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    if task_rows and task_rows[0].get("leadId"):
        update_lead(
            {
                "dbId": task_rows[0]["leadId"],
                "queued": False,
                "tungeeStatus": "未找到" if "未找到企业" in message else "执行失败",
            }
        )
    intake_clause = ""
    if task_rows and task_rows[0].get("intakeId"):
        intake_clause = (
            f"UPDATE `investigation_intake` SET `status`='failed', `error_text`={sql_literal(message)} "
            f"WHERE `id`={int(task_rows[0]['intakeId'])} LIMIT 1; "
        )
    run_mysql(
        "UPDATE `investigation_tasks` SET `status`='failed', "
        f"`result_json`={sql_literal(json.dumps(diagnostic, ensure_ascii=False))}, "
        f"`error_text`={sql_literal(message)}, `finished_at`=NOW() WHERE `id`={int(task_id)} LIMIT 1; " +
        intake_clause +
        "UPDATE `tungee_seats` SET `state`='idle', "
        f"`last_error`={sql_literal(message[:1000])} WHERE `id`={int(seat_id)} LIMIT 1",
        config,
    )


def retry_investigation_task(body: dict) -> dict:
    """Requeue an old failed task without changing its filing/company data."""
    config = load_config()
    ensure_investigation_tables(config)
    task_id = int(body.get("taskId") or 0)
    if not task_id:
        raise ApiError("缺少任务编号。")
    rows = json_rows(
        "SELECT JSON_OBJECT('status',t.`status`,'platforms',t.`platforms`,'companyName',t.`company_name`,"
        "'error',IFNULL(t.`error_text`,''),'leadId',IFNULL(CAST(t.`lead_id` AS CHAR),''),"
        "'intakeId',IFNULL(CAST(i.`id` AS CHAR),'')) "
        "FROM `investigation_tasks` t LEFT JOIN `investigation_intake` i ON i.`task_id`=t.`id` "
        f"WHERE t.`id`={task_id} LIMIT 1",
        config,
    )
    if not rows:
        raise ApiError("任务不存在。")
    task = rows[0]
    if task.get("status") not in {"failed", "research_failed"}:
        raise ApiError("只有失败任务可以重试。")
    error = str(task.get("error") or "")
    if task.get("status") == "research_failed":
        next_status = "doubao_queued"
    elif "豆包" in str(task.get("platforms") or "") and "探迹" not in error:
        next_status = "doubao_queued"
    else:
        next_status = "tungee_queued"
    intake_status = "doubao_queued" if next_status == "doubao_queued" else "tungee_queued"
    cleaned_company = normalize_enterprise_target(task.get("companyName", ""))
    run_mysql(
        f"UPDATE `investigation_tasks` SET `status`={sql_literal(next_status)},"
        f"`company_name`={sql_literal(cleaned_company)},"
        "`assigned_seat_id`=NULL,`started_at`=NULL,`finished_at`=NULL,`error_text`='' "
        f"WHERE `id`={task_id} LIMIT 1; "
        + (f"UPDATE `investigation_intake` SET `status`={sql_literal(intake_status)},`error_text`='' "
           f"WHERE `id`={int(task['intakeId'])} LIMIT 1; " if task.get("intakeId") else "")
        + (f"UPDATE `sales_leads` SET `queued`=1 WHERE `id`={int(task['leadId'])} LIMIT 1; " if task.get("leadId") else ""),
        config,
    )
    start_auto_dispatch()
    return {"ok": True, "taskId": str(task_id)}


def run_tungee_task(task_id: str, seat_id: str) -> None:
    driver = None
    try:
        config = load_config()
        rows = json_rows(
            "SELECT JSON_OBJECT('companyName',t.`company_name`,'projectName',t.`project_name`,"
            "'result',IFNULL(CAST(t.`result_json` AS CHAR),'{}'),'profileKey',s.`profile_key`) FROM `investigation_tasks` t "
            "JOIN `tungee_seats` s ON s.`id`=t.`assigned_seat_id` "
            f"WHERE t.`id`={int(task_id)} AND s.`id`={int(seat_id)} LIMIT 1",
            config,
        )
        if not rows:
            raise ApiError("任务或席位不存在。")
        task = rows[0]
        try:
            task_result = json.loads(task.get("result") or "{}")
        except Exception:
            task_result = {}
        company_name = normalize_enterprise_target(
            valid_public_value(task_result.get("tungee_target")) or valid_public_value(task.get("companyName"))
        )
        if not looks_like_enterprise_name(company_name):
            raise ApiError(
                f"探迹查询目标不是有效企业全称，已停止执行。项目名称“{task.get('projectName') or ''}”仅供豆包调查。"
            )
        if str(MODULE_CODE_DIR) not in sys.path:
            sys.path.insert(0, str(MODULE_CODE_DIR))
        from enrich_tungee import (
            TUNGEE_SEARCH,
            attach_driver,
            enter_bidding_product,
            enter_enterprise_query,
            investigate_company_chain,
            is_logged_in,
            is_recoverable_browser_error,
            restore_session_cookies,
            save_session_cookies,
        )

        port, profile_dir = seat_runtime_config(seat_id, task.get("profileKey", ""))
        profile_dir.mkdir(parents=True, exist_ok=True)
        session_path = profile_dir / "tungee_session.json"
        legacy_session_path = PROJECT_CODE_DIR / "tungee_session.json"
        result = {}
        for attempt in range(2):
            try:
                driver = attach_driver(port, profile_dir)
                logged_in = is_logged_in(driver)
                if not logged_in and session_path.exists():
                    logged_in = restore_session_cookies(driver, session_path)
                if not logged_in and int(seat_id) == 1 and legacy_session_path.exists():
                    logged_in = restore_session_cookies(driver, legacy_session_path)
                if not logged_in:
                    raise ApiError(f"席位未登录探迹，请先打开席位登录窗口（端口 {port}）。")
                save_session_cookies(driver, session_path)

                # A valid logged-in session can open enterprise search directly.
                # Only fall back to the product selector when direct navigation fails.
                if not enter_enterprise_query(driver):
                    if not enter_bidding_product(driver) or not enter_enterprise_query(driver):
                        raise ApiError("探迹已登录，但无法进入“拓客 · 招投标版”的企业查询。")
                result = investigate_company_chain(driver, company_name)
                result = annotate_tungee_target_source(result, task_result)
                break
            except Exception as attempt_error:
                recoverable = is_recoverable_browser_error(attempt_error) or "无法进入“拓客" in str(attempt_error)
                if driver is not None:
                    try:
                        driver.service.stop()
                    except Exception:
                        pass
                    driver = None
                if attempt == 0 and recoverable:
                    time.sleep(2)
                    continue
                raise
        if not result:
            raise ApiError(f"探迹未找到企业：{company_name}")
        if not any(result.values()):
            raise ApiError("探迹详情页未提取到可用字段。")
        finish_tungee_task(task_id, seat_id, result)
    except Exception as exc:
        fail_tungee_task(task_id, seat_id, str(exc))
    finally:
        if driver is not None:
            try:
                driver.service.stop()
            except Exception:
                pass


def doubao_profile_dir() -> Path:
    return PROJECT_CODE_DIR.parent / "data" / "browser_profiles" / "doubao"


def open_doubao_login() -> dict:
    if str(MODULE_CODE_DIR) not in sys.path:
        sys.path.insert(0, str(MODULE_CODE_DIR))
    from llm_doubao_web import DOUBAO_DEBUG_PORT, DOUBAO_URL, chrome_running, launch_chrome

    profile_dir = doubao_profile_dir()
    profile_dir.mkdir(parents=True, exist_ok=True)
    if not chrome_running(DOUBAO_DEBUG_PORT):
        launch_chrome(DOUBAO_DEBUG_PORT, profile_dir)
    open_debug_browser_url(DOUBAO_DEBUG_PORT, DOUBAO_URL)
    return {
        "ok": True,
        "port": DOUBAO_DEBUG_PORT,
        "message": "豆包登录窗口已在服务器打开，请完成一次登录。",
        **fetch_investigation_center(),
    }


def fetch_lead_for_research(lead_id: str) -> dict:
    config = load_config()
    query = config.get("query", {})
    table = query.get("table")
    field_map = query.get("field_map", {})
    if not table:
        raise ApiError("项目表配置缺失。")
    json_args = ["'dbId'", "CAST(`id` AS CHAR)"]
    for key in EXPORT_FIELDS + META_FIELDS:
        json_args.extend([f"'{key}'", value_expr(field_map, key)])
    rows = json_rows(
        f"SELECT JSON_OBJECT({', '.join(json_args)}) FROM {quote_identifier(table)} "
        f"WHERE `id`={int(lead_id)} LIMIT 1",
        config,
    )
    if not rows:
        raise ApiError("待调查项目不存在。")
    return rows[0]


def project_to_doubao_row(project: dict) -> dict:
    return {
        "company_name": project.get("projectName", ""),
        "project_unit": project.get("projectCompany", ""),
        "credit_code": project.get("recordCode", ""),
        "record_date": project.get("recordDate", ""),
        "contact_person": project.get("projectContact", ""),
        "phone": project.get("projectPhone", ""),
        "address": project.get("projectAddress", ""),
        "main_company": project.get("mainCompany", ""),
        "investor": project.get("investor", ""),
        "project_desc": project.get("projectSituation", ""),
        "project_status": project.get("projectProgress", ""),
        "remark": project.get("remark", ""),
    }


def valid_public_value(value) -> str:
    text = str(value or "").strip()
    return "" if text in {"[未公开]", "[未确认]", "未公开", "未确认"} else text


def normalize_enterprise_target(value: str) -> str:
    """Remove legacy narrative prefixes while preserving the enterprise full name."""
    text = unicodedata.normalize("NFKC", valid_public_value(value))
    prefixes = (
        "实际业主均为", "实际业主为", "最终主体为", "主体公司为",
        "一级股东（", "一级股东(", "一级股东为", "控股股东为", "由", "为",
    )
    changed = True
    while text and changed:
        changed = False
        for prefix in prefixes:
            if text.startswith(prefix):
                text = text[len(prefix):].strip()
                changed = True
                break
    text = text.strip(" \t\r\n：:，,；;")
    while text.endswith(("）", ")")) and text.count("（") + text.count("(") < text.count("）") + text.count(")"):
        text = text[:-1].rstrip()
    # Chinese legal names are displayed with full-width parentheses on Tungee.
    return text.replace("(", "（").replace(")", "）")


def enterprise_name_key(value: str) -> str:
    text = unicodedata.normalize("NFKC", normalize_enterprise_target(value))
    return re.sub(r"[\s\u200b\ufeff]+", "", text).casefold()


def archive_equivalent_failed_tasks(task_id: str | int, company_name: str, config: dict | None = None) -> int:
    """Archive older failures once the same normalized company is searchable."""
    target_key = enterprise_name_key(company_name)
    if not target_key:
        return 0
    config = config or load_config()
    rows = json_rows(
        "SELECT JSON_OBJECT('id',CAST(`id` AS CHAR),'companyName',`company_name`) "
        "FROM `investigation_tasks` WHERE `status`='failed' "
        f"AND `id`<{int(task_id)}",
        config,
    )
    ids = [int(row["id"]) for row in rows if enterprise_name_key(row.get("companyName", "")) == target_key]
    if not ids:
        return 0
    run_mysql(
        "UPDATE `investigation_tasks` SET `status`='superseded',"
        f"`error_text`={sql_literal('同一企业后续查询已成功，本条历史失败已归档（企业名称已重新匹配）')} "
        f"WHERE `id` IN ({','.join(map(str, ids))})",
        config,
    )
    return len(ids)


def archive_resolved_historical_failures() -> int:
    """Repair legacy logs after deployment without hiding newer failures."""
    config = load_config()
    rows = json_rows(
        "SELECT JSON_OBJECT('id',CAST(`id` AS CHAR),'companyName',`company_name`,'status',`status`) "
        "FROM `investigation_tasks` WHERE `status` IN ('failed','completed','tungee_done','research_running','research_failed')",
        config,
    )
    latest_success: dict[str, tuple[int, str]] = {}
    for row in rows:
        if row.get("status") == "failed":
            continue
        key = enterprise_name_key(row.get("companyName", ""))
        task_id = int(row.get("id") or 0)
        if key and task_id > latest_success.get(key, (0, ""))[0]:
            latest_success[key] = (task_id, row.get("companyName", ""))
    ids = []
    for row in rows:
        if row.get("status") != "failed":
            continue
        success = latest_success.get(enterprise_name_key(row.get("companyName", "")))
        if success and int(row.get("id") or 0) < success[0]:
            ids.append(int(row["id"]))
    if not ids:
        return 0
    run_mysql(
        "UPDATE `investigation_tasks` SET `status`='superseded',"
        f"`error_text`={sql_literal('同一企业后续查询已成功，本条历史失败已归档（企业名称已重新匹配）')} "
        f"WHERE `id` IN ({','.join(map(str, ids))})",
        config,
    )
    return len(ids)


def friendly_tungee_error(error: str) -> str:
    """Convert Selenium diagnostics into a short employee-facing conclusion."""
    raw = str(error or "").strip()
    lower = raw.lower()
    if not raw:
        return ""
    if "同一主体公司重复补全" in raw or "同一控股公司重复补全" in raw:
        return "同一控股公司重复补全，已自动合并"
    if "stale element" in lower or "target frame detached" in lower:
        return "探迹页面刷新导致元素失效（技术异常，可重试）"
    if any(marker in lower for marker in ("httpconnectionpool", "read timed out", "connectionreseterror", "connection aborted")):
        return "探迹浏览器连接超时（技术异常，可重试）"
    if "未登录探迹" in raw:
        match = re.search(r"端口\s*(\d+)", raw)
        suffix = f"（端口 {match.group(1)}）" if match else ""
        return f"探迹登录已失效，请管理员重新登录{suffix}"
    if "无法进入“拓客" in raw or "产品权限" in raw:
        return "探迹登录有效，但企业查询页面暂时无法打开；系统已尝试直达并重试"
    if "未找到企业" in raw:
        return raw.splitlines()[0][:240]
    first_line = raw.splitlines()[0].removeprefix("Message:").strip()
    return first_line[:240] or "探迹任务执行失败"


def looks_like_enterprise_name(value: str) -> bool:
    """Prevent project titles from ever being submitted to Tungee enterprise search."""
    text = normalize_enterprise_target(value)
    if not text or len(text) > 160:
        return False
    if re.search(r"(?:工程|项目|充电站|光伏发电|停车场)$", text):
        return False
    return bool(re.search(r"(?:有限责任公司|有限公司|股份公司|股份有限公司|集团|分公司|合伙企业(?:（有限合伙）)?|企业管理中心（有限合伙）|公司|企业)$", text))


def fetch_baidu_public_clues(project: dict) -> list[dict]:
    from bs4 import BeautifulSoup

    company = project.get("projectCompany") or project.get("projectName") or ""
    project_name = project.get("projectName") or ""
    queries = [
        f"{company} 光伏 招标 EPC",
        f"{company} 融资 扩建 新能源",
        f"{project_name} 备案 项目进展",
    ]
    clues = []
    for query in queries:
        url = "https://www.baidu.com/s?" + urlencode({"wd": query})
        request = urllib.request.Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124 Safari/537.36",
                "Accept-Language": "zh-CN,zh;q=0.9",
            },
        )
        try:
            with urllib.request.urlopen(request, timeout=12) as response:
                html = response.read().decode("utf-8", errors="replace")
            soup = BeautifulSoup(html, "lxml")
            query_count = 0
            for block in soup.select("div.result, div.c-container"):
                heading = block.find("h3")
                link = heading.find("a") if heading else None
                title = heading.get_text(" ", strip=True) if heading else ""
                href = link.get("href", "") if link else ""
                snippet = block.get_text(" ", strip=True)
                if not title or not href:
                    continue
                clues.append({"source": "百度", "query": query, "title": title[:300], "snippet": snippet[:700], "url": href[:1000]})
                query_count += 1
                if query_count >= 5:
                    break
            if query_count:
                continue
            page_title = soup.title.get_text(" ", strip=True) if soup.title else ""
            clues.append(
                {
                    "source": "百度",
                    "query": query,
                    "error": "百度安全验证限制，已自动切换公共搜索备选" if "安全验证" in page_title else "百度未返回可解析结果，已自动切换公共搜索备选",
                }
            )
        except Exception as exc:
            clues.append({"source": "百度", "query": query, "error": str(exc)[:300]})

        fallback_url = "https://html.duckduckgo.com/html/?" + urlencode({"q": query})
        fallback_request = urllib.request.Request(
            fallback_url,
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124 Safari/537.36"},
        )
        try:
            with urllib.request.urlopen(fallback_request, timeout=12) as response:
                fallback_html = response.read().decode("utf-8", errors="replace")
            fallback_soup = BeautifulSoup(fallback_html, "lxml")
            for block in fallback_soup.select(".result")[:5]:
                link = block.select_one(".result__a")
                snippet_node = block.select_one(".result__snippet")
                if not link:
                    continue
                clues.append(
                    {
                        "source": "公共搜索备选",
                        "query": query,
                        "title": link.get_text(" ", strip=True)[:300],
                        "snippet": (snippet_node.get_text(" ", strip=True) if snippet_node else "")[:700],
                        "url": str(link.get("href") or "")[:1000],
                    }
                )
        except Exception as exc:
            clues.append({"source": "公共搜索备选", "query": query, "error": str(exc)[:300]})
    return clues


def doubao_result_to_project(result: dict, db_id: str, baidu_clues: list[dict] | None = None) -> dict:
    normalized = {str(key).strip(): value for key, value in result.items()} if isinstance(result, dict) else {}
    equity = normalized.get("股权穿透") if isinstance(normalized.get("股权穿透"), dict) else {}
    contacts = normalized.get("关键对接人") if isinstance(normalized.get("关键对接人"), list) else []
    contact_name = ""
    contact_phone = ""
    for contact in contacts:
        if not isinstance(contact, dict):
            continue
        if not contact_name:
            contact_name = valid_public_value(contact.get("姓名"))
        phone = valid_public_value(contact.get("电话"))
        if phone:
            contact_phone = phone
            break
    final_subject = suggested_tungee_company(normalized)
    final_controller = valid_public_value(equity.get("最终控股方"))
    actual_controller = normalized_result_value(normalized, "控股人姓名") or valid_public_value(equity.get("实际控制人")) or contact_name
    controller_phone = normalized_result_value(normalized, "控股人手机号") or contact_phone
    relation_payload = {
        "股权穿透": equity,
        "关联企业": result.get("关联企业", []),
    }
    remark_payload = {
        "备案逻辑分析": result.get("备案逻辑分析", ""),
        "历史招投标": result.get("历史招投标", []),
        "光伏EPC记录": result.get("光伏EPC记录", []),
        "潜在线索": result.get("潜在线索", []),
        "风险提示": result.get("风险提示", ""),
        "销售建议": result.get("销售建议", ""),
        "信息源": result.get("信息源", []),
        "百度公开检索线索": baidu_clues or [],
    }
    update = {
        "dbId": db_id,
        "projectSituation": normalized_result_value(normalized, "项目情况", "现在项目情况"),
        "projectProgress": normalized_result_value(normalized, "项目进展"),
        "projectLocation": normalized_result_value(normalized, "项目建设地点", "项目所在地", "项目地址"),
        "relationGraph": json.dumps(relation_payload, ensure_ascii=False, indent=2),
        "investor": " / ".join(item for item in [final_controller, valid_public_value(result.get("公司性质"))] if item),
        "remark": json.dumps(remark_payload, ensure_ascii=False, indent=2),
        "mainCompany": final_subject,
        "mainContact": actual_controller,
        "mainPhone": controller_phone,
        "reviewStatus": "待人工审核",
        "queued": False,
    }
    return {key: value for key, value in update.items() if key in {"dbId", "reviewStatus", "queued"} or value}


def finish_research_task(task_id: str, result: dict, baidu_clues: list[dict] | None = None) -> None:
    config = load_config()
    rows = json_rows(
        "SELECT JSON_OBJECT('leadId',IFNULL(CAST(t.`lead_id` AS CHAR),''),'intakeId',IFNULL(CAST(i.`id` AS CHAR),''),'existing',IFNULL(CAST(t.`result_json` AS CHAR),'{}')) "
        "FROM `investigation_tasks` t LEFT JOIN `investigation_intake` i ON i.`task_id`=t.`id` "
        f"WHERE t.`id`={int(task_id)} LIMIT 1",
        config,
    )
    if not rows:
        raise ApiError("综合调查任务不存在。")
    existing_text = rows[0].get("existing") or "{}"
    try:
        combined = json.loads(existing_text)
    except Exception:
        combined = {}
    if not isinstance(combined, dict):
        combined = {"tungee": combined}
    combined["doubao"] = result
    combined["baidu"] = baidu_clues or []
    lead_id = rows[0].get("leadId", "")
    intake_id = rows[0].get("intakeId", "")
    if lead_id:
        update_lead(doubao_result_to_project(result, lead_id, baidu_clues))
    elif intake_id:
        intake_rows = json_rows(
            f"SELECT JSON_OBJECT('projectJson',CAST(`project_json` AS CHAR)) FROM `investigation_intake` WHERE `id`={int(intake_id)} LIMIT 1",
            config,
        )
        if not intake_rows:
            raise ApiError("待调查项目不存在。")
        project = hydrate_intake_project(json.loads(intake_rows[0].get("projectJson") or "{}"), config)
        doubao_update = doubao_result_to_project(result, "", baidu_clues)
        doubao_update.pop("dbId", None)
        project.update(doubao_update)
        # Tungee is authoritative for the subject company's person and phone.
        project.update(tungee_result_to_project(combined.get("tungee", {})))
        formal = save_leads([project], auto_investigate=False)
        formal_saved = formal.get("saved") or []
        if not formal_saved or not formal_saved[0].get("dbId"):
            raise ApiError("调查完成，但正式项目线索入库失败。")
        lead_id = formal_saved[0]["dbId"]
    run_mysql(
        "UPDATE `investigation_tasks` SET `status`='completed', "
        f"`result_json`={sql_literal(json.dumps(combined, ensure_ascii=False))}, `finished_at`=NOW(), `error_text`='' "
        f"WHERE `id`={int(task_id)} LIMIT 1; "
        + (f"UPDATE `investigation_tasks` SET `lead_id`={int(lead_id)} WHERE `id`={int(task_id)} LIMIT 1; " if lead_id else "")
        + (f"UPDATE `investigation_intake` SET `status`='completed', `error_text`='' WHERE `id`={int(intake_id)} LIMIT 1" if intake_id else ""),
        config,
    )
    try:
        backfill_shared_company_data()
    except Exception:
        pass


def finish_doubao_stage(task_id: str, result: dict, baidu_clues: list[dict] | None = None, target_override: str = "") -> None:
    """Persist Doubao's owner analysis, then queue Tungee against that owner."""
    config = load_config()
    rows = json_rows(
        "SELECT JSON_OBJECT('intakeId',IFNULL(CAST(i.`id` AS CHAR),''),'companyName',t.`company_name`,"
        "'projectJson',IFNULL(CAST(i.`project_json` AS CHAR),'{}'),'existing',IFNULL(CAST(t.`result_json` AS CHAR),'{}')) "
        "FROM `investigation_tasks` t LEFT JOIN `investigation_intake` i ON i.`task_id`=t.`id` "
        f"WHERE t.`id`={int(task_id)} LIMIT 1",
        config,
    )
    if not rows or not rows[0].get("intakeId"):
        raise ApiError("备案调查任务不存在。")
    try:
        combined = json.loads(rows[0].get("existing") or "{}")
    except Exception:
        combined = {}
    if not isinstance(combined, dict):
        combined = {}
    combined["doubao"] = result
    combined["baidu"] = baidu_clues or []
    target = normalize_enterprise_target(suggested_tungee_company(result, target_override))
    target_source = "doubao"
    if not looks_like_enterprise_name(target):
        try:
            filing_project = hydrate_intake_project(json.loads(rows[0].get("projectJson") or "{}"), config)
        except Exception:
            filing_project = {}
        filing_company = normalize_enterprise_target(filing_project.get("projectCompany", ""))
        if looks_like_enterprise_name(filing_company):
            target = filing_company
            target_source = "filing_company_fallback"
    if not looks_like_enterprise_name(target):
        raise ApiError("豆包未识别出有效企业全称，备案数据也没有可查询的企业全称；项目名称仍未发送到探迹。")
    combined["tungee_target"] = target
    combined["tungee_target_source"] = target_source
    if target_source == "filing_company_fallback":
        combined["tungee_target_note"] = "豆包未确认最终主体，系统改用备案公司作为探迹股权穿透起点；该公司不是已确认的最终主体。"
    run_mysql(
        "UPDATE `investigation_tasks` SET `status`='tungee_queued', "
        f"`company_name`={sql_literal(target)}, `result_json`={sql_literal(json.dumps(combined, ensure_ascii=False))}, "
        f"`error_text`='' WHERE `id`={int(task_id)} LIMIT 1; "
        f"UPDATE `investigation_intake` SET `status`='tungee_queued', `error_text`='' WHERE `id`={int(rows[0]['intakeId'])} LIMIT 1",
        config,
    )
    start_auto_dispatch()


def submit_manual_doubao_result(body: dict) -> dict:
    """Fallback for a Doubao web session that accepts input but returns no reply."""
    task_id = str(body.get("taskId") or "").strip()
    if not task_id.isdigit():
        raise ApiError("任务 ID 无效。")
    raw = body.get("result")
    if isinstance(raw, dict):
        result = raw
    else:
        text = str(raw or "").strip()
        if not text:
            raise ApiError("请粘贴豆包返回的 JSON。")
        try:
            from llm_doubao_web import extract_json
            result = extract_json(text)
        except Exception:
            result = None
        if not isinstance(result, dict):
            try:
                result = json.loads(text)
            except Exception as exc:
                raise ApiError("豆包结果不是可解析的 JSON，请保留完整大括号内容后重试。") from exc
    target = suggested_tungee_company(result, str(body.get("tungeeTarget") or "").strip())
    if not target:
        raise ApiError("结果中没有识别到控股公司，请补充“建议探迹查询控股公司”或“实际投资方”。")
    finish_doubao_stage(task_id, result, [], target)
    start_auto_dispatch()
    return {"ok": True, "taskId": task_id, "tungeeTarget": target, **fetch_investigation_center()}


def fail_research_task(task_id: str, error: str) -> None:
    config = load_config()
    intake = json_rows(
        "SELECT JSON_OBJECT('intakeId',IFNULL(CAST(i.`id` AS CHAR),'')) "
        "FROM `investigation_tasks` t LEFT JOIN `investigation_intake` i ON i.`task_id`=t.`id` "
        f"WHERE t.`id`={int(task_id)} LIMIT 1",
        config,
    )
    run_mysql(
        "UPDATE `investigation_tasks` SET `status`='research_failed', "
        f"`error_text`={sql_literal(str(error)[:2000])}, `finished_at`=NOW() WHERE `id`={int(task_id)} LIMIT 1; "
        + (f"UPDATE `investigation_intake` SET `status`='research_failed', `error_text`={sql_literal(str(error)[:2000])} WHERE `id`={int(intake[0]['intakeId'])} LIMIT 1" if intake and intake[0].get("intakeId") else ""),
        config,
    )


def fallback_subject_result(project: dict) -> dict:
    """Keep the pipeline moving when Doubao web returns no response."""
    project_name = valid_public_value(project.get("projectName"))
    location = valid_public_value(project.get("projectLocation")) or infer_project_location(project_name)
    situation = valid_public_value(project.get("projectSituation"))
    if not situation and project_name:
        situation = f"项目已完成备案，建设内容以“{project_name}”备案信息为准，当前施工进度待电话核实。"
    return {
        "最终控股公司": "[未确认]",
        "建议探迹查询控股公司": "[未确认]",
        "控股人姓名": "[未确认]",
        "控股人手机号": "[未公开]",
        "项目建设地点": location or "[未确认]",
        "项目情况": situation or "已完成备案，当前建设进度待电话核实。",
        "股权穿透": {"最终控股方": "", "实际控制人": ""},
        "关键对接人": [],
        "调查说明": "豆包网页本次无有效响应。备案公司已保留为调查起点，未确认实际主体，因此不进入探迹。",
    }


def hydrate_intake_project(project: dict, config: dict) -> dict:
    """Restore the filing company when an older intake row lost projectCompany."""
    if not isinstance(project, dict) or project.get("projectCompany"):
        return project
    source_id = str(project.get("sourceId") or "").strip()
    if not source_id:
        return project
    try:
        rows = json_rows(
            "SELECT JSON_OBJECT('projectUnit',`project_unit`,'projectName',`project_name`) "
            f"FROM `filing_projects` WHERE `source_project_id`={sql_literal(source_id)} LIMIT 1",
            config,
        )
        if rows:
            project["projectCompany"] = rows[0].get("projectUnit", "")
            if not project.get("projectName"):
                project["projectName"] = rows[0].get("projectName", "")
    except Exception:
        pass
    return project


def fetch_task_project(task_id: str, config: dict) -> tuple[str, dict]:
    rows = json_rows(
        "SELECT JSON_OBJECT('leadId',IFNULL(CAST(t.`lead_id` AS CHAR),''),'intakeId',IFNULL(CAST(i.`id` AS CHAR),''),'projectJson',IFNULL(CAST(i.`project_json` AS CHAR),'{}')) "
        "FROM `investigation_tasks` t LEFT JOIN `investigation_intake` i ON i.`task_id`=t.`id` "
        f"WHERE t.`id`={int(task_id)} LIMIT 1",
        config,
    )
    if not rows:
        raise ApiError("调查任务不存在。")
    if rows[0].get("leadId"):
        return rows[0]["leadId"], fetch_lead_for_research(rows[0]["leadId"])
    try:
        return "", hydrate_intake_project(json.loads(rows[0].get("projectJson") or "{}"), config)
    except Exception as exc:
        raise ApiError("待调查项目原始数据损坏。") from exc


def release_research_driver_and_dispatch(driver) -> None:
    if driver is not None:
        try:
            driver.quit()
        except Exception:
            pass
    start_auto_dispatch()


def run_doubao_research_task(task_id: str) -> None:
    driver = None
    with DOUBAO_LOCK:
        try:
            config = load_config()
            lead_id, project = fetch_task_project(task_id, config)
            if str(MODULE_CODE_DIR) not in sys.path:
                sys.path.insert(0, str(MODULE_CODE_DIR))
            from llm_doubao_web import DOUBAO_DEBUG_PORT, DOUBAO_URL, attach_driver, call_doubao
            from llm_summary import SYSTEM_PROMPT, build_user_prompt

            profile_dir = doubao_profile_dir()
            profile_dir.mkdir(parents=True, exist_ok=True)
            driver = attach_driver(DOUBAO_DEBUG_PORT, profile_dir)
            # 豆包公开对话可匿名使用，不把“未登录”作为硬失败条件。
            if "doubao.com" not in driver.current_url:
                driver.get(DOUBAO_URL)
                time.sleep(2)
            # The primary handoff is Doubao -> Tungee. Do not block the
            # subject-company decision on optional public-search enrichment.
            baidu_clues = []
            prompt = (
                f"[系统设定] {SYSTEM_PROMPT}\n\n{build_user_prompt(project_to_doubao_row(project))}"
            )
            result = call_doubao(driver, prompt, max_retry=2)
            if not result:
                result = fallback_subject_result(project)
            baidu_clues = fetch_baidu_public_clues(project)
            intake_rows = json_rows(
                f"SELECT JSON_OBJECT('intakeId',IFNULL(CAST(`id` AS CHAR),'')) FROM `investigation_intake` WHERE `task_id`={int(task_id)} LIMIT 1",
                config,
            )
            if intake_rows and intake_rows[0].get("intakeId"):
                # Keep the parsed response even when the subject-company
                # handoff needs manual confirmation.
                run_mysql(
                    f"UPDATE `investigation_tasks` SET `result_json`={sql_literal(json.dumps({'doubao': result, 'baidu': baidu_clues}, ensure_ascii=False))} WHERE `id`={int(task_id)} LIMIT 1",
                    config,
                )
                finish_doubao_stage(task_id, result, baidu_clues)
            else:
                finish_research_task(task_id, result, baidu_clues)
        except Exception as exc:
            fail_research_task(task_id, str(exc))
        finally:
            release_research_driver_and_dispatch(driver)


def run_next_research_task() -> dict:
    config = load_config()
    ensure_investigation_tables(config)
    task_id = run_mysql(
        "SELECT `id` FROM `investigation_tasks` WHERE `status` IN ('doubao_queued','tungee_done','research_failed') "
        "ORDER BY `priority`, `id` LIMIT 1",
        config,
    ).strip()
    if not task_id:
        raise ApiError("当前没有待综合调查任务。")
    if not start_research_task(task_id):
        raise ApiError("该任务当前不能启动综合调查。")
    return {"ok": True, "taskId": task_id, **fetch_investigation_center()}


def start_research_task(task_id: str) -> bool:
    config = load_config()
    with TASK_ASSIGN_LOCK:
        status = run_mysql(
            f"SELECT `status` FROM `investigation_tasks` WHERE `id`={int(task_id)} LIMIT 1",
            config,
        ).strip()
        if status not in {"doubao_queued", "tungee_done", "research_failed"}:
            return False
        active_research = run_mysql(
            "SELECT COUNT(*) FROM `investigation_tasks` "
            "WHERE `status` IN ('doubao_running','research_running')",
            config,
        ).strip()
        if active_research not in {"", "0"}:
            return False
        next_status = "doubao_running" if status == "doubao_queued" else "research_running"
        run_mysql(
            f"UPDATE `investigation_tasks` SET `status`={sql_literal(next_status)}, `error_text`='', `finished_at`=NULL "
            f"WHERE `id`={int(task_id)} LIMIT 1",
            config,
        )
        worker = threading.Thread(target=run_doubao_research_task, args=(task_id,), daemon=True)
        worker.start()
    return True


def assign_next_investigation_task() -> dict:
    config = load_config()
    ensure_investigation_tables(config)
    with TASK_ASSIGN_LOCK:
        seat_id = run_mysql(
            "SELECT `id` FROM `tungee_seats` WHERE `status`='active' AND `state`='idle' "
            "ORDER BY IFNULL(`last_used_at`,'1970-01-01'), `id` LIMIT 1",
            config,
        ).strip()
        if not seat_id:
            raise ApiError("当前没有空闲探迹席位。")
        task_id = run_mysql(
            "SELECT `id` FROM `investigation_tasks` WHERE `status` IN ('queued','tungee_queued') "
            "ORDER BY `priority`, `id` LIMIT 1",
            config,
        ).strip()
        if not task_id:
            raise ApiError("当前没有待分配任务。")
        run_mysql(
            "UPDATE `investigation_tasks` SET `status`='running', "
            f"`assigned_seat_id`={int(seat_id)}, `started_at`=NOW(), `error_text`='' WHERE `id`={int(task_id)} LIMIT 1; "
            "UPDATE `tungee_seats` SET `state`='busy', `today_count`=`today_count`+1, "
            f"`last_used_at`=NOW(), `last_error`='' WHERE `id`={int(seat_id)} LIMIT 1",
            config,
        )
        worker = threading.Thread(target=run_tungee_task, args=(task_id, seat_id), daemon=True)
        worker.start()
    return {"ok": True, "taskId": task_id, "seatId": seat_id, **fetch_investigation_center()}


def auto_dispatch_available_tasks() -> None:
    if not AUTO_DISPATCH_LOCK.acquire(blocking=False):
        return
    try:
        config = load_config()
        # Doubao uses one shared browser session, so only one analysis may run.
        active_doubao = run_mysql(
            "SELECT COUNT(*) FROM `investigation_tasks` WHERE `status` IN ('doubao_running','research_running')",
            config,
        ).strip()
        if active_doubao == "0":
            next_doubao = run_mysql(
                "SELECT `id` FROM `investigation_tasks` "
                "WHERE `status` IN ('doubao_queued','tungee_done','research_failed') "
                "ORDER BY `priority`,`id` LIMIT 1",
                config,
            ).strip()
            if next_doubao:
                start_research_task(next_doubao)

        while True:
            try:
                assign_next_investigation_task()
            except Exception:
                return
    finally:
        AUTO_DISPATCH_LOCK.release()


def start_auto_dispatch() -> None:
    threading.Thread(target=auto_dispatch_available_tasks, daemon=True).start()


def automatic_dispatch_loop() -> None:
    """Keep existing waiting tasks moving even when an event-driven wake-up is missed."""
    while True:
        try:
            auto_dispatch_available_tasks()
        except Exception:
            # A transient browser or database error must not permanently stop dispatch.
            pass
        time.sleep(AUTO_DISPATCH_INTERVAL_SECONDS)


def auto_collect_previous_day() -> dict:
    """Collect yesterday's finished filings and enqueue enrichment automatically."""
    date_text = (date.today() - timedelta(days=1)).isoformat()
    payload = fetch_filings(date_text, AUTO_FILING_KEYWORDS, False)
    rows = payload.get("rows", []) if isinstance(payload, dict) else []
    intake = submit_intake_projects(rows) if rows else {"queued": 0, "duplicates": 0}
    return {
        "date": date_text,
        "sourceMode": payload.get("readMode", "") if isinstance(payload, dict) else "",
        "projects": len(rows),
        "queued": intake.get("queued", 0),
        "duplicates": intake.get("duplicates", 0),
    }


def automatic_pipeline_loop() -> None:
    """Run once per day; retries every 15 minutes after a transient failure."""
    last_successful_date = ""
    while True:
        target_date = (date.today() - timedelta(days=1)).isoformat()
        if target_date != last_successful_date:
            try:
                auto_collect_previous_day()
                last_successful_date = target_date
            except Exception:
                # The next interval retries without affecting employee usage.
                pass
        try:
            backfill_shared_company_data()
            queue_incomplete_company_refreshes(limit=20)
        except Exception:
            # Missing-field refresh must not stop the daily filing collector.
            pass
        time.sleep(900)


def recover_investigation_state() -> None:
    config = load_config()
    ensure_investigation_tables(config)
    run_mysql(
            "UPDATE `investigation_tasks` SET `status`=IF(`lead_id` IS NULL,'tungee_queued','queued'), `assigned_seat_id`=NULL, `started_at`=NULL "
            "WHERE `status`='running'; "
            "UPDATE `investigation_tasks` SET `status`=IF(`lead_id` IS NULL,'doubao_queued','tungee_done') WHERE `status` IN ('research_running','doubao_running'); "
        "UPDATE `tungee_seats` SET `state`=IF(`status`='active','idle','offline') WHERE `state`='busy'",
        config,
    )


def fetch_projects() -> dict:
    config = load_config()
    query = config.get("query", {})
    table = query.get("table")
    if not table:
        raise ApiError("config.local.json 缺少 query.table。")
    ensure_sales_workflow_columns(config, table)

    limit = int(query.get("limit", 500))
    limit = max(1, min(limit, 5000))
    field_map = query.get("field_map", {})
    fields = EXPORT_FIELDS + META_FIELDS
    json_args = ["'dbId'", "CAST(`id` AS CHAR)"]
    for key in fields:
        json_args.append(f"'{key}'")
        json_args.append(value_expr(field_map, key))
    for key, column in WORKFLOW_COLUMNS.items():
        json_args.append(f"'{key}'")
        if key == "queued":
            json_args.append(f"IFNULL(CAST({quote_identifier(column)} AS UNSIGNED), 0)")
        else:
            json_args.append(f"IFNULL(CAST({quote_identifier(column)} AS CHAR), '')")

    sql = f"SELECT JSON_OBJECT({', '.join(json_args)}) FROM {quote_identifier(table)}"
    source_column = field_map.get("source")
    if source_column:
        sql += f" WHERE COALESCE({quote_identifier(source_column)}, '') <> 'Excel日期分组'"
    order_by = query.get("order_by")
    if order_by:
        order = str(query.get("order", "DESC")).upper()
        if order not in {"ASC", "DESC"}:
            order = "DESC"
        sql += f" ORDER BY {quote_identifier(order_by)} {order}"
    sql += f" LIMIT {limit}"

    output = run_mysql(sql, config)
    projects = []
    for line in output.splitlines():
        line = line.strip()
        if not line:
            continue
        item = json.loads(line)
        inferred_tungee = "已取信息" if any(item.get(key) for key in ("projectContact", "projectPhone", "mainCompany")) else "待查探迹"
        item["tungeeStatus"] = item.get("tungeeStatus") or inferred_tungee
        item["reviewStatus"] = item.get("reviewStatus") or ("待人工审核" if item["tungeeStatus"] == "已取信息" else "待查探迹")
        item["queued"] = bool(item.get("queued"))
        item["accuracy"] = assess_project_accuracy(item)
        projects.append(item)

    # Show raw filing leads while enrichment is still running. This keeps the
    # employee workspace useful before a task becomes a formal sales_leads row.
    pending_intake = json_rows(
        "SELECT JSON_SET(`project_json`, '$.dbId', '', '$.sourceId', `source_id`, "
        "'$.reviewStatus', CASE WHEN `status`='completed' THEN '待修复入库' "
        "WHEN `status`='research_failed' THEN '待人工确认' ELSE '待补全' END, "
        "'$.tungeeStatus', IF(`status`='completed','已取信息',`status`), '$.queued', TRUE, '$.intakeStatus', `status`) "
        "FROM `investigation_intake` WHERE `status` <> 'cancelled' ORDER BY `id` DESC LIMIT 5000",
        config,
    )
    existing_source_ids = {str(item.get("sourceId") or "") for item in projects if item.get("sourceId")}
    existing_record_codes = {str(item.get("recordCode") or "").strip() for item in projects if str(item.get("recordCode") or "").strip()}
    existing_project_pairs = {
        (str(item.get("projectName") or "").strip(), str(item.get("projectCompany") or "").strip())
        for item in projects
        if str(item.get("projectName") or "").strip() or str(item.get("projectCompany") or "").strip()
    }
    for item in pending_intake:
        if not isinstance(item, dict):
            continue
        source_id = str(item.get("sourceId") or "")
        record_code = str(item.get("recordCode") or "").strip()
        project_pair = (str(item.get("projectName") or "").strip(), str(item.get("projectCompany") or "").strip())
        if (
            source_id in existing_source_ids
            or (record_code and record_code in existing_record_codes)
            or project_pair in existing_project_pairs
        ):
            continue
        item["dbId"] = ""
        item["category"] = item.get("category") or infer_project_category(item)
        item["queued"] = True
        item["source"] = item.get("source") or FILING_SOURCE_NAME
        item["accuracy"] = assess_project_accuracy(item)
        projects.append(item)

    mysql = config.get("mysql", {})
    return {
        "ok": True,
        "source": "mysql",
        "database": mysql.get("database", ""),
        "table": table,
        "count": len(projects),
        "projects": projects,
    }


def fetch_available_dates(config: dict, table: str, date_column: str) -> list[str]:
    sql = (
        f"SELECT DISTINCT LEFT(CAST({quote_identifier(date_column)} AS CHAR), 10) "
        f"FROM {quote_identifier(table)} "
        f"WHERE {quote_identifier(date_column)} IS NOT NULL AND CAST({quote_identifier(date_column)} AS CHAR) <> '' "
        f"ORDER BY 1 DESC LIMIT 5"
    )
    return [line.strip() for line in run_mysql(sql, config).splitlines() if line.strip()]


def filter_filing_rows(rows: list[dict], keywords: list[str]) -> list[dict]:
    if not keywords:
        return rows
    return [
        row for row in rows
        if any(keyword in filing_search_text(row) for keyword in keywords)
    ]


def deduplicate_filing_rows(rows: list[dict]) -> list[dict]:
    """Collapse imported copies of official filings while preserving stable order."""
    result: list[dict] = []
    positions: dict[tuple[str, ...], int] = {}
    for row in rows:
        code = str(row.get("projectCode") or "").strip()
        name = str(row.get("projectName") or "").strip()
        company = str(row.get("projectUnit") or "").strip()
        key = ("code", code) if code else ("pair", name, company)
        if not any(key[1:]):
            result.append(row)
            continue
        if key not in positions:
            positions[key] = len(result)
            result.append(row)
            continue
        current_index = positions[key]
        current = result[current_index]
        current_official = str(current.get("source") or "") == FILING_SOURCE_NAME
        incoming_official = str(row.get("source") or "") == FILING_SOURCE_NAME
        if incoming_official and not current_official:
            result[current_index] = row
    return result


def fetch_filings(date_text: str, keyword_text: str, force_refresh: bool = False) -> dict:
    date_text = date_text.strip()
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_text):
        raise ApiError("立项日期格式必须是 YYYY-MM-DD。")

    keywords = split_keywords(keyword_text)
    today_text = time.strftime("%Y-%m-%d")
    is_today = date_text == today_text
    # 当天数据可能在下午继续发布，因此当天始终实时采集；历史日期优先读库。
    if not force_refresh and not is_today:
        try:
            cache_payload = fetch_filing_cache(date_text, keywords)
            if cache_payload["hasDate"]:
                return {
                    "ok": True,
                    "source": "sz_investment_publicity",
                    "sourceLabel": FILING_SOURCE_NAME,
                    "sourceUrl": FILING_SOURCE_URL,
                    "readMode": "filing_database",
                    "date": date_text,
                    "keywords": keywords,
                    "availableDates": cache_payload["availableDates"],
                    "count": len(cache_payload["rows"]),
                    "totalCount": cache_payload["totalCount"],
                    "rows": cache_payload["rows"],
                    "allRows": cache_payload.get("allRows", cache_payload["rows"]),
                }
        except Exception:
            pass
    try:
        live_rows, crawler_logs = crawl_filings_live(date_text, [])
        cache_write_error = ""
        try:
            store_filing_rows(live_rows, keyword_text)
        except Exception as cache_exc:
            cache_write_error = str(cache_exc)
        filtered_rows = filter_filing_rows(live_rows, keywords)
        return {
            "ok": True,
            "source": "sz_investment_publicity",
            "sourceLabel": FILING_SOURCE_NAME,
            "sourceUrl": FILING_SOURCE_URL,
            "readMode": "live_crawler",
            "date": date_text,
            "keywords": keywords,
            "availableDates": [],
            "count": len(filtered_rows),
            "totalCount": len(live_rows),
            "rows": filtered_rows,
            "allRows": live_rows,
            "crawlerLogs": crawler_logs,
            "cacheWriteError": cache_write_error,
        }
    except Exception as crawler_exc:
        crawler_error = str(crawler_exc)

    try:
        cache_payload = fetch_filing_cache(date_text, keywords)
        if cache_payload["rows"] or cache_payload["hasDate"]:
            return {
                "ok": True,
                "source": "sz_investment_publicity",
                "sourceLabel": FILING_SOURCE_NAME,
                "sourceUrl": FILING_SOURCE_URL,
                "readMode": "filing_database_fallback",
                "crawlerError": crawler_error,
                "date": date_text,
                "keywords": keywords,
                "availableDates": cache_payload["availableDates"],
                "count": len(cache_payload["rows"]),
                "totalCount": cache_payload["totalCount"],
                "rows": cache_payload["rows"],
                "allRows": cache_payload.get("allRows", cache_payload["rows"]),
            }
    except Exception:
        pass

    config = load_config()
    query = config.get("query", {})
    table = query.get("table")
    if not table:
        raise ApiError("config.local.json 缺少 query.table。")

    field_map = query.get("field_map", {})
    date_column = field_map.get("recordDate")
    if not date_column:
        raise ApiError("config.local.json 缺少 recordDate 字段映射。")

    filing_field_map = {target: field_map.get(source) for target, source in FILING_FIELDS}
    json_args = []
    for target, _source in FILING_FIELDS:
        json_args.append(f"'{target}'")
        json_args.append(value_expr(filing_field_map, target))

    limit = int(query.get("filing_limit", 1000))
    limit = max(1, min(limit, 5000))
    date_filter = f"LEFT(CAST({quote_identifier(date_column)} AS CHAR), 10) = {sql_literal(date_text)}"
    sql = f"SELECT JSON_OBJECT({', '.join(json_args)}) FROM {quote_identifier(table)} WHERE {date_filter}"
    order_by = query.get("order_by")
    if order_by:
        order = str(query.get("order", "DESC")).upper()
        if order not in {"ASC", "DESC"}:
            order = "DESC"
        sql += f" ORDER BY {quote_identifier(order_by)} {order}"
    sql += f" LIMIT {limit}"

    all_rows = []
    for line in run_mysql(sql, config).splitlines():
        line = line.strip()
        if not line:
            continue
        item = json.loads(line)
        item["source"] = FILING_SOURCE_NAME
        item["sourceId"] = item.get("sourceId") or item.get("projectCode") or item.get("projectName")
        item["sourceUrl"] = FILING_SOURCE_URL
        all_rows.append(item)

    rows = filter_filing_rows(all_rows, keywords)

    if crawler_error and not all_rows:
        raise ApiError(f"实时备案采集失败，且本地没有 {date_text} 的可用数据：{crawler_error}")

    mysql = config.get("mysql", {})
    return {
        "ok": True,
        "source": "sz_investment_publicity",
        "sourceLabel": FILING_SOURCE_NAME,
        "sourceUrl": FILING_SOURCE_URL,
        "readMode": "local_cache",
        "crawlerError": crawler_error,
        "database": mysql.get("database", ""),
        "table": table,
        "date": date_text,
        "keywords": keywords,
        "availableDates": fetch_available_dates(config, table, date_column),
        "count": len(rows),
        "rows": rows,
        "totalCount": len(all_rows),
        "allRows": all_rows,
    }


def create_filing_job(
    date_text: str,
    keyword_text: str,
    force_refresh: bool = False,
    requested_by: str = "",
) -> dict:
    date_text = date_text.strip()
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_text):
        raise ApiError("立项日期格式必须是 YYYY-MM-DD。")

    keywords = split_keywords(keyword_text)
    request_key = (date_text, tuple(sorted(keywords)), bool(force_refresh))
    with JOB_LOCK:
        for existing in FILING_JOBS.values():
            existing_key = (
                existing.get("date", ""),
                tuple(sorted(existing.get("keywords", []))),
                bool(existing.get("forceRefresh")),
            )
            if existing_key == request_key and existing.get("status") in {"queued", "running"}:
                payload = public_job(existing)
                payload["shared"] = True
                payload["message"] = "相同查询正在处理中，已共享现有任务"
                return payload

    job_id = uuid.uuid4().hex[:12]
    job = {
        "id": job_id,
        "type": "filing_crawl",
        "status": "queued",
        "date": date_text,
        "keywords": keywords,
        "keywordText": keyword_text,
        "forceRefresh": force_refresh,
        "requestedBy": requested_by,
        "sourceLabel": FILING_SOURCE_NAME,
        "sourceUrl": FILING_SOURCE_URL,
        "readMode": "",
        "message": "已提交到服务器采集队列",
        "createdAt": now_text(),
        "updatedAt": now_text(),
        "finishedAt": "",
        "count": 0,
        "rows": [],
        "allRows": [],
        "crawlerLogs": [],
        "error": "",
    }
    with JOB_LOCK:
        FILING_JOBS[job_id] = job
    persist_filing_job(job)

    thread = threading.Thread(target=run_filing_job, args=(job_id,), daemon=True)
    thread.start()
    return public_job(job)


def public_job(job: dict) -> dict:
    payload = dict(job)
    payload["ok"] = True
    return payload


def update_job(job_id: str, **changes) -> None:
    with JOB_LOCK:
        job = FILING_JOBS.get(job_id)
        if not job:
            return
        job.update(changes)
        job["updatedAt"] = now_text()
        snapshot = dict(job)
    persist_filing_job(snapshot)


def get_filing_job(job_id: str) -> dict:
    with JOB_LOCK:
        job = FILING_JOBS.get(job_id)
        if job:
            return public_job(job)
    job = filing_job_from_db(job_id)
    if not job:
        raise ApiError("采集任务不存在。")
    return public_job(job)


def run_filing_job(job_id: str) -> None:
    with JOB_LOCK:
        job = FILING_JOBS.get(job_id)
        if not job:
            return
        date_text = job["date"]
        keyword_text = job["keywordText"]
        force_refresh = bool(job.get("forceRefresh"))
    update_job(
        job_id,
        status="running",
        startedAt=now_text(),
        message=(
            "正在实时采集今天的深圳投资项目公示"
            if force_refresh or date_text == time.strftime("%Y-%m-%d")
            else "正在读取采投标数据库"
        ),
    )
    try:
        payload = fetch_filings(date_text, keyword_text, force_refresh)
        update_job(
            job_id,
            status="done",
            message="采集完成",
            readMode=payload.get("readMode", ""),
            count=payload.get("count", 0),
            rows=payload.get("rows", []),
            allRows=payload.get("allRows", payload.get("rows", [])),
            crawlerLogs=payload.get("crawlerLogs", []),
            crawlerError=payload.get("crawlerError", ""),
            cacheWriteError=payload.get("cacheWriteError", ""),
            availableDates=payload.get("availableDates", []),
            totalCount=payload.get("totalCount", payload.get("count", 0)),
            finishedAt=now_text(),
        )
    except Exception as exc:
        update_job(job_id, status="failed", message="采集失败", error=str(exc), finishedAt=now_text())


def fetch_schema() -> dict:
    config = load_config()
    query = config.get("query", {})
    table = query.get("table")
    if not table:
        raise ApiError("config.local.json 缺少 query.table。")
    rows = run_mysql(f"SHOW COLUMNS FROM {quote_identifier(table)}", config).splitlines()
    columns = []
    for row in rows:
        parts = row.split("\t")
        if parts:
            columns.append(parts[0])
    return {"ok": True, "table": table, "columns": columns}


def infer_project_category(project: dict) -> str:
    text = " ".join(str(project.get(key, "")) for key in ("projectName", "projectCompany", "investor", "mainCompany"))
    if re.search(r"储能|光储充|kwh|mwh|电池", text, re.IGNORECASE):
        return "storage_charge"
    if re.search(r"车棚|停车场|充电站|超充|充电桩", text, re.IGNORECASE):
        return "carport"
    if re.search(r"国企|华润|中建|中铁|城建|招商|深能|国电|华能|南网|国家电投", text):
        return "roof_state"
    return "roof_private"


def project_nature_label(project: dict) -> str:
    labels = {
        "roof_state": "屋顶光伏（国企）",
        "storage_charge": "光储充",
        "carport": "光伏车棚",
        "roof_private": "屋顶光伏（民企）",
    }
    category = project.get("category") or infer_project_category(project)
    return labels.get(category, str(category or "待分类"))


def fetch_original_customers_for_export(date_filter: str = "") -> list[dict]:
    config = load_config()
    ensure_filing_cache_table(config)
    where = ""
    if date_filter == "__undated__":
        where = " WHERE `record_date`=''"
    elif date_filter:
        where = f" WHERE LEFT(CAST(`record_date` AS CHAR),10)={sql_literal(date_filter)}"
    return json_rows(
        "SELECT JSON_OBJECT('projectName',`project_name`,'projectCompany',`project_unit`,"
        "'recordCode',`project_code`,'recordDate',`record_date`) FROM `filing_projects`"
        f"{where} ORDER BY `record_date` DESC,`id` DESC LIMIT 5000",
        config,
    )


def build_export_workbook(date_filter: str = "") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    projects = fetch_projects()["projects"]
    if date_filter == "__undated__":
        projects = [project for project in projects if not str(project.get("recordDate") or "").strip()]
    elif date_filter:
        projects = [
            project for project in projects
            if str(project.get("recordDate") or "")[:10] == date_filter
        ]
    projects.sort(
        key=lambda project: (
            str(project.get("recordDate") or ""),
            str(project.get("recordCode") or ""),
            str(project.get("projectName") or ""),
        ),
        reverse=True,
    )
    originals = fetch_original_customers_for_export(date_filter)
    met_projects = [
        project for project in projects
        if re.search(r"见面|拜访|面谈|到访", str(project.get("phoneFeedback") or ""))
    ]

    workbook = Workbook()
    workbook.remove(workbook.active)

    def style_sheet(sheet, widths: list[int]) -> None:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(widths))}{max(1, sheet.max_row)}"
        sheet.sheet_view.showGridLines = False
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="176B61")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.row_dimensions[1].height = 30

    original_sheet = workbook.create_sheet("原始客户")
    original_sheet.append(["项目信息", "备案公司", "项目备案编号"])
    for row in originals:
        original_sheet.append([row.get("projectName", ""), row.get("projectCompany", ""), row.get("recordCode", "")])
    style_sheet(original_sheet, [50, 38, 26])
    for cell in original_sheet["C"]:
        cell.number_format = "@"

    target_sheet = workbook.create_sheet("目标客户")
    target_sheet.append(SALES_RESULT_LABELS)
    for project in projects:
        row = dict(project)
        row["projectNature"] = project_nature_label(project)
        target_sheet.append([row.get(key, "") for key in SALES_RESULT_FIELDS])
    style_sheet(target_sheet, [48, 32, 20, 18, 22, 32, 20, 22, 38, 48, 20, 48, 48, 28, 26, 32, 40, 40])
    for column in ("E", "H", "O"):
        for cell in target_sheet[column]:
            cell.number_format = "@"

    met_sheet = workbook.create_sheet("已见面客户")
    met_fields = [
        "projectName", "projectCompany", "projectContact", "projectPhone", "mainCompany",
        "projectSituation", "projectProgress", "projectLocation", "projectSummary", "phoneFeedback",
    ]
    met_sheet.append(["项目信息", "项目单位", "项目公司联系人", "手机号", "控股公司", "项目情况", "项目进展", "项目所在地", "项目总结", "拜访反馈"])
    for project in met_projects:
        met_sheet.append([project.get(key, "") for key in met_fields])
    style_sheet(met_sheet, [48, 32, 18, 22, 32, 48, 48, 28, 40, 48])
    for cell in met_sheet["D"]:
        cell.number_format = "@"

    performance_sheet = workbook.create_sheet("业绩进度表")
    contacted = sum(bool(project.get("projectPhone") or project.get("mainPhone")) for project in projects)
    feedback = sum(bool(project.get("phoneFeedback")) for project in projects)
    performance_sheet.append(["指标", "数量"])
    for label, value in (
        ("原始客户", len(originals)),
        ("目标客户", len(projects)),
        ("已获取联系方式", contacted),
        ("已电话反馈", feedback),
        ("已见面客户", len(met_projects)),
    ):
        performance_sheet.append([label, value])
    style_sheet(performance_sheet, [28, 16])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(BASE_DIR), **kwargs)

    def send_json(self, status: int, payload: dict, extra_headers: dict | None = None) -> None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        for name, value in (extra_headers or {}).items():
            self.send_header(name, value)
        self.end_headers()
        self.wfile.write(data)

    def end_headers(self) -> None:
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", "SAMEORIGIN")
        self.send_header("Referrer-Policy", "same-origin")
        # Employees must not keep an old frontend after ERP workflow updates.
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        super().end_headers()

    def read_json_body(self) -> dict:
        length = int(self.headers.get("Content-Length") or 0)
        if length <= 0:
            return {}
        raw = self.rfile.read(length).decode("utf-8", errors="replace")
        if not raw.strip():
            return {}
        return json.loads(raw)

    def send_bytes(self, status: int, data: bytes, content_type: str, filename: str) -> None:
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(filename)}")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        if path == "/api/session":
            required = bool(configured_access_code() or configured_users())
            user = session_user(self.headers)
            self.send_json(200, {"ok": True, "authRequired": required, "authenticated": bool(user), "user": user or {}})
            return
        if path == "/api/health":
            self.send_json(200, {"ok": True, "app": "鑫众ERP系统"})
            return
        if path.startswith("/api/") and not session_is_valid(self.headers):
            self.send_json(401, {"ok": False, "error": "请先登录鑫众ERP系统。"})
            return
        if path == "/api/export.xlsx":
            try:
                params = parse_qs(parsed.query)
                date_filter = params.get("date", [""])[0].strip()
                if date_filter and date_filter != "__undated__" and not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_filter):
                    raise ApiError("导出日期格式必须是 YYYY-MM-DD。")
                filename_date = "未标日期" if date_filter == "__undated__" else date_filter or "全部日期"
                filename = f"鑫众ERP系统_{filename_date}.xlsx"
                self.send_bytes(
                    200,
                    build_export_workbook(date_filter),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    filename,
                )
            except Exception as exc:
                self.send_json(500, {"ok": False, "error": str(exc)})
            return
        if path == "/api/projects":
            try:
                self.send_json(200, fetch_projects())
            except Exception as exc:
                self.send_json(500, {"ok": False, "error": str(exc)})
            return
        if path == "/api/filings":
            params = parse_qs(parsed.query)
            try:
                self.send_json(200, fetch_filings(params.get("date", [""])[0], params.get("keywords", [""])[0]))
            except Exception as exc:
                self.send_json(500, {"ok": False, "error": str(exc)})
            return
        if path == "/api/filing-jobs":
            params = parse_qs(parsed.query)
            try:
                if params.get("recent", [""])[0]:
                    self.send_json(200, recent_filing_jobs(params.get("limit", ["20"])[0]))
                else:
                    self.send_json(200, get_filing_job(params.get("id", [""])[0]))
            except Exception as exc:
                self.send_json(404, {"ok": False, "error": str(exc)})
            return
        if path == "/api/schema":
            try:
                self.send_json(200, fetch_schema())
            except Exception as exc:
                self.send_json(500, {"ok": False, "error": str(exc)})
            return
        if path == "/api/investigation-center":
            try:
                if not user_can_access(session_user(self.headers) or {}, "sales"):
                    raise ApiError("当前账号无权访问电销调查中心。")
                self.send_json(200, fetch_investigation_center())
            except Exception as exc:
                self.send_json(500, {"ok": False, "error": str(exc)})
            return
        if path not in PUBLIC_STATIC_PATHS:
            self.send_error(404, "Not Found")
            return
        super().do_GET()

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        if path == "/api/login":
            try:
                access_code = configured_access_code()
                body = self.read_json_body()
                username = str(body.get("username") or "").strip()
                submitted = str(body.get("password") or body.get("accessCode") or "")
                matched_user = None
                for item in configured_users():
                    if hmac.compare_digest(username, str(item.get("username") or "")) and hmac.compare_digest(submitted, str(item.get("password") or "")):
                        matched_user = {
                            "username": username,
                            "name": str(item.get("name") or username),
                            "role": str(item.get("role") or "employee"),
                            "department": str(item.get("department") or "sales"),
                        }
                        break
                if matched_user is None and access_code and hmac.compare_digest(submitted, access_code):
                    matched_user = {"username": username or "admin", "name": "系统管理员", "role": "admin", "department": "all"}
                if matched_user is None:
                    raise ApiError("访问码错误。")
                token = secrets.token_urlsafe(32)
                with SESSION_LOCK:
                    SESSIONS[token] = {"expiresAt": time.time() + 12 * 60 * 60, "user": matched_user}
                self.send_json(
                    200,
                    {"ok": True, "user": matched_user},
                    {"Set-Cookie": f"xinzhong_session={token}; HttpOnly; SameSite=Strict; Path=/; Max-Age=43200"},
                )
            except Exception as exc:
                self.send_json(401, {"ok": False, "error": str(exc)})
            return
        if path == "/api/logout":
            token = session_token_from_headers(self.headers)
            with SESSION_LOCK:
                SESSIONS.pop(token, None)
            self.send_json(
                200,
                {"ok": True},
                {"Set-Cookie": "xinzhong_session=; HttpOnly; SameSite=Strict; Path=/; Max-Age=0"},
            )
            return
        if path.startswith("/api/") and not session_is_valid(self.headers):
            self.send_json(401, {"ok": False, "error": "请先登录鑫众ERP系统。"})
            return
        user = session_user(self.headers) or {}
        admin_paths = {
            "/api/tungee-seats", "/api/tungee-seats/toggle", "/api/tungee-seats/open-login",
            "/api/research/open-doubao-login", "/api/investigation-tasks/assign-next",
            "/api/investigation-tasks/run-research-next", "/api/investigation-tasks/manual-doubao-result",
            "/api/investigation-tasks/retry",
        }
        sales_paths = {"/api/filing-jobs", "/api/leads", "/api/leads/update", "/api/investigation-tasks"}
        if path in admin_paths and str(user.get("role")) != "admin":
            self.send_json(403, {"ok": False, "error": "该操作仅管理员可用。"})
            return
        if path in sales_paths and not user_can_access(user, "sales"):
            self.send_json(403, {"ok": False, "error": "当前账号无权操作电销数据。"})
            return
        if path == "/api/filing-jobs":
            try:
                body = self.read_json_body()
                self.send_json(
                    202,
                    create_filing_job(
                        str(body.get("date", "")),
                        str(body.get("keywords", "")),
                        bool(body.get("forceRefresh")),
                        str(body.get("requestedBy", "")),
                    ),
                )
            except Exception as exc:
                self.send_json(400, {"ok": False, "error": str(exc)})
            return
        if path == "/api/leads":
            try:
                body = self.read_json_body()
                projects = body.get("projects", [])
                if not isinstance(projects, list):
                    raise ApiError("projects 必须是数组。")
                self.send_json(200, submit_intake_projects(projects))
            except Exception as exc:
                self.send_json(400, {"ok": False, "error": str(exc)})
            return
        if path == "/api/leads/update":
            try:
                body = self.read_json_body()
                project = body.get("project", {})
                if not isinstance(project, dict):
                    raise ApiError("project 必须是对象。")
                self.send_json(200, update_lead(project))
            except Exception as exc:
                self.send_json(400, {"ok": False, "error": str(exc)})
            return
        if path == "/api/investigation-tasks":
            try:
                self.send_json(200, create_investigation_task(self.read_json_body()))
            except Exception as exc:
                self.send_json(400, {"ok": False, "error": str(exc)})
            return
        if path == "/api/investigation-tasks/assign-next":
            try:
                self.send_json(200, assign_next_investigation_task())
            except Exception as exc:
                self.send_json(400, {"ok": False, "error": str(exc)})
            return
        if path == "/api/investigation-tasks/run-research-next":
            try:
                self.send_json(200, run_next_research_task())
            except Exception as exc:
                self.send_json(400, {"ok": False, "error": str(exc)})
            return
        if path == "/api/investigation-tasks/retry":
            try:
                self.send_json(200, retry_investigation_task(self.read_json_body()))
            except Exception as exc:
                self.send_json(400, {"ok": False, "error": str(exc)})
            return
        if path == "/api/investigation-tasks/manual-doubao-result":
            try:
                self.send_json(200, submit_manual_doubao_result(self.read_json_body()))
            except Exception as exc:
                self.send_json(400, {"ok": False, "error": str(exc)})
            return
        if path == "/api/tungee-seats":
            try:
                self.send_json(200, create_tungee_seat(self.read_json_body()))
            except Exception as exc:
                self.send_json(400, {"ok": False, "error": str(exc)})
            return
        if path == "/api/tungee-seats/toggle":
            try:
                self.send_json(200, toggle_tungee_seat(self.read_json_body()))
            except Exception as exc:
                self.send_json(400, {"ok": False, "error": str(exc)})
            return
        if path == "/api/tungee-seats/open-login":
            try:
                self.send_json(200, open_tungee_seat_login(self.read_json_body()))
            except Exception as exc:
                self.send_json(400, {"ok": False, "error": str(exc)})
            return
        if path == "/api/research/open-doubao-login":
            try:
                self.send_json(200, open_doubao_login())
            except Exception as exc:
                self.send_json(400, {"ok": False, "error": str(exc)})
            return
        self.send_json(404, {"ok": False, "error": "接口不存在"})


def main() -> None:
    port = int(os.environ.get("XINZHONG_ERP_PORT", "8765"))
    host = os.environ.get("XINZHONG_ERP_HOST", "0.0.0.0")
    recover_filing_jobs()
    recover_investigation_state()
    normalize_queued_investigation_targets()
    archive_resolved_historical_failures()
    backfill_shared_company_data()
    threading.Thread(target=automatic_pipeline_loop, daemon=True).start()
    threading.Thread(target=automatic_dispatch_loop, daemon=True).start()
    server = ThreadingHTTPServer((host, port), Handler)
    print(f"鑫众ERP系统 running at http://{host}:{port}")
    server.serve_forever()


if __name__ == "__main__":
    main()
