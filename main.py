# -*- coding: utf-8 -*-
"""
新能源项目智能爬虫工具 v2.0
按日期抓取深圳投资网 + 探迹，输出Excel
"""
import os
import sys
import json
import time
import queue
import threading
import traceback
import re
import sqlite3
import subprocess
import random
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

# GUI
import tkinter as tk
from tkinter import (
    ttk, messagebox, filedialog, END, DISABLED, NORMAL,
    StringVar, IntVar, BooleanVar
)
from tkinter.scrolledtext import ScrolledText

# 爬虫依赖
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from selenium import webdriver
from selenium.webdriver import ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import (
    NoSuchElementException, TimeoutException, WebDriverException,
    StaleElementReferenceException
)

# ============================================================
# 配置
# ============================================================
ROOT_DIR = Path(__file__).resolve().parent
CONFIG_FILE = ROOT_DIR / "config.json"
DB_FILE = ROOT_DIR / "projects.db"

TUNGEE_LOGIN_URL = "https://user.tungee.com/users/sign-in"
TUNGEE_SEARCH_URL = "https://bidding.tungee.com/search-enterprise/home"
TUNGEE_FILTER_URL = "https://bidding.tungee.com/customer-seeking/advanced-filter/enterprise"
SZ_BASE_URL = "https://wsbs.sz.gov.cn"

EXCEL_HEADERS = [
    "项目信息", "项目单位", "项目公司联系人", "手机号", "地址",
    "主体公司", "主体公司联系人", "手机号", "主体公司地址",
    "关系图谱", "投资人", "项目情况", "项目进展", "项目所在地",
    "项目备案编号", "备注", "项目总结"
]

def load_config() -> dict:
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_config(cfg: dict) -> None:
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

DEFAULT_CONFIG = {
    "tungee_phone": os.environ.get("TUNGEE_PHONE", ""),
    "tungee_password": os.environ.get("TUNGEE_PASSWORD", ""),
    "export_dir": str(Path.home() / "Desktop"),
    "request_interval": "3",
    "chrome_path": r"C:\Program Files\Google\Chrome\Application\chrome.exe",
    "debug_port": "9222",
    "headless": "0",
    "filter_keywords": "车棚 超充站 光储充 分布式光伏 屋顶分布式光伏 光伏EPC 充电站 储能 光伏发电 光伏电站",
}
APP_CONFIG = {**DEFAULT_CONFIG, **load_config()}

# ============================================================
# 工具函数
# ============================================================
def clean_text(text: str) -> str:
    if not text: return ''
    return re.sub(r'\s+', ' ', text).strip()

def extract_phone(text: str) -> str:
    if not text: return ''
    match = re.search(r'1[3-9]\d{9}', text)
    return match.group(0) if match else ''

def extract_sz_district(addr: str) -> str:
    for d in ['福田','罗湖','南山','宝安','龙岗','龙华','坪山','光明','盐田','大鹏','深汕']:
        if d in addr: return d
    return '深圳'

def sanitize_filename(s: str) -> str:
    s = clean_text(s)
    s = re.sub(r'[\\/:*?"<>|]+', '_', s)
    return s.strip(' ._')[:80] or 'file'

# ============================================================
# 数据库
# ============================================================
def init_db():
    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    # WAL 模式：支持并发读写，避免 "database is locked"
    c.execute("PRAGMA journal_mode=WAL")
    c.execute("""
        CREATE TABLE IF NOT EXISTS companies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_name TEXT NOT NULL,
            project_unit TEXT DEFAULT '',
            contact_person TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            address TEXT DEFAULT '',
            main_company TEXT DEFAULT '',
            main_contact TEXT DEFAULT '',
            main_phone TEXT DEFAULT '',
            main_address TEXT DEFAULT '',
            investor TEXT DEFAULT '',
            project_desc TEXT DEFAULT '',
            project_status TEXT DEFAULT '',
            location TEXT DEFAULT '',
            credit_code TEXT DEFAULT '',
            relation_graph TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            summary TEXT DEFAULT '',
            source TEXT DEFAULT '',
            record_date TEXT DEFAULT '',
            created_at TEXT DEFAULT '',
            UNIQUE(company_name, source, record_date)
        )
    """)
    # 迁移旧数据库：添加新增列（如果不存在）
    for col, default in [
        ('project_unit', ''),
        ('relation_graph', ''),
        ('remark', ''),
        ('summary', ''),
        ('phone_feedback', ''),
    ]:
        try:
            c.execute(f"ALTER TABLE companies ADD COLUMN {col} TEXT DEFAULT '{default}'")
        except sqlite3.OperationalError:
            pass  # 列已存在
    c.execute("""
        CREATE TABLE IF NOT EXISTS scrape_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT NOT NULL,
            record_date TEXT NOT NULL,
            started_at TEXT DEFAULT '',
            finished_at TEXT DEFAULT '',
            total_count INTEGER DEFAULT 0,
            status TEXT DEFAULT 'running',
            error_msg TEXT DEFAULT ''
        )
    """)
    conn.commit()
    conn.close()

def get_db_records(record_date: str = None, source: str = None) -> list:
    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    if record_date and source:
        c.execute("SELECT * FROM companies WHERE record_date=? AND source=? ORDER BY id DESC",
                  (record_date, source))
    elif record_date:
        c.execute("SELECT * FROM companies WHERE record_date=? ORDER BY id DESC", (record_date,))
    elif source:
        c.execute("SELECT * FROM companies WHERE source=? ORDER BY id DESC", (source,))
    else:
        c.execute("SELECT * FROM companies ORDER BY id DESC")
    rows = c.fetchall()
    conn.close()
    return rows

def insert_records(records: list):
    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    for r in records:
        c.execute("""
            INSERT OR REPLACE INTO companies
            (company_name, project_unit, contact_person, phone, address,
             main_company, main_contact, main_phone, main_address, investor,
             project_desc, project_status, location, credit_code,
             relation_graph, remark, summary,
             source, record_date, created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            r.get('company_name',''),
            r.get('project_unit',''),
            r.get('contact_person',''),
            r.get('phone',''),
            r.get('address',''),
            r.get('main_company',''),
            r.get('main_contact',''),
            r.get('main_phone',''),
            r.get('main_address',''),
            r.get('investor',''),
            r.get('project_desc',''),
            r.get('project_status',''),
            r.get('location',''),
            r.get('credit_code',''),
            r.get('relation_graph',''),
            r.get('remark',''),
            r.get('summary',''),
            r.get('source',''),
            r.get('record_date',''),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        ))
    conn.commit()
    conn.close()

def log_start(source: str, record_date: str) -> int:
    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    c.execute("INSERT INTO scrape_logs (source, record_date, started_at, status) VALUES (?,?,?,'running')",
              (source, record_date, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    log_id = c.lastrowid
    conn.commit()
    conn.close()
    return log_id

def log_finish(log_id: int, count: int, status: str, err: str = ''):
    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    c.execute("UPDATE scrape_logs SET finished_at=?, total_count=?, status=?, error_msg=? WHERE id=?",
              (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), count, status, err, log_id))
    conn.commit()
    conn.close()

# ============================================================
# 深圳投资网爬虫（Selenium + API 拦截）
# ============================================================
class SZInvestCrawler:
    """深圳投资网爬虫 - Selenium 驱动，拦截 AJAX API"""

    def __init__(self, log_queue: queue.Queue, interval: int = 3,
                 chrome_path: str = None, debug_port: int = 9222,
                 headless: bool = False, filter_keywords: list = None,
                 stop_event=None):
        self.log_queue = log_queue
        self.interval = interval
        self.chrome_path = chrome_path or APP_CONFIG.get('chrome_path')
        self.debug_port = debug_port
        self.stop_event = stop_event
        self.headless = headless
        self.driver: Optional[webdriver.Chrome] = None
        self._api_responses: list = []
        self._signature = ""
        self.filter_keywords = filter_keywords  # None=全量, list=只保留匹配的
        self._api_key = ""
        self._cookies = {}

    def log(self, msg: str):
        ts = datetime.now().strftime('%H:%M:%S')
        self.log_queue.put(f"[深圳投资网 {ts}] {msg}")

    def _build_driver(self) -> webdriver.Chrome:
        import urllib.request
        try:
            with urllib.request.urlopen(f"http://127.0.0.1:{self.debug_port}/json/version", timeout=2):
                pass
        except Exception:
            self.log("Chrome 未检测到，正在启动...")
            self._launch_chrome()

        options = ChromeOptions()
        if self.chrome_path:
            options.binary_location = self.chrome_path
        options.add_experimental_option("debuggerAddress", f"127.0.0.1:{self.debug_port}")
        if self.headless:
            options.add_argument("--headless=new")
        self.driver = webdriver.Chrome(options=options)
        return self.driver

    def _launch_chrome(self) -> None:
        import subprocess
        profile_dir = ROOT_DIR / "browser_debug_profile"
        profile_dir.mkdir(exist_ok=True)
        chrome_exe = self.chrome_path or r"C:\Program Files\Google\Chrome\Application\chrome.exe"
        cmd = [
            chrome_exe,
            f"--remote-debugging-port={self.debug_port}",
            f"--user-data-dir={profile_dir}",
            "--new-window",
            "about:blank",
        ]
        try:
            subprocess.Popen(cmd, shell=False, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            time.sleep(3)
            self.log("Chrome 已启动")
        except Exception as e:
            self.log(f"启动 Chrome 失败: {e}")

    def _generate_sig_key(self, sig: str) -> str:
        chars = '0123456789abcdef'
        key = ''
        key_index = -1
        for i in range(6):
            c = sig[key_index + 1] if key_index + 1 < len(sig) else sig[0]
            key += c
            idx = chars.find(c)
            key_index = idx if idx >= 0 else i
        return key

    def _call_api(self, params: dict) -> dict:
        """通过 Chrome JS fetch 调用深圳投资网 API（绕过 Python SSL 问题）"""
        if not self.driver:
            self.log("_call_api: driver 未初始化")
            return {}

        sig = self._signature or 'f795401781746567084'
        key = self._generate_sig_key(sig)
        ts = str(random.randint(1000, 9999)) + '_' + key + '_' + str(int(time.time() * 1000))
        t = ts.replace('+', '_')
        url = (f"{SZ_BASE_URL}/investment/api-v2/investment.project.ProjectCmd/execute"
               f"?s={sig}&t={t}")

        # 用 Chrome 的 XMLHttpRequest 发请求（Chrome 处理 SSL，不受 Python SSL 限制）
        js_code = """
        var url = arguments[0];
        var params = arguments[1];
        var done = false;
        var result = null;
        var xhr = new XMLHttpRequest();
        xhr.open('POST', url, false);  // 同步模式
        xhr.setRequestHeader('Content-Type', 'application/json');
        xhr.setRequestHeader('X-Requested-With', 'XMLHttpRequest');
        try {
            xhr.send(JSON.stringify(params));
            if (xhr.status === 200) {
                return JSON.parse(xhr.responseText);
            }
        } catch(e) {
            return {error: e.message};
        }
        return {error: 'status_' + xhr.status};
        """

        for attempt in range(3):
            try:
                result = self.driver.execute_script(js_code, url, params)
                if result and isinstance(result, dict) and 'error' not in result:
                    return result
                elif result and 'error' in result:
                    raise Exception(result['error'])
                else:
                    raise Exception(f"Unexpected: {result}")
            except Exception as e:
                if attempt < 2:
                    self.log(f"API调用失败(重试 {attempt + 1}/3): {e}")
                    time.sleep(2)
                else:
                    self.log(f"API调用失败(已重试3次): {e}")
                    return {}
        return {}

    def crawl_by_date(self, start_date: str, end_date: str) -> list[dict]:
        """按日期抓取深圳招采网（照搬 caitoubiao 已验证方案）"""
        all_records = []

        try:
            driver = self._build_driver()
            self.driver = driver
            self.log("Chrome 已连接")

            # 1. 访问深圳招采网，点击"项目公示"（caitoubiao 同款流程）
            driver.get(f"{SZ_BASE_URL}/investment/pubInformation/index")
            time.sleep(3)
            try:
                driver.find_element(By.CSS_SELECTOR, 'a[type="btn1"]').click()
                self.log("已打开'项目公示'页面")
                time.sleep(2)
            except Exception as e:
                self.log(f"点击项目公示标签失败: {e}")

            # 2. 翻页抓取（caitoubiao 同款逻辑）
            page = 1
            max_pages = 200
            while page <= max_pages:
                if self.stop_event and self.stop_event.is_set():
                    break

                self.log(f"正在抓取第 {page} 页...")

                try:
                    WebDriverWait(driver, 15).until(
                        lambda d: d.find_elements(By.CSS_SELECTOR, "#noticeList tr.tr_222"))
                except Exception:
                    pass

                rows = driver.find_elements(By.CSS_SELECTOR, "#noticeList tr.tr_222")
                if not rows:
                    self.log("当前页没有数据")
                    break

                page_has_target = False
                page_dates = set()

                # 首页调试：打印原始日期文本
                if page == 1 and rows:
                    try:
                        tds = rows[0].find_elements(By.TAG_NAME, "td")
                        raw = driver.execute_script("return arguments[0].innerText;", tds[4]) if len(tds) > 4 else ''
                        parsed = self._parse_date_text(raw)
                        self.log(f"  [DEBUG] raw_date='{raw}' → parsed='{parsed}' target='{start_date}'")
                    except Exception:
                        pass

                for row in rows:
                    if self.stop_event and self.stop_event.is_set():
                        break
                    try:
                        tds = row.find_elements(By.TAG_NAME, "td")
                        if len(tds) < 5:
                            continue

                        raw_date = driver.execute_script("return arguments[0].innerText;", tds[4])
                        clean_date = self._parse_date_text(raw_date)
                        if clean_date:
                            page_dates.add(clean_date)

                        if not (start_date <= clean_date <= end_date):
                            continue

                        project_code = driver.execute_script("return arguments[0].innerText;", tds[0]).strip()
                        project_name = driver.execute_script("return arguments[0].innerText;", tds[1]).strip()
                        company = driver.execute_script("return arguments[0].innerText;", tds[2]).strip()
                        project_type = driver.execute_script("return arguments[0].innerText;", tds[3]).strip()

                        if self.filter_keywords:
                            search_text = project_name + " " + company
                            if not any(kw in search_text for kw in self.filter_keywords):
                                continue

                        page_has_target = True
                        address = extract_address_from_name(project_name)

                        rec = {
                            'company_name': company,
                            'project_unit': company,
                            'contact_person': '', 'phone': '',
                            'address': address,
                            'main_company': '', 'main_contact': '', 'main_phone': '', 'main_address': '',
                            'investor': '',
                            'project_desc': project_name,
                            'project_status': project_type,
                            'location': extract_sz_district(project_name) or extract_sz_district(address),
                            'credit_code': project_code,
                            'relation_graph': '', 'remark': '', 'summary': '',
                            'source': '深圳投资网',
                            'record_date': clean_date,
                        }
                        all_records.append(rec)
                        self.log(f"  [+] {project_name[:40]}")
                    except Exception:
                        continue

                # 日志：本页日期范围
                if page_dates:
                    sd = sorted(page_dates)
                    self.log(f"  日期范围: {sd[0]} ~ {sd[-1]}")

                self.log(f"第 {page} 页: 累计 {len(all_records)} 条")

                # 停止判断（必须在翻页前！数据按时间降序，本页最大日期 < start → 已翻过）
                if not page_has_target and page_dates:
                    page_max = max(page_dates)
                    if page_max < start_date:
                        self.log(f"已翻过目标日期（本页最大{page_max} < 起始{start_date}），停止")
                        break

                # 翻页（caitoubiao 同款）
                try:
                    all_links = driver.find_elements(By.CSS_SELECTOR, ".pages a")
                    next_link = None
                    for link in all_links:
                        if "下一页" in link.text:
                            next_link = link
                            break
                    if next_link:
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_link)
                        time.sleep(0.5)
                        driver.execute_script("arguments[0].click();", next_link)
                        time.sleep(2)
                        page += 1
                    else:
                        self.log("没有下一页，停止")
                        break
                except Exception:
                    self.log("翻页异常，停止")
                    break

            self.log(f"共抓取 {len(all_records)} 条记录")

        except Exception as e:
            self.log(f"爬取出错: {e}")
            traceback.print_exc()

        finally:
            if self.driver:
                try:
                    self.driver.quit()
                except Exception:
                    pass

        return all_records

    @staticmethod
    def _parse_date_text(text):
        """caitoubiao 同款日期解析"""
        if not text:
            return ""
        for pattern in [
            r'(\d{4})-(\d{1,2})-(\d{1,2})',
            r'(\d{4})/(\d{1,2})/(\d{1,2})',
            r'(\d{4})年(\d{1,2})月(\d{1,2})日',
        ]:
            m = re.search(pattern, text)
            if m:
                return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
        return text.strip()

    def _crawl_by_dom(self, driver: webdriver.Chrome,
                       start_date: str, end_date: str) -> list[dict]:
        """Selenium DOM 翻页抓取（回退方案）

        深圳投资网表格列结构（从 HTML template 确认）：
        col[0] 国家编码  → credit_code / PROJECT_CODE
        col[1] 项目名称  → company_name / PROJECT_NAME
        col[2] 项目单位  → project_unit / XMSBDWMC    ← 重要字段，之前代码丢失
        col[3] 立项类型  → project_status
        col[4] 立项时间  → record_date / APPLY_DATE
        """
        records = []
        page = 1

        while True:
            self.log(f"DOM抓取 第 {page} 页...")

            try:
                # 等待表格加载 —— 修正：使用正确的选择器 tbody#noticeList tr
                WebDriverWait(driver, 10).until(
                    lambda d: d.find_elements(By.CSS_SELECTOR, 'tbody#noticeList tr, table.table tbody tr'))
                time.sleep(1)

                # 修正选择器：从 .listTable tbody tr（0行）改为 tbody#noticeList tr（正确）
                rows = driver.find_elements(By.CSS_SELECTOR,
                    'tbody#noticeList tr, table.table tbody tr')

                if not rows:
                    # 兜底：任意 table 里的 tr
                    rows = driver.find_elements(By.CSS_SELECTOR, 'table tbody tr')

                if not rows:
                    self.log("未找到表格行，停止")
                    break

                for row in rows:
                    try:
                        cells = row.find_elements(By.TAG_NAME, 'td')
                        if len(cells) < 5:
                            continue

                        # 用精确列索引提取（根据 HTML template 映射）
                        # col[0] 国家编码
                        credit_code = clean_text(cells[0].text) if len(cells) > 0 else ''
                        # col[1] 项目名称
                        company_name = clean_text(cells[1].text) if len(cells) > 1 else ''
                        # col[2] 项目单位 ← 之前完全丢失的字段！
                        project_unit = clean_text(cells[2].text) if len(cells) > 2 else ''
                        # col[3] 立项类型
                        project_status_text = clean_text(cells[3].text) if len(cells) > 3 else '备案'
                        # col[4] 立项时间
                        record_date = clean_text(cells[4].text) if len(cells) > 4 else start_date

                        if not company_name:
                            continue

                        # 项目描述：从项目名称中提取（名称通常含项目内容）
                        project_desc = company_name

                        # 从项目名称推断地址区划（深圳各区）
                        location = extract_sz_district(company_name)

                        rec = {
                            'company_name': company_name,
                            'project_unit': project_unit,
                            'contact_person': '',
                            'phone': '',
                            'address': '',
                            'main_company': '',
                            'main_contact': '',
                            'main_phone': '',
                            'main_address': '',
                            'investor': '',
                            'project_desc': project_desc,
                            'project_status': project_status_text if project_status_text else '备案',
                            'location': location if location else '深圳',
                            'credit_code': credit_code,
                            'relation_graph': '',
                            'remark': '',
                            'summary': '',
                            'source': '深圳投资网',
                            'record_date': record_date if record_date else start_date,
                        }

                        # 关键：如果本行日期超出指定范围，停止爬取（不再翻页）
                        if record_date:
                            row_date = record_date.strip()
                            if row_date < start_date.strip() or row_date > end_date.strip():
                                self.log(f"日期 {row_date} 超出范围 [{start_date}, {end_date}]，停止")
                                # 立即返回，不翻页
                                return records

                        records.append(rec)

                    except Exception:
                        continue

                self.log(f"第 {page} 页: {len(rows)} 条")

                # 翻页：查找下一页按钮
                # 页面结构: <a href="javascript:void(0);" onclick="pagesList.pageIndex(n);onQuery()">下一页</a>
                try:
                    next_btn = driver.find_element(
                        By.XPATH, "//a[contains(text(),'下一页')]")
                    if next_btn.is_enabled():
                        driver.execute_script("arguments[0].click();", next_btn)
                        time.sleep(2)
                        page += 1
                    else:
                        break
                except (NoSuchElementException, Exception):
                    self.log("没有下一页，停止")
                    break

            except TimeoutException:
                self.log("等待页面加载超时")
                break
            except Exception as e:
                self.log(f"DOM抓取第 {page} 页出错: {e}")
                break

        return records

    def _parse_api_item(self, item: dict, record_date: str) -> Optional[dict]:
        """解析 API 返回的项目项"""
        try:
            # 尝试多种可能的字段名（深圳投资网 API 字段名不固定）
            company_name = (
                clean_text(item.get('projectName', '') or item.get('companyName', '')
                or item.get('name', '') or item.get('enterpriseName', '')
                or item.get('company_name', ''))
            )

            # 项目单位：申报该项目的主体（如"深圳市发改委"或具体公司）
            project_unit = (
                clean_text(item.get('projectUnit', '') or item.get('projectUnitName', '')
                or item.get('unitName', '') or item.get('declareUnit', '')
                or item.get('申报单位', '') or item.get('investUnit', ''))
            )

            address = (
                clean_text(item.get('address', '') or item.get('projectAddress', '')
                or item.get('projectAddr', '') or item.get('建地址', ''))
            )

            credit_code = (
                clean_text(item.get('projectCode', '') or item.get('code', '')
                or item.get('creditCode', '') or item.get('备案号', ''))
            )

            # 项目情况：项目具体内容描述
            project_desc = (
                clean_text(item.get('projectDesc', '') or item.get('content', '')
                or item.get('projectContent', '') or item.get('建设内容', '')
                or item.get('description', ''))
            )

            project_status = (
                clean_text(item.get('status', '') or item.get('projectStatus', '')
                or item.get('state', '') or '已备案')
            )

            phone = extract_phone(str(item))

            # 光伏关键词筛选（如果配置了关键词，不匹配的跳过）
            if self.filter_keywords:
                search_text = f"{company_name} {project_unit} {project_desc} {item}"
                if not any(kw in search_text for kw in self.filter_keywords):
                    return None  # 不匹配光伏关键词，跳过

            rec = {
                'company_name': company_name,
                'project_unit': project_unit,
                'contact_person': '',
                'phone': phone,
                'address': address,
                'main_company': '',
                'main_contact': '',
                'main_phone': '',
                'main_address': '',
                'investor': '',
                'project_desc': project_desc,
                'project_status': project_status,
                'location': extract_sz_district(address),
                'credit_code': credit_code,
                'relation_graph': '',
                'remark': '',
                'summary': '',
                'source': '深圳投资网',
                'record_date': record_date,
            }
            return rec
        except Exception as e:
            self.log(f"解析项目出错: {e}")
            return None

    def quit(self):
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = None


# ============================================================
# 天眼查爬虫（公开数据补全，无需登录）
# ============================================================
def extract_address_from_name(project_name: str) -> str:
    """从项目名称中提取地址"""
    if not project_name:
        return ''
    patterns = [
        r'(深圳市?\s*[^\s]{0,5}区\s*[^\s]{0,10}(?:街道|镇)\s*[^\s]{0,30}(?:路|街|大道)\s*[^\s]{0,10}号)',
        r'([^\s]{0,5}区\s*[^\s]{0,10}(?:街道|镇)\s*[^\s]{0,30}(?:路|街|大道)\s*[^\s]{0,10}号)',
        r'([^\s]{0,5}区\s*[^\s]{0,10}(?:街道|镇)\s*[^\s]{0,20}(?:社区|村)\s*[^\s]{0,20}(?:路|街|大道)\s*[^\s]{0,10}号)',
        r'([^\s]{0,5}(?:街道|镇)\s*[^\s]{0,20}(?:路|街)\s*[^\s]{0,10}号)',
        r'(深圳市?\s*[^\s]{0,5}区\s*[^\s]{0,10}(?:街道|镇)\s*[^\s]{0,20}(?:工业园|产业园|科技园|园区))',
        r'([^\s]{0,5}区\s*[^\s]{0,10}(?:街道|镇)\s*[^\s]{0,20}(?:工业园|产业园|科技园|园区))',
    ]
    for pat in patterns:
        m = re.search(pat, project_name)
        if m:
            addr = m.group(1).strip()
            if len(addr) >= 6:
                return addr
    qu_idx = project_name.find('区')
    for end_kw in ['号', '园', '大厦', '栋', '楼']:
        end_idx = project_name.rfind(end_kw)
        if 0 < qu_idx < end_idx < 100:
            return project_name[:end_idx + 1] if qu_idx == 0 else project_name[max(0,qu_idx-4):end_idx + 1]
    return ''


class TianyanchaEnricher:
    """天眼查企业信息补全 - 通过公开页面抓取法定代表人、地址等

    天眼查详情页无需登录即可查看：
    - 法定代表人/联系人
    - 企业地址
    - 股东信息（部分）
    - 注册资本等基本信息

    需要登录才能查看：手机号（付费数据）
    """

    TYC_SEARCH_URL = "https://www.tianyancha.com/search"
    TYC_COMPANY_URL = "https://www.tianyancha.com/company/"

    def __init__(self, log_queue: queue.Queue,
                 interval: int = 3, chrome_path: str = None,
                 debug_port: int = 9222, stop_event=None):
        self.log_queue = log_queue
        self.interval = interval
        self.chrome_path = chrome_path or APP_CONFIG.get('chrome_path')
        self.debug_port = debug_port
        self.driver: Optional[webdriver.Chrome] = None
        self.stop_event = stop_event

    def log(self, msg: str):
        ts = datetime.now().strftime('%H:%M:%S')
        self.log_queue.put(f"[天眼查 {ts}] {msg}")

    def _build_driver(self) -> webdriver.Chrome:
        import urllib.request
        try:
            with urllib.request.urlopen(
                    f"http://127.0.0.1:{self.debug_port}/json/version", timeout=2):
                pass
        except Exception:
            self.log("Chrome 未运行，正在启动...")
            self._launch_chrome()

        options = ChromeOptions()
        if self.chrome_path:
            options.binary_location = self.chrome_path
        profile_dir = ROOT_DIR / "browser_debug_profile_tyc"
        profile_dir.mkdir(exist_ok=True)
        options.add_argument(f"--user-data-dir={profile_dir}")
        options.add_experimental_option("debuggerAddress", f"127.0.0.1:{self.debug_port}")
        self.driver = webdriver.Chrome(options=options)
        return self.driver

    def _launch_chrome(self) -> None:
        import subprocess
        chrome_exe = self.chrome_path or r"C:\Program Files\Google\Chrome\Application\chrome.exe"
        profile_dir = ROOT_DIR / "browser_debug_profile_tyc"
        profile_dir.mkdir(exist_ok=True)
        cmd = [
            chrome_exe,
            f"--remote-debugging-port={self.debug_port}",
            f"--user-data-dir={profile_dir}",
            "--new-window",
            "about:blank",
        ]
        try:
            subprocess.Popen(cmd, shell=False,
                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            import time
            time.sleep(3)
            self.log("Chrome 已启动")
        except Exception as e:
            self.log(f"启动 Chrome 失败: {e}")

    def enrich(self, company_name: str) -> dict:
        """根据公司名从天眼查获取补全信息

        Returns:
            dict 包含: legal_person, address, phone(可能为空), shareholder
        """
        result = {
            'contact_person': '',   # 法定代表人/联系人
            'phone': '',             # 手机号（天眼查公开页通常看不到）
            'address': '',          # 企业地址
            'main_company': '',      # 主体公司（如有）
            'investor': '',         # 投资人/股东
            'remark': '',
        }

        try:
            import urllib.parse
            query = urllib.parse.quote(company_name[:30])
            search_url = f"{self.TYC_SEARCH_URL}?key={query}"

            self.log(f"搜索: {company_name[:15]}...")
            self.driver.get(search_url)
            import time
            time.sleep(2)

            # 找公司详情链接
            links = self.driver.find_elements(
                By.CSS_SELECTOR, f'a[href^="{self.TYC_COMPANY_URL}"]')
            if not links:
                self.log(f"  未找到: {company_name[:15]}")
                return result

            detail_url = links[0].get_attribute('href')
            self.log(f"  进入详情: {detail_url[-20:]}")

            # 访问详情页 + 快速滚动触发懒加载
            self.driver.get(detail_url)
            time.sleep(2)
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(0.5)
            self.driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(0.3)

            # 用 BeautifulSoup 解析 HTML（比纯正则精确）
            page_source = self.driver.page_source
            soup = BeautifulSoup(page_source, 'lxml')

            # ========== 方法1：CSS 选择器精确抓取 ==========
            def find_label_value(soup, *label_texts):
                """查找 label 文本后跟的 value"""
                for label_text in label_texts:
                    # 方式A: <span class="label">XXX</span><span>值</span>
                    label = soup.find(['span', 'div', 'td'],
                                   string=re.compile(label_text))
                    if label:
                        parent = label.parent
                        if parent:
                            sibling = parent.find_next_sibling()
                            if sibling:
                                text = clean_text(sibling.get_text())
                                if text:
                                    return text
                            value_el = parent.find(['span', 'div'],
                                                  class_=re.compile('value|content', re.I))
                            if value_el:
                                return clean_text(value_el.get_text())

                    # 方式B: <div>XXX: 值</div>
                    div = soup.find(['div', 'td'],
                                  string=re.compile(label_text + '[：:]'))
                    if div:
                        text = div.get_text()
                        parts = re.split(r'[：:]', text, 1)
                        if len(parts) > 1:
                            val = clean_text(parts[1])
                            if val:
                                return val

                    # 方式C: dt/dd 结构
                    for dt in soup.find_all(['dt', 'th'],
                                         string=re.compile(label_text)):
                        dd = dt.find_next_sibling(['dd', 'td'])
                        if dd:
                            return clean_text(dd.get_text())
                return ''

            # 1. 法定代表人
            legal = find_label_value(soup, '法定代表人', '法人代表', '法人')
            if legal:
                result['contact_person'] = legal
            else:
                # 兜底：正则
                text = self.driver.execute_script("return document.body.innerText")
                for pat in [
                    r'法定代表人[：:]\s*([^\s]{2,20})',
                    r'法人代表[：:]\s*([^\s]{2,20})',
                    r'法人[：:]\s*([^\s]{2,20})',
                ]:
                    m = re.search(pat, text)
                    if m:
                        result['contact_person'] = m.group(1).strip()
                        break

            # 2. 企业地址
            addr = find_label_value(soup, '地址', '企业地址', '公司地址')
            if addr:
                result['address'] = addr
            else:
                text = self.driver.execute_script("return document.body.innerText")
                for pat in [r'地址[：:]\s*([^\n]{5,100})']:
                    m = re.search(pat, text)
                    if m:
                        result['address'] = m.group(1).strip()
                        break

            # 3. 股东/投资人
            result['investor'] = find_label_value(soup, '股东', '投资人', '主要人员')

            # 4. 手机号（公开页可能部分可见）
            if not result['phone']:
                text = self.driver.execute_script("return document.body.innerText")
                phones = re.findall(r'(?<![0-9A-Z])1[3-9]\d{9}(?![0-9A-Z])', text)
                for p in phones:
                    if p and len(p) == 11 and not p.startswith(('400', '800')):
                        result['phone'] = p
                        break

            # 5. 标注备注（VIP 状态）
            page_text = self.driver.execute_script("return document.body.innerText")
            if 'VIP' in page_text or '升级会员' in page_text or '立即开通' in page_text:
                if not result['contact_person']:
                    result['remark'] = '天眼查需VIP查看'
                elif not result['phone']:
                    result['remark'] = '手机号需VIP'

            self.log(f"  完成: 法代={result['contact_person'][:10] if result['contact_person'] else '-'}, "
                    f"地址={result['address'][:15] if result['address'] else '-'}"
                    f", 手机={result['phone'] or '-'}")

        except Exception as e:
            self.log(f"  出错: {e}")

        return result

    def enrich_batch(self, company_names: list[str]) -> list[dict]:
        """批量补全公司信息"""
        all_results = []
        driver = self._build_driver()
        self.driver = driver

        try:
            total = len(company_names)
            for idx, name in enumerate(company_names):
                if self.stop_event and self.stop_event.is_set():
                    self.log("收到停止信号，退出")
                    break

                self.log(f"[{idx+1}/{total}] {name[:15]}...")
                info = self.enrich(name)
                info['company_name'] = name
                all_results.append(info)

                import time
                time.sleep(self.interval)

        finally:
            self.quit()

        self.log(f"天眼查补全完成: {len(all_results)} 条")
        return all_results

    def quit(self):
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = None


# ============================================================
# 搜索引擎补全（Bing + 百度，纯 requests，无需浏览器）
# ============================================================
class SearchEnricher:
    """搜索引擎补全 - 用 requests 调用 Bing/百度搜索，从公开招标公告中提取项目信息

    数据来源：政府采购网、各地公共资源交易中心等政府公开招标平台
    无需浏览器、无需账号、支持无头运行
    """

    # 搜索查询模板（依次尝试，从精确到宽泛）
    SEARCH_QUERIES = [
        '"{company}" 光伏 中标 招标',
        '"{company}" 新能源 项目',
        '光伏 中标 "{company}"',
    ]

    # 全局发现搜索（不限于特定公司，发现新的光伏项目）
    DISCOVERY_QUERIES = [
        '分布式光伏 中标公告 联系人 电话',
        '光伏车棚 超充站 中标 有限公司',
        '光储充一体化 充电站 项目 中标 2026',
    ]

    HEADERS = {
        "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"),
        "Accept-Language": "zh-CN,zh;q=0.9",
    }

    def __init__(self, log_queue: queue.Queue,
                 interval: float = 1.5, stop_event=None,
                 filter_keywords: list = None):
        self.log_queue = log_queue
        self.interval = interval
        self.stop_event = stop_event
        self.filter_keywords = filter_keywords or PV_FILTER_KEYWORDS
        self._session = None

    def log(self, msg: str):
        ts = datetime.now().strftime('%H:%M:%S')
        self.log_queue.put(f"[搜索补全 {ts}] {msg}")

    @property
    def session(self):
        if self._session is None:
            self._session = requests.Session()
            self._session.headers.update(self.HEADERS)
        return self._session

    def _search_web(self, query: str) -> list[dict]:
        """搜索（DuckDuckGo Lite，不会被封），返回 [{title, url, snippet}]"""
        import urllib.parse
        # DuckDuckGo Lite: 极简HTML，从不要求验证码
        url = f"https://lite.duckduckgo.com/lite/?q={urllib.parse.quote(query)}"

        try:
            resp = requests.get(url, headers=self.HEADERS, timeout=15)
            if resp.status_code != 200:
                self.log(f"  搜索 HTTP {resp.status_code}")
                return []
        except Exception as e:
            self.log(f"  搜索请求失败: {e}")
            return []

        soup = BeautifulSoup(resp.text, 'lxml')
        results = []

        # DuckDuckGo Lite 结构: table 中包含 tr 结果行
        for tr in soup.select('table tbody tr'):
            try:
                link_el = tr.select_one('a.result-link, a[rel="nofollow"]')
                if not link_el:
                    link_el = tr.select_one('a')
                title = clean_text(link_el.get_text()) if link_el else ''
                link = link_el.get('href', '') if link_el else ''
                snippet_el = tr.select_one('td.result-snippet, .result-snippet')
                snippet = clean_text(snippet_el.get_text()) if snippet_el else ''
                # 也尝试从下一个 td 获取摘要
                if not snippet:
                    all_tds = tr.select('td')
                    if len(all_tds) >= 2:
                        snippet = clean_text(all_tds[-1].get_text())
                        # 去掉标题部分
                        if title and snippet.startswith(title[:10]):
                            snippet = snippet[len(title):].strip()
                if title or snippet:
                    results.append({'title': title, 'url': link, 'snippet': snippet})
            except Exception:
                continue

        # 如果 DuckDuckGo 没结果，回退到 Bing（可能被限流）
        if not results:
            try:
                bing_url = f"https://www.bing.com/search?q={urllib.parse.quote(query)}&setlang=zh-cn"
                resp2 = requests.get(bing_url, headers=self.HEADERS, timeout=10)
                if resp2.status_code == 200:
                    soup2 = BeautifulSoup(resp2.text, 'lxml')
                    for li in soup2.select('li.b_algo')[:10]:
                        try:
                            h2 = li.select_one('h2 a')
                            t = clean_text(h2.get_text()) if h2 else ''
                            l = h2.get('href', '') if h2 else ''
                            se = li.select_one('.b_caption p, .b_lineclamp2')
                            s = clean_text(se.get_text()) if se else ''
                            if t or s:
                                results.append({'title': t, 'url': l, 'snippet': s})
                        except Exception:
                            continue
            except Exception:
                pass

        return results

    def _extract_info(self, results: list[dict], company_name: str = '') -> dict:
        """从搜索结果中提取结构化信息"""
        info = {
            'contact_person': '',
            'phone': '',
            'address': '',
            'project_desc': '',
            'project_status': '',
            'investor': '',
            'main_company': '',
            'main_contact': '',
            'main_phone': '',
            'main_address': '',
            'remark': '',
            'summary': '',
        }

        # 合并所有文本用于提取
        all_text_parts = []
        for r in results:
            all_text_parts.append(r['title'])
            all_text_parts.append(r['snippet'])
        all_text = '\n'.join(all_text_parts)

        if not all_text:
            return info

        # 1. 提取手机号（从招标公告中常有联系人电话）
        phones = re.findall(r'1[3-9]\d{9}', all_text)
        # 过滤400/800开头的
        valid_phones = [p for p in phones if not p.startswith(('400', '800'))]
        if valid_phones:
            info['phone'] = valid_phones[0]

        # 同时也提取固话
        landlines = re.findall(r'(?:电话|联系电话|联系人电话|Tel)[：:]\s*([\d\-]{7,15})', all_text)
        if landlines and not info['phone']:
            info['phone'] = landlines[0].strip()

        # 2. 提取联系人
        contact_pats = [
            r'(?:联系人|项目联系人)[：:]\s*([^\s\n]{2,10})',
            r'联系人[：:]\s*([^\s\n]{2,10})',
        ]
        for pat in contact_pats:
            m = re.search(pat, all_text)
            if m:
                cand = m.group(1).strip()
                if len(cand) >= 2 and not cand[0].isdigit():
                    info['contact_person'] = cand
                    break

        # 3. 提取项目描述（从标题和摘要中收集光伏相关文本）
        pv_kws = ['光伏', '充电站', '车棚', '储能', '光储充', '新能源', '分布式',
                   '超充', 'EPC', 'BIPV', '电站', '逆变器', '太阳能']
        desc_parts = []
        for r in results:
            title = r['title']
            snippet = r['snippet']
            combined = title + ' ' + snippet
            if any(kw in combined for kw in pv_kws):
                # 取标题+摘要中有信息量的部分
                if len(title) > 10:
                    desc_parts.append(title)
                if len(snippet) > 20:
                    desc_parts.append(snippet[:200])

        if desc_parts:
            info['project_desc'] = ' | '.join(desc_parts[:4])[:500]

        # 4. 提取项目状态
        if any(kw in all_text for kw in ['中标', '成交公告', '中标公告']):
            info['project_status'] = '中标'
        elif any(kw in all_text for kw in ['招标', '招标公告']):
            info['project_status'] = '招标中'
        elif any(kw in all_text for kw in ['已投产', '并网发电', '已建成']):
            info['project_status'] = '已投产'
        elif any(kw in all_text for kw in ['在建', '施工中', '建设中', '开工建设']):
            info['project_status'] = '在建'
        elif any(kw in all_text for kw in ['备案', '获批', '核准', '公示', '环评']):
            info['project_status'] = '规划中'

        # 5. 提取中标单位/投资人
        investor_pats = [
            r'(?:中标单位|中标人|成交供应商)[：:]\s*([^\s\n]{4,40})',
            r'(?:招标人|采购人)[：:]\s*([^\s\n]{4,40})',
        ]
        for pat in investor_pats:
            m = re.search(pat, all_text)
            if m:
                cand = clean_text(m.group(1))
                if cand and len(cand) > 3:
                    info['investor'] = cand[:40]
                    break

        # 6. 提取地址
        addr_pats = [
            r'地址[：:]\s*([^\n]{5,60})',
            r'地点[：:]\s*([^\n]{5,60})',
        ]
        for pat in addr_pats:
            m = re.search(pat, all_text)
            if m:
                cand = clean_text(m.group(1))
                if any(k in cand for k in ['路', '号', '区', '镇', '街道', '园', '大厦', '省', '市']):
                    info['address'] = cand[:60]
                    break

        # 7. 标记数据来源
        urls = [r['url'] for r in results if r['url']]
        if urls:
            info['remark'] = urls[0][:80]  # 保存第一个结果URL作为参考

        return info

    def enrich(self, company_name: str) -> dict:
        """搜索单个公司，返回补全信息"""
        result = {
            'contact_person': '', 'phone': '', 'address': '',
            'project_desc': '', 'project_status': '', 'investor': '',
            'main_company': '', 'main_contact': '', 'main_phone': '',
            'main_address': '', 'remark': '', 'summary': '',
        }

        try:
            all_results = []
            for query_template in self.SEARCH_QUERIES:
                if self.stop_event and self.stop_event.is_set():
                    break
                query = query_template.format(company=company_name[:30])
                results = self._search_web(query)
                all_results.extend(results)
                time.sleep(0.8)

            if all_results:
                info = self._extract_info(all_results, company_name)
                result.update(info)

            if result['project_desc']:
                self.log(f"  ✓ {company_name[:15]}... ({len(result['project_desc'])}字)"
                         + (f" 电话:{result['phone']}" if result['phone'] else ""))
            else:
                self.log(f"  - {company_name[:15]}... 未找到")

        except Exception as e:
            self.log(f"  出错: {company_name[:15]} - {e}")

        return result

    def enrich_batch(self, company_names: list[str]) -> list[dict]:
        """批量补全公司信息（无需浏览器，纯requests）"""
        all_results = []

        try:
            total = len(company_names)
            for idx, name in enumerate(company_names):
                if self.stop_event and self.stop_event.is_set():
                    self.log("收到停止信号，退出")
                    break

                self.log(f"[{idx+1}/{total}] {name[:15]}...")

                info = self.enrich(name)
                info['company_name'] = name
                info['source'] = '搜索补全'
                all_results.append(info)

                time.sleep(self.interval)

        except Exception as e:
            self.log(f"搜索补全异常: {e}")
            traceback.print_exc()

        # 统计
        with_phone = sum(1 for r in all_results if r.get('phone'))
        with_desc = sum(1 for r in all_results if r.get('project_desc'))
        self.log(f"搜索补全完成: {len(all_results)} 条 "
                 f"(手机号:{with_phone}, 项目描述:{with_desc})")
        return all_results

    def quit(self):
        """无需清理（无浏览器资源）"""
        if self._session:
            try:
                self._session.close()
            except Exception:
                pass
            self._session = None


# ============================================================
# 探迹爬虫（复用现有工具逻辑）
# ============================================================
# 探迹光伏项目筛选关键词
PV_FILTER_KEYWORDS = [
    "车棚", "超充站", "光储充", "分布式光伏", "屋顶分布式光伏",
    "光伏EPC", "充电站", "光伏发电", "光伏项目", "光伏电站",
    "BIPV", "光伏车棚", "储能", "光伏组件", "逆变器",
]

class TungeeCrawler:
    """探迹爬虫 - Selenium 自动登录"""

    def __init__(self, log_queue: queue.Queue, phone: str, password: str,
                 interval: int = 3, chrome_path: str = None,
                 debug_port: int = 9222, stop_event=None,
                 filter_keywords: list = None, headless: bool = False):
        self.log_queue = log_queue
        self.phone = phone
        self.password = password
        self.interval = interval
        self.chrome_path = chrome_path or APP_CONFIG.get('chrome_path')
        self.debug_port = debug_port
        self.driver: Optional[webdriver.Chrome] = None
        self._logged_in = False
        self.stop_event = stop_event
        self.filter_keywords = filter_keywords if filter_keywords is not None else PV_FILTER_KEYWORDS
        self.headless = headless

    def log(self, msg: str):
        ts = datetime.now().strftime('%H:%M:%S')
        self.log_queue.put(f"[探迹 {ts}] {msg}")

    def _build_driver(self) -> webdriver.Chrome:
        """创建 Chrome 实例（支持无头模式和远程调试模式）"""
        import urllib.request

        options = ChromeOptions()
        if self.chrome_path:
            options.binary_location = self.chrome_path

        if self.headless:
            # 无头模式：始终启动独立的 headless Chrome（不复用已有实例）
            self.log("无头模式：创建独立 headless Chrome 实例...")
            options.add_argument("--headless=new")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument(
                "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            )
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option("useAutomationExtension", False)
            # 每个实例用独立 profile 目录，避免端口冲突
            profile_dir = ROOT_DIR / f"chrome_profile_{os.getpid()}_{threading.get_ident()}_{id(self)}"
            options.add_argument(f"--user-data-dir={profile_dir}")
        else:
            # 非无头模式：启动独立的可见 Chrome 实例
            profile_dir = ROOT_DIR / f"browser_debug_profile_{os.getpid()}_{threading.get_ident()}_{id(self)}"
            profile_dir.mkdir(exist_ok=True)
            options.add_argument(f"--user-data-dir={profile_dir}")
            options.add_argument("--new-window")
            self.log("非无头模式：启动独立可见 Chrome 实例...")

        self.driver = webdriver.Chrome(options=options)
        return self.driver

    def _launch_chrome(self) -> None:
        import subprocess
        profile_dir = ROOT_DIR / "browser_debug_profile"
        profile_dir.mkdir(exist_ok=True)
        chrome_exe = self.chrome_path or r"C:\Program Files\Google\Chrome\Application\chrome.exe"
        cmd = [
            chrome_exe,
            f"--remote-debugging-port={self.debug_port}",
            f"--user-data-dir={profile_dir}",
            "--new-window",
            "about:blank",
        ]
        try:
            subprocess.Popen(cmd, shell=False, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            time.sleep(3)
            self.log("Chrome 已启动")
        except Exception as e:
            self.log(f"启动 Chrome 失败: {e}")

    def _do_login(self, driver: webdriver.Chrome) -> bool:
        """执行登录（优先密码表单，否则扫码）"""
        try:
            # 先检查是不是已经登录了（直接访问home页）
            driver.get("https://user.tungee.com/home")
            time.sleep(3)
            cur_url = driver.current_url
            body_text = driver.execute_script("return document.body.innerText;") or ""

            # 已经登录：在个人中心页面
            if '/home' in cur_url or '退出' in body_text or '个人中心' in body_text:
                if '退出' in body_text or '退出账户' in body_text:
                    self.log("检测到已登录状态（session有效），跳过登录")
                    self._save_session(driver)
                    return True

            # 未登录：导航到登录页
            self.log("未登录，导航到登录页...")
            driver.delete_all_cookies()
            driver.get(TUNGEE_LOGIN_URL)
            time.sleep(4)

            # 优先检测密码表单（通过找输入框元素，不依赖 placeholder 文本）
            try:
                phone_inp = driver.find_element(By.CSS_SELECTOR, "input[placeholder='请输入手机号码']")
                pw_inp = driver.find_element(By.CSS_SELECTOR, "input[type='password']")
                if phone_inp.is_displayed() and pw_inp.is_displayed():
                    self.log("检测到密码表单，执行自动登录...")
                    ok = self._login_with_password(driver)
                    if ok:
                        return True
                    self.log("自动登录失败")
            except Exception:
                pass

            # 没有密码表单 → 点击"账号登录"Tab后再尝试
            try:
                # 用JS点击账号登录Tab
                driver.execute_script('''
                    var tabs = document.querySelectorAll("._2me12");
                    for(var t of tabs) {
                        if(t.innerText && t.innerText.includes("账号")) {
                            t.click(); return;
                        }
                    }
                ''')
                time.sleep(2)
                phone_inp = driver.find_element(By.CSS_SELECTOR, "input[placeholder='请输入手机号码']")
                pw_inp = driver.find_element(By.CSS_SELECTOR, "input[type='password']")
                if phone_inp.is_displayed() and pw_inp.is_displayed():
                    self.log("点击账号登录Tab后检测到表单，执行自动登录...")
                    ok = self._login_with_password(driver)
                    if ok:
                        return True
                    self.log("自动登录失败")
            except Exception:
                pass

            # 自动登录失败 → 等用户手动登录
            self.log("=" * 50)
            self.log("探迹需要手动登录（可能有验证码）")
            self.log("请在Chrome窗口完成登录，工具会自动继续")
            self.log("注意：如果看不到浏览器窗口，请手动打开 Chrome")
            self.log("访问 https://user.tungee.com/users/sign-in 完成登录")
            self.log("登录成功后此窗口会自动继续...")
            self.log("（等待最多10分钟，超时后自动退出）")
            self.log("=" * 50)
            driver.get(TUNGEE_LOGIN_URL)

            # 10分钟超时（600秒），给用户足够时间解决验证码
            for i in range(600):
                time.sleep(1)
                try:
                    cur = driver.current_url
                    body = driver.execute_script("return document.body.innerText;") or ""
                    if '/home' in cur or '退出' in body:
                        self.log("登录成功！session已保存")
                        self._save_session(driver)
                        return True
                except Exception:
                    pass
                if self.stop_event and self.stop_event.is_set():
                    return False
                # 每30秒打印一次状态
                if i > 0 and i % 30 == 0:
                    self.log(f"  等待登录中... ({i}秒)")

            self.log("手动登录超时（10分钟）")
            return False

        except Exception as e:
            self.log(f"登录异常: {e}")
            return False

    def _login_with_password(self, driver: webdriver.Chrome) -> bool:
        """用手机号+密码登录（表单已显示，直接填值提交）"""
        try:
            # 等待表单元素出现
            from selenium.webdriver.support.ui import WebDriverWait
            try:
                WebDriverWait(driver, 15).until(
                    lambda d: d.find_element(By.CSS_SELECTOR, "input[type='password']"))
            except Exception:
                self.log("等待密码输入框超时")

            time.sleep(1)

            # 3. 填写表单（React Ant Design 专用策略）
            # 手机号输入框：优先用 placeholder 定位
            phone_inp = None
            for sel in [
                "input[placeholder='请输入手机号码']",
                "input.ant-input[type='text']",
                "input[placeholder*='手机']",
            ]:
                try:
                    el = driver.find_element(By.CSS_SELECTOR, sel)
                    if el.is_displayed():
                        phone_inp = el
                        break
                except Exception:
                    continue

            if phone_inp:
                # JS填值+React事件 → 确保React状态更新
                driver.execute_script("""
                    var el = arguments[0], val = arguments[1];
                    var nativeInputValueSetter = Object.getOwnPropertyDescriptor(
                        window.HTMLInputElement.prototype, 'value').set;
                    nativeInputValueSetter.call(el, val);
                    el.dispatchEvent(new Event('input', {bubbles: true}));
                    el.dispatchEvent(new Event('change', {bubbles: true}));
                """, phone_inp, self.phone)
                self.log(f"已填写手机号: {self.phone[:3]}***")
            else:
                self.log("未找到手机号输入框")
                return False

            time.sleep(0.3)

            # 密码输入框
            pw_inp = None
            for sel in [
                "input[placeholder*='密码']",
                "input[type='password']",
            ]:
                try:
                    el = driver.find_element(By.CSS_SELECTOR, sel)
                    if el.is_displayed():
                        pw_inp = el
                        break
                except Exception:
                    continue

            if pw_inp:
                driver.execute_script("""
                    var el = arguments[0], val = arguments[1];
                    var nativeInputValueSetter = Object.getOwnPropertyDescriptor(
                        window.HTMLInputElement.prototype, 'value').set;
                    nativeInputValueSetter.call(el, val);
                    el.dispatchEvent(new Event('input', {bubbles: true}));
                    el.dispatchEvent(new Event('change', {bubbles: true}));
                """, pw_inp, self.password)
                self.log("已填写密码")
            else:
                self.log("未找到密码输入框")
                return False

            time.sleep(0.5)

            # 4. 提交：多种方式兼容React
            try:
                buttons = driver.find_elements(By.CSS_SELECTOR, "button.ant-btn-primary")
                for btn in buttons:
                    if '登' in btn.text or '登' in btn.text:
                        if btn.is_enabled() and btn.is_displayed():
                            driver.execute_script("arguments[0].click();", btn)
                            self.log("已点击登录按钮")
                            break
            except Exception:
                pass

            # 兜底：按回车
            try:
                pw_inp = driver.find_element(By.CSS_SELECTOR, "input[type='password']")
                pw_inp.send_keys(Keys.RETURN)
                self.log("已按回车提交")
            except Exception:
                pass

            time.sleep(3)

            # 检测登录错误提示
            try:
                body = driver.execute_script("return document.body.innerText;") or ""
                if '密码错误' in body or '账号不存在' in body:
                    self.log(f"登录失败: 密码错误或账号不存在")
                    return False
                if '验证码' in body or '滑块' in body or '图形验证' in body:
                    self.log("登录需要验证码/滑块验证，请手动登录后重试")
                    return False
            except Exception:
                pass

            # 等待登录成功跳转
            for i in range(10):
                if '/home' in driver.current_url:
                    self.log("登录成功，跳转到个人中心")
                    self._save_session(driver)
                    return True
                time.sleep(1)
            self.log(f"登录后 URL: {driver.current_url}")
            # 截图保存登录失败状态用于调试
            try:
                driver.save_screenshot('debug_output/tungee_login_failed.png')
                self.log("登录失败截图已保存: debug_output/tungee_login_failed.png")
            except Exception:
                pass
            return False

        except Exception as e:
            self.log(f"密码登录异常: {e}")
            return False

    def _login_with_qr(self, driver: webdriver.Chrome) -> bool:
        """等待扫码登录"""
        try:
            try:
                ss_path = Path.home() / "Desktop" / "tungee_qrcode.png"
                driver.save_screenshot(str(ss_path))
                self.log(f"二维码截图已保存: {ss_path}")
            except Exception:
                pass

            last_url = driver.current_url
            for i in range(24):
                time.sleep(5)
                try:
                    current_url = driver.current_url
                    current_body = driver.execute_script("return document.body.innerText;")
                    if 'sign-in' not in current_url and 'user.tungee' not in current_url:
                        self.log(f"扫码登录成功！({(i+1)*5}s)")
                        self._save_session(driver)
                        return True
                    if current_url == last_url and '扫码' not in current_body:
                        time.sleep(5)
                        if 'sign-in' not in driver.current_url:
                            self.log("登录成功")
                            self._save_session(driver)
                            return True
                    last_url = current_url
                    self.log(f"等待扫码中... ({(i+1)*5}s)")
                except Exception:
                    pass
            self.log("扫码超时")
            return False
        except Exception as e:
            self.log(f"扫码异常: {e}")
            return False

    def _save_session(self, driver: webdriver.Chrome) -> None:
        """保存当前 session 的 cookies 到文件"""
        try:
            cookies = driver.get_cookies()
            session_file = ROOT_DIR / "tungee_session.json"
            with open(session_file, 'w', encoding='utf-8') as f:
                json.dump(cookies, f, ensure_ascii=False, indent=2)
            self.log(f"Session 已保存: {session_file} ({len(cookies)} cookies)")
        except Exception as e:
            self.log(f"Session 保存失败: {e}")

    def _load_session(self, driver: webdriver.Chrome) -> bool:
        """加载已保存的 session cookies"""
        try:
            session_file = ROOT_DIR / "tungee_session.json"
            if not session_file.exists():
                return False
            with open(session_file, 'r', encoding='utf-8') as f:
                cookies = json.load(f)
            for cookie in cookies:
                try:
                    driver.add_cookie(cookie)
                except Exception:
                    pass
            self.log(f"已加载 {len(cookies)} 个 cookies")
            return True
        except Exception as e:
            self.log(f"Session 加载失败: {e}")
            return False

    def ensure_logged_in(self) -> Optional[webdriver.Chrome]:
        """确保已登录，返回 driver 或 None

        流程：尝试session → 失败则调用 _do_login → 进入 bidding
        """
        try:
            driver = self._build_driver()
            self.driver = driver

            # Step 0: 尝试加载已有 session
            self.log("尝试加载已有 session...")
            driver.get("https://user.tungee.com/home")
            time.sleep(3)

            body = driver.execute_script("return document.body.innerText") or ""
            # 检测是否已登录（个人中心页面有"退出"/"退出账户"等文字）
            if '退出' in body and ('个人中心' in body or '手机' in body):
                self.log("Session 有效，已登录")
                self._logged_in = True
                self._save_session(driver)
                # 直接进入 bidding
                driver.get("https://bidding.tungee.com/")
                time.sleep(3)
                if 'bidding' in driver.current_url or 'tungee' in driver.current_url:
                    self.log("bidding 平台访问成功")
                    return driver
                # bidding 可能跳转了，不管怎样继续
                self.log("bidding 访问完成")
                return driver

            # Step 1: Session无效，重新登录
            self.log("Session 无效，重新登录...")

            # 如果是 headless 模式，session 过期后需要切换到可见模式才能手动登录
            if self.headless:
                self.log("检测到 headless 模式 + session 过期...")
                self.log("切换到可见模式以便完成验证码登录...")
                driver.quit()
                time.sleep(1)
                # 重新创建可见的 driver
                self.headless = False
                driver = self._build_driver()
                self.driver = driver
                self.log("已切换到可见模式，请在浏览器中完成登录")

            ok = self._do_login(driver)
            if not ok:
                self.log("登录失败")
                return None

            self._logged_in = True
            self._save_session(driver)

            # Step 2: 在 home 页点击"拓客·招投标版"进入 bidding
            self.log('点击"拓客·招投标版"进入 bidding...')
            driver.get("https://user.tungee.com/home")
            time.sleep(3)

            bidding_clicked = False
            for l in driver.find_elements(By.TAG_NAME, 'a'):
                try:
                    if '招投标' in l.text and l.is_displayed():
                        driver.execute_script("arguments[0].click();", l)
                        self.log(f"已点击: {l.text.strip()}")
                        bidding_clicked = True
                        time.sleep(5)
                        break
                except Exception:
                    continue

            if not bidding_clicked:
                # 直接尝试 URL
                self.log("未找到招投标链接，直接访问 bidding...")
                driver.get("https://bidding.tungee.com/")
                time.sleep(4)

            # Step 3: 导航到 search-enterprise/home
            driver.get(TUNGEE_SEARCH_URL)
            time.sleep(4)
            self._dismiss_popup(driver)
            time.sleep(2)

            # 最终验证
            body_final = driver.execute_script("return document.body.innerText;") or ""
            if '请输入企业名' in body_final or '查询一下' in body_final:
                self.log(f"搜索页加载成功: {driver.current_url}")
                return driver
            elif '请输入手机号码' in body_final or '账号登录' in body_final:
                self.log("仍在登录页，登录未成功")
                return None
            else:
                self.log(f"搜索页就绪（URL: {driver.current_url}）")
                return driver

        except WebDriverException as e:
            err = str(e).lower()
            if 'cannot connect' in err or 'chrome not reachable' in err:
                self.log("Chrome 连接失败")
            else:
                self.log(f"WebDriver 错误: {e}")
            return None
        except Exception as e:
            self.log(f"登录异常: {e}")
            traceback.print_exc()
            return None

    def _dismiss_popup(self, driver: webdriver.Chrome) -> None:
        """关闭强制下线等弹窗"""
        try:
            body = driver.execute_script("return document.body.innerText;")
            if '另一地点登录' in body or '已被迫下线' in body:
                self.log("检测到强制下线弹窗，关闭...")
                # 尝试多种方式关闭
                for selector in [
                    "//button[contains(text(),'我知道了')]",
                    "//button[contains(text(),'确定')]",
                    ".ant-modal-close",
                    "button.sentinelClose",
                ]:
                    try:
                        els = driver.find_elements(By.XPATH, selector)
                        for el in els:
                            if el.is_displayed():
                                driver.execute_script("arguments[0].click();", el)
                                time.sleep(1)
                                break
                    except Exception:
                        continue
                time.sleep(2)
        except Exception:
            pass

    def crawl_by_date(self, start_date: str, end_date: str) -> list[dict]:
        """抓取探迹高级筛选结果页"""
        all_records = []

        # 复用已有 driver，避免重复创建 Chrome 实例
        driver = self.driver
        if driver is None:
            driver = self.ensure_logged_in()
        if not driver:
            self.log("登录失败，退出")
            return []

        try:
            # 1. 导航到高级筛选页
            self.log("导航到高级筛选页...")
            driver.get(TUNGEE_FILTER_URL)
            time.sleep(4)

            # 检查 URL 是否包含高级筛选路径
            if "advanced-filter" not in driver.current_url:
                self.log("URL 不在高级筛选页，尝试从当前页导航...")
                # 从投标平台首页尝试
                for url in [
                    "https://bidding.tungee.com/",
                    "https://bidding.tungee.com/customer-seeking/advanced-filter/enterprise",
                ]:
                    driver.get(url)
                    time.sleep(3)
                    if "advanced-filter" in driver.current_url:
                        break

            if "advanced-filter" not in driver.current_url:
                self.log(f"无法到达高级筛选页，当前 URL: {driver.current_url}")
                return []

            self.log(f"已进入高级筛选页: {driver.current_url}")

            # 2. 等待结果加载（与 tanji 一致的选择器）
            self._wait_for_table(driver, timeout=20)

            # 3. 尝试设置日期筛选
            date_set = self._set_date_filter(driver, start_date, end_date)
            if not date_set:
                self.log("日期筛选控件未找到，将抓取全量")
            else:
                self.log("日期筛选已应用")

            # 3.5 尝试设置关键词筛选（页面级）
            if self.filter_keywords:
                kw_str = " ".join(self.filter_keywords[:3])  # 前3个关键词合并搜索
                kw_set = self._set_keyword_filter(driver, kw_str)
                if kw_set:
                    self.log(f"页面关键词筛选: {self.filter_keywords}")
                else:
                    self.log("页面关键词筛选未成功，将启用客户端关键词过滤")

            # 4. 翻页抓取
            page = 1
            max_pages = 100
            while page <= max_pages:
                if self.stop_event and self.stop_event.is_set():
                    self.log("收到停止信号")
                    break
                self.log(f"正在抓取第 {page} 页...")

                # 等待表格加载（与 tanji 一致）
                self._wait_for_table(driver, timeout=15)

                # 获取当前页所有行（与 tanji 一致的选择器）
                rows = driver.find_elements(By.CSS_SELECTOR, ".ant-table-tbody > tr")
                visible_rows = []
                for row in rows:
                    try:
                        if row.is_displayed():
                            visible_rows.append(row)
                    except Exception:
                        continue
                rows = visible_rows

                if not rows:
                    # 检查是否有"无数据"提示
                    placeholders = driver.find_elements(By.CSS_SELECTOR, ".ant-table-placeholder")
                    empty_text = placeholders[0].text if placeholders else ""
                    if "暂无数据" in empty_text or not empty_text:
                        self.log("当前页无数据，停止")
                        break
                    time.sleep(2)
                    rows = driver.find_elements(By.CSS_SELECTOR, ".ant-table-tbody > tr")

                page_records = 0
                for row in rows:
                    try:
                        cells = row.find_elements(By.CSS_SELECTOR, "td")
                        if len(cells) < 2:
                            continue

                        # 解析公司名（与 tanji 一致）
                        company_name = ""
                        for cell in cells[1:]:
                            text = clean_text(cell.text)
                            if not text:
                                continue
                            for line in cell.text.splitlines():
                                line = clean_text(line)
                                if not line or line in {"查看", "详情", "更多"}:
                                    continue
                                if line.isdigit():
                                    continue
                                company_name = line
                                break
                            if company_name:
                                break

                        if not company_name:
                            company_name = clean_text(cells[0].text)

                        if not any(k in company_name for k in ['有限', '集团', '股份', '科技', '公司']):
                            continue

                        # 客户端关键词过滤：行文本必须包含至少一个光伏关键词
                        if self.filter_keywords:
                            row_text = clean_text(row.text)
                            if not any(kw in row_text for kw in self.filter_keywords):
                                continue

                        # 提取手机号
                        phone = extract_phone(clean_text(row.text))

                        # 提取地址
                        address = ""
                        for cell in cells:
                            ct = clean_text(cell.text)
                            if any(k in ct for k in ['街道', '镇', '区', '路', '号', '园', '楼', '栋']):
                                address = ct
                                break

                        rec = {
                            'company_name': company_name,
                            'project_unit': '',
                            'contact_person': '',
                            'phone': phone,
                            'address': address,
                            'main_company': '',
                            'main_contact': '',
                            'main_phone': '',
                            'main_address': '',
                            'investor': '',
                            'project_desc': '',
                            'project_status': '',
                            'location': extract_sz_district(address),
                            'credit_code': '',
                            'relation_graph': '',
                            'remark': '',
                            'summary': '',
                            'source': '探迹',
                            'record_date': start_date,
                        }
                        all_records.append(rec)
                        page_records += 1

                    except Exception:
                        continue

                self.log(f"第 {page} 页: {page_records} 条，累计 {len(all_records)} 条")

                # 5. 翻页（与 tanji 一致的选择器）
                next_btns = driver.find_elements(
                    By.CSS_SELECTOR, "li.ant-pagination-next:not(.ant-pagination-disabled)")
                if not next_btns:
                    self.log("没有下一页按钮，停止")
                    break

                # 记录当前页第一条签名
                prev_first = ""
                try:
                    first_row = driver.find_element(By.CSS_SELECTOR, ".ant-table-tbody > tr")
                    prev_first = clean_text(first_row.text)
                except Exception:
                    pass

                # 点击下一页
                try:
                    next_btn = next_btns[0]
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", next_btn)
                    time.sleep(0.3)
                    driver.execute_script("arguments[0].click();", next_btn)
                except Exception:
                    try:
                        next_btn.click()
                    except Exception:
                        self.log("点击下一页失败，停止")
                        break

                # 等待页码变化
                time.sleep(3)
                max_wait = 20
                waited = 0
                while waited < max_wait:
                    try:
                        new_first = ""
                        try:
                            new_row = driver.find_element(By.CSS_SELECTOR, ".ant-table-tbody > tr")
                            new_first = clean_text(new_row.text)
                        except Exception:
                            pass
                        if new_first and new_first != prev_first:
                            break
                    except Exception:
                        pass
                    time.sleep(1)
                    waited += 1

                if waited >= max_wait:
                    self.log("翻页等待超时，停止")
                    break

                page += 1
                time.sleep(self.interval)

        except Exception as e:
            self.log(f"爬取异常: {e}")
            traceback.print_exc()

        finally:
            self.quit()

        self.log(f"共抓取 {len(all_records)} 条")
        return all_records

    def crawl_by_keywords(self, company_names: list[str]) -> list[dict]:
        """按公司名列表在探迹普通搜索页查询，补全字段（性能优化版）"""
        all_records = []

        # 复用已有 driver
        driver = self.driver
        if driver is None:
            driver = self.ensure_logged_in()
        if not driver:
            self.log("登录失败，退出")
            return []

        try:
            first_search = True
            for idx, company_name in enumerate(company_names):
                if self.stop_event and self.stop_event.is_set():
                    self.log("收到停止信号，退出")
                    break

                self.log(f"[{idx+1}/{len(company_names)}] 搜索: {company_name}")

                try:
                    # 1. 导航到搜索页（首次完整加载，后续快速刷新）
                    wait_time = 2 if first_search else 1
                    driver.get(TUNGEE_SEARCH_URL)
                    if first_search:
                        first_search = False
                    time.sleep(wait_time)

                    # 2. 等待搜索框出现
                    try:
                        search_input = WebDriverWait(driver, 8).until(
                            lambda d: d.find_element(By.CSS_SELECTOR,
                                "input[placeholder*='请输入企业名']"))
                    except Exception:
                        self.log(f"搜索框未找到，跳过: {company_name}")
                        continue

                    # 3. 快速清空并填写关键词
                    search_input.click()
                    driver.execute_script(
                        "arguments[0].value=''; arguments[0].dispatchEvent(new Event('input', {bubbles: true}));",
                        search_input)
                    search_input.send_keys(company_name)
                    search_input.send_keys(Keys.RETURN)
                    time.sleep(1.5)

                    # 4. 等待结果加载
                    try:
                        WebDriverWait(driver, 8).until(
                            lambda d: d.find_elements(By.CSS_SELECTOR, "a[href*='/enterprise-details/']"))
                    except Exception:
                        pass

                    # 5. 解析结果列表，找到匹配公司名的链接
                    result_links = driver.find_elements(
                        By.CSS_SELECTOR, "a[href*='/enterprise-details/']")
                    if not result_links:
                        self.log(f"  未找到结果: {company_name}")
                        time.sleep(0.5)
                        continue

                    # 精确匹配
                    matched_link = None
                    for link in result_links:
                        try:
                            link_text = link.text.strip()
                            if link_text and company_name[:6] in link_text:
                                matched_link = link
                                break
                        except Exception:
                            continue

                    first_link = matched_link or result_links[0]
                    detail_url = first_link.get_attribute('href')
                    self.log(f"  进入详情: {detail_url[-40:]}")
                    driver.get(detail_url)

                    # 6. 等待详情页加载
                    time.sleep(2)
                    try:
                        WebDriverWait(driver, 8).until(
                            lambda d: '企业信息' in (d.execute_script("return document.body.innerText;") or ''))
                    except Exception:
                        pass

                    # 7. 解锁CRM获取联系方式（付费会员）
                    self._unlock_crm(driver)

                    # 8. 解析详情页
                    rec = self._parse_detail_page(driver, company_name)
                    if rec:
                        all_records.append(rec)
                        self.log(f"  完成: {rec.get('company_name', company_name)}")

                    time.sleep(0.5)  # 简短的节流

                except Exception as e:
                    self.log(f"  处理出错: {company_name} - {e}")
                    continue

        except Exception as e:
            self.log(f"爬取异常: {e}")
            traceback.print_exc()

        finally:
            self.quit()

        self.log(f"关键词搜索完成，共获取 {len(all_records)} 条")
        return all_records

    def _unlock_crm(self, driver: webdriver.Chrome) -> bool:
        """解锁CRM查看联系方式

        基于实际探索发现的稳定选择器：
        - 解锁按钮: //button[contains(text(),'解锁')]
        - 解锁后出现Tab: "公司联系方式" / "招投标联系方式"
        """
        try:
            body_text = driver.execute_script("return document.body.innerText;") or ""

            # 如果已经有手机号可见，跳过
            if re.search(r'1[3-9]\d{9}', body_text):
                self.log("    联系方式已可见，跳过解锁")
                return True

            # 方法1: XPath文本定位找"解锁"按钮（最稳定）
            unlock_xpaths = [
                "//button[contains(text(),'解锁')]",
                "//button[contains(.,'解锁')]",
            ]
            clicked = False
            for xpath in unlock_xpaths:
                try:
                    els = driver.find_elements(By.XPATH, xpath)
                    for el in els:
                        if el.is_displayed() and el.is_enabled():
                            driver.execute_script("arguments[0].click();", el)
                            self.log("    已点击解锁按钮")
                            clicked = True
                            time.sleep(2)
                            break
                except Exception:
                    pass
                if clicked:
                    break

            # 方法2: Ant Design主按钮（兜底）
            if not clicked:
                try:
                    buttons = driver.find_elements(By.CSS_SELECTOR, "button.ant-btn-primary")
                    for btn in buttons:
                        if '解' in (btn.text or '') and btn.is_displayed():
                            driver.execute_script("arguments[0].click();", btn)
                            self.log("    已点击主要按钮解锁")
                            clicked = True
                            time.sleep(2)
                            break
                except Exception:
                    pass

            if not clicked:
                self.log("    未找到解锁元素，可能已解锁或账号权限不足")
                return False

            # 点击后，切换到"公司联系方式"Tab
            try:
                contact_tabs = driver.find_elements(By.XPATH, "//*[contains(text(),'公司联系方式')]")
                for tab in contact_tabs:
                    if tab.is_displayed():
                        driver.execute_script("arguments[0].click();", tab)
                        self.log("    已切换到公司联系方式")
                        time.sleep(2)
                        break
            except Exception:
                pass

            # 等待手机号出现
            try:
                WebDriverWait(driver, 10).until(
                    lambda d: bool(re.search(r'1[3-9]\d{9}',
                        d.execute_script("return document.body.innerText;") or ''))
                )
                self.log("    联系方式已解锁")
                return True
            except Exception:
                self.log("    解锁后未检测到联系方式")
                return False

        except Exception as e:
            self.log(f"    解锁CRM异常: {e}")
            return False

    def _parse_detail_page(self, driver: webdriver.Chrome, company_name: str) -> Optional[dict]:
        """解析探迹企业详情页，提取联系人和主体公司等信息

        基于实际探索发现的稳定选择器：
        - 联系方式Tab: //*[contains(text(),'公司联系方式')]
        - 关系图谱Tab: //*[contains(text(),'企业图谱')]
        - 手机号: 正则扫描innerText
        """
        try:
            rec = {
                'company_name': company_name,
                'project_unit': '',
                'contact_person': '',
                'phone': '',
                'address': '',
                'main_company': '',
                'main_contact': '',
                'main_phone': '',
                'main_address': '',
                'investor': '',
                'project_desc': '',
                'project_status': '',
                'location': '',
                'credit_code': '',
                'relation_graph': '',
                'remark': '',
                'summary': '',
                'source': '探迹',
                'record_date': '',
            }

            # 等待页面加载 + 滚动触发懒加载
            time.sleep(1)
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(0.5)
            driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(0.3)

            # 保存诊断HTML
            try:
                debug_dir = ROOT_DIR / "debug_output"
                debug_dir.mkdir(exist_ok=True)
                static_file = debug_dir / "tungee_detail_page.html"
                if not static_file.exists():
                    with open(static_file, 'w', encoding='utf-8') as f:
                        f.write(driver.page_source)
                    self.log(f"[诊断] 探迹详情页HTML已保存")
            except Exception:
                pass

            # ========== 1. 切换到"公司联系方式"Tab并提取手机号 ==========
            try:
                # Tab文字在 <i class="tgt-capitalize-first-letter">公司联系方式</i>
                contact_tab = driver.find_element(By.XPATH, "//i[contains(text(),'公司联系方式')]")
                driver.execute_script("arguments[0].parentNode.parentNode.click();", contact_tab)
                time.sleep(2)
            except Exception:
                pass

            body_text = driver.execute_script("return document.body.innerText;") or ""

            # 提取所有手机号
            all_phones = re.findall(r'1[3-9]\d{9}', body_text)
            valid_phones = [p for p in all_phones if not p.startswith(('400', '800'))]
            if valid_phones:
                rec['phone'] = valid_phones[0]
                rec['main_phone'] = valid_phones[1] if len(valid_phones) > 1 else ''

            # ========== 2. 联系人 ==========
            # 从正文正则匹配
            for pat in [
                r'联系人[：:]\s*([^\s\n]{2,15})',
                r'企业联系人[：:]\s*([^\s\n]{2,15})',
                r'负责人[：:]\s*([^\s\n]{2,15})',
            ]:
                m = re.search(pat, body_text)
                if m and m.group(1).strip():
                    rec['contact_person'] = m.group(1).strip()
                    break

            # 兜底法定代表人
            if not rec['contact_person']:
                for pat in [
                    r'法定代表人[：:]\s*([^\s\n]{2,15})',
                    r'法人代表[：:]\s*([^\s\n]{2,15})',
                    r'法人[：:]\s*([^\s\n]{2,15})',
                ]:
                    m = re.search(pat, body_text)
                    if m and m.group(1).strip():
                        rec['contact_person'] = m.group(1).strip()
                        break

            # ========== 3. 地址 ==========
            for pat in [r'地址[：:]\s*([^\n]{5,80})', r'所在地[：:]\s*([^\n]{5,80})']:
                m = re.search(pat, body_text)
                if m:
                    addr = clean_text(m.group(1))
                    if addr and len(addr) > 5:
                        rec['address'] = addr
                        rec['location'] = extract_sz_district(addr)
                        break

            # ========== 4. 关系图谱 - 切换到"企业图谱"Tab ==========
            try:
                # 方式1: 直接找包含"企业图谱"文字的tab div
                tab_divs = driver.find_elements(By.XPATH,
                    "//div[@role='tab' and .//i[contains(text(),'企业图谱')]]")
                if tab_divs:
                    driver.execute_script("arguments[0].click();", tab_divs[0])
                    self.log("    已点击企业图谱Tab")
                    time.sleep(3)

                # 等待表格加载
                WebDriverWait(driver, 10).until(
                    lambda d: d.find_elements(By.CSS_SELECTOR, "table.ant-table"))
                time.sleep(1)

                # 提取股东信息表格 - 方式1: 找包含"股东"列的表格
                tables = driver.find_elements(By.CSS_SELECTOR, "table.ant-table")
                shareholders = []
                for table in tables:
                    try:
                        # 检查表头是否包含"股东"
                        ths = table.find_elements(By.CSS_SELECTOR, "thead th")
                        has_shareholder = False
                        for th in ths:
                            th_text = th.text.strip()
                            th_title = th.get_attribute('title') or ''
                            if '股东' in th_text or '股东' in th_title:
                                has_shareholder = True
                                break
                        if has_shareholder:
                            # 提取股东名（表格第二列）
                            rows = table.find_elements(By.CSS_SELECTOR, "tbody tr")
                            for row in rows:
                                cells = row.find_elements(By.TAG_NAME, "td")
                                if len(cells) >= 2:
                                    name = cells[1].text.strip()
                                    # 过滤：排除纯数字、纯符号、太短的
                                    if name and len(name) >= 2 and not name.isdigit() and name not in ['-', '/', '\\']:
                                        shareholders.append(name)
                    except Exception:
                        continue

                if shareholders:
                    rec['relation_graph'] = ' | '.join(shareholders[:10])
                    self.log(f"    关系图谱提取到 {len(shareholders)} 个股东")
                else:
                    self.log("    未找到股东信息表")
            except Exception as e:
                self.log(f"    关系图谱提取异常: {e}")

            # ========== 5. 主体公司 ==========
            for pat in [r'主体公司[：:]\s*([^\s\n]{4,30})', r'所属企业[：:]\s*([^\s\n]{4,30})']:
                m = re.search(pat, body_text)
                if m:
                    rec['main_company'] = clean_text(m.group(1))
                    break

            # ========== 6. 信用代码 ==========
            codes = re.findall(r'[0-9A-Z]{18}', body_text)
            for code in codes:
                if code[:2] in ('91', '92', '93'):
                    rec['credit_code'] = code
                    break

            # ========== 7. 投资人/股东 ==========
            for pat in [r'股东[：:]\s*([^\s\n]{2,30})', r'投资人[：:]\s*([^\s\n]{2,30})']:
                m = re.search(pat, body_text)
                if m:
                    rec['investor'] = clean_text(m.group(1))
                    break

            # ========== 8. 备注/摘要 ==========
            if not rec.get('phone') and not rec.get('contact_person'):
                rec['remark'] = '联系方式未提取到'

            email_match = re.search(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', body_text)
            if email_match:
                rec['remark'] = (rec.get('remark', '') + f" 邮箱:{email_match.group(0)}").strip()

            if not rec.get('summary'):
                rec['summary'] = clean_text(body_text)[:200]

            return rec

        except Exception as e:
            self.log(f"解析详情页出错: {e}")
            traceback.print_exc()
            return None

    def _wait_for_table(self, driver: webdriver.Chrome, timeout: int = 15) -> None:
        """等待表格加载（与 tanji 一致的选择器）"""
        try:
            WebDriverWait(driver, timeout).until(
                lambda d: d.find_elements(By.CSS_SELECTOR, ".ant-table-wrapper"))
        except Exception:
            pass
        time.sleep(1)
        try:
            WebDriverWait(driver, timeout).until(
                lambda d: bool(d.find_elements(By.CSS_SELECTOR, ".ant-table-tbody > tr")) or
                          bool(d.find_elements(By.CSS_SELECTOR, ".ant-table-placeholder")))
        except Exception:
            pass

    def _set_date_filter(self, driver: webdriver.Chrome,
                         start_date: str, end_date: str) -> bool:
        """设置日期筛选"""
        try:
            # 方法1: 查找日期输入框 (Ant Design RangePicker)
            date_selectors = [
                "input[placeholder*='开始日期']",
                "input[placeholder*='结束日期']",
                "input[placeholder*='日期']",
                "input[placeholder*='时间']",
                ".ant-picker input",
                ".ant-calendar-picker input",
                ".date-picker input",
            ]

            date_inputs = []
            for sel in date_selectors:
                try:
                    inputs = driver.find_elements(By.CSS_SELECTOR, sel)
                    for inp in inputs:
                        if inp.is_displayed() and inp.is_enabled():
                            date_inputs.append(inp)
                except Exception:
                    continue

            # 方法2: 找所有可见的 input，排除搜索框
            if not date_inputs:
                all_inputs = driver.find_elements(By.CSS_SELECTOR, "input:not([type='hidden'])")
                for inp in all_inputs:
                    try:
                        if not inp.is_displayed():
                            continue
                        placeholder = (inp.get_attribute('placeholder') or '').lower()
                        if any(kw in placeholder for kw in ['日期', '时间', 'date', '开始', '结束', 'start', 'end']):
                            date_inputs.append(inp)
                    except Exception:
                        continue

            if not date_inputs:
                # 保存截图用于诊断
                try:
                    driver.save_screenshot('debug_output/tungee_filter_debug.png')
                    self.log("日期控件未找到，已保存截图: debug_output/tungee_filter_debug.png")
                except Exception:
                    self.log("未找到日期控件")
                return False

            self.log(f"找到 {len(date_inputs)} 个日期输入框")

            # 填入开始日期
            date_inputs[0].click()
            time.sleep(0.5)
            active = driver.switch_to.active_element
            active.send_keys(Keys.CONTROL, 'a')
            active.send_keys(Keys.DELETE)
            active.send_keys(start_date)
            time.sleep(0.3)
            active.send_keys(Keys.ENTER)
            time.sleep(0.5)

            # 填入结束日期
            if len(date_inputs) > 1:
                date_inputs[1].click()
                time.sleep(0.5)
                active = driver.switch_to.active_element
                active.send_keys(Keys.CONTROL, 'a')
                active.send_keys(Keys.DELETE)
                active.send_keys(end_date)
                active.send_keys(Keys.ENTER)
                time.sleep(0.5)

            # 点击查询/筛选按钮
            for btn_text in ['查询', '筛选', '搜 索', '搜索', '确 定', '确定']:
                try:
                    btns = driver.find_elements(By.XPATH,
                        f"//button[contains(normalize-space(),'{btn_text}')] | //span[contains(text(),'{btn_text}')]/parent::button")
                    for btn in btns:
                        if btn.is_displayed() and btn.is_enabled():
                            driver.execute_script("arguments[0].click();", btn)
                            time.sleep(2)
                            self.log(f"日期筛选已应用（点击了'{btn_text}'）")
                            return True
                except Exception:
                    continue

            # 兜底：按回车触发查询
            try:
                active = driver.switch_to.active_element
                active.send_keys(Keys.ENTER)
                time.sleep(2)
                self.log("按回车触发查询")
                return True
            except Exception:
                pass

            return False

        except Exception as e:
            self.log(f"设置日期失败: {e}")
            return False

    def _set_keyword_filter(self, driver: webdriver.Chrome, keyword: str) -> bool:
        """在高级筛选页设置关键词筛选（项目类型/关键词输入框）"""
        try:
            # 尝试多种选择器定位关键词输入框
            keyword_selectors = [
                "input[placeholder*='关键词']",
                "input[placeholder*='项目类型']",
                "input[placeholder*='项目名称']",
                "input[placeholder*='请输入']",
                "input.ant-input[type='text']:not([readonly])",
            ]

            kw_input = None
            for sel in keyword_selectors:
                try:
                    inputs = driver.find_elements(By.CSS_SELECTOR, sel)
                    for inp in inputs:
                        if inp.is_displayed() and inp.is_enabled():
                            # 跳过日期输入框
                            placeholder = inp.get_attribute('placeholder') or ''
                            if any(t in placeholder for t in ['日期', '时间', '年', '月']):
                                continue
                            kw_input = inp
                            break
                    if kw_input:
                        break
                except Exception:
                    continue

            if not kw_input:
                self.log("未找到关键词输入框，跳过页面级关键词筛选")
                return False

            # 清空并填入关键词
            kw_input.click()
            time.sleep(0.3)
            kw_input.send_keys(Keys.CONTROL, 'a')
            kw_input.send_keys(Keys.DELETE)
            time.sleep(0.1)
            kw_input.send_keys(keyword)
            time.sleep(0.3)
            kw_input.send_keys(Keys.RETURN)
            time.sleep(2)

            self.log(f"页面关键词筛选已设置: {keyword}")
            return True

        except Exception as e:
            self.log(f"设置关键词筛选失败: {e}")
            return False

    def quit(self):
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = None


# ============================================================
# Excel 导出
# ============================================================
def export_to_excel(records: list, filepath: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "项目信息"

    hdr_font = Font(name='Microsoft YaHei', bold=True, size=10, color='FFFFFF')
    hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(vertical='top', wrap_text=True)

    for col, h in enumerate(EXCEL_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align

    # 数据映射（17列 ←→ EXCEL_HEADERS 顺序完全对应）
    FIELD_MAP = [
        'company_name',   # 1. 项目信息
        'project_unit',   # 2. 项目单位          ← 修正：之前错误用 company_name
        'contact_person', # 3. 项目公司联系人
        'phone',          # 4. 手机号
        'address',        # 5. 地址
        'main_company',   # 6. 主体公司
        'main_contact',  # 7. 主体公司联系人
        'main_phone',    # 8. 手机号            ← 修正：之前漏了 main_phone
        'main_address',   # 9. 主体公司地址
        'relation_graph', # 10. 关系图谱         ← 修正：之前为空
        'investor',      # 11. 投资人
        'project_desc',   # 12. 项目情况
        'project_status', # 13. 项目进展
        'location',      # 14. 项目所在地
        'credit_code',    # 15. 项目备案编号
        'remark',         # 16. 备注             ← 修正：之前为空
        'summary',        # 17. 项目总结         ← 修正：之前为空
    ]

    for row_idx, rec in enumerate(records, 2):
        for col_idx, field in enumerate(FIELD_MAP, 1):
            val = rec.get(field, '') if field else ''
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.alignment = cell_align

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 42
    ws.column_dimensions['F'].width = 28
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 42
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 16
    ws.column_dimensions['L'].width = 45
    ws.column_dimensions['M'].width = 30
    ws.column_dimensions['N'].width = 10
    ws.column_dimensions['O'].width = 28
    ws.column_dimensions['P'].width = 20
    ws.column_dimensions['Q'].width = 30
    ws.freeze_panes = 'A2'

    wb.save(filepath)
    return filepath


# ============================================================
# GUI
# ============================================================
class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("新能源项目智能爬虫 v2.0")
        self.root.geometry("960x740")
        self.root.minsize(860, 640)

        self.running = False
        self.stop_event = threading.Event()

        # 配置变量
        self.cfg_phone = StringVar(value=APP_CONFIG.get('tungee_phone',''))
        self.cfg_password = StringVar(value=APP_CONFIG.get('tungee_password',''))
        self.cfg_export_dir = StringVar(value=APP_CONFIG.get('export_dir', str(Path.home() / "Desktop")))
        self.cfg_interval = StringVar(value=APP_CONFIG.get('request_interval','3'))
        self.cfg_chrome_path = StringVar(value=APP_CONFIG.get('chrome_path',''))
        self.cfg_debug_port = StringVar(value=APP_CONFIG.get('debug_port','9222'))
        self.cfg_headless = StringVar(value=APP_CONFIG.get('headless','0'))
        self.cfg_filter_keywords = StringVar(value=APP_CONFIG.get('filter_keywords',
            '车棚 超充站 光储充 分布式光伏 屋顶分布式光伏 光伏EPC 充电站 储能 光伏发电 光伏电站'))

        # 日期
        self.date_start = StringVar(value=(datetime.now()-timedelta(days=7)).strftime('%Y-%m-%d'))
        self.date_end = StringVar(value=datetime.now().strftime('%Y-%m-%d'))

        # 数据源
        self.var_sz = BooleanVar(value=True)
        self.var_tungee = BooleanVar(value=True)
        self.var_tianyancha = BooleanVar(value=True)  # 天眼查补全

        self.log_queue: queue.Queue[str] = queue.Queue()

        self._build_ui()
        self.root.after(200, self._poll_log)
        init_db()
        self._refresh_stats()

    def _build_ui(self):
        main = ttk.Frame(self.root)
        main.pack(fill='both', expand=True, padx=10, pady=10)

        # === 左侧面板 ===
        left = ttk.LabelFrame(main, text=" 配置 ", padding=10)
        left.pack(side='left', fill='y', padx=(0, 10))

        # 日期
        df = ttk.LabelFrame(left, text=" 日期范围 ", padding=8)
        df.pack(fill='x', pady=(0, 8))

        ttk.Label(df, text="开始:").grid(row=0, column=0, sticky='w', pady=3)
        ttk.Entry(df, textvariable=self.date_start, width=13).grid(row=0, column=1, padx=5, pady=3)
        ttk.Label(df, text="截止:").grid(row=1, column=0, sticky='w', pady=3)
        ttk.Entry(df, textvariable=self.date_end, width=13).grid(row=1, column=1, padx=5, pady=3)

        btnrow = ttk.Frame(df)
        btnrow.grid(row=2, column=0, columnspan=2, pady=(4,0), sticky='w')
        for label, days in [("今",0),("3天",3),("7天",7),("30天",30)]:
            ttk.Button(btnrow, text=label, width=4,
                command=lambda d=days: self._set_dates(d)
            ).pack(side='left', padx=2)

        # 数据源
        sf = ttk.LabelFrame(left, text=" 数据源 ", padding=8)
        sf.pack(fill='x', pady=(0, 8))
        ttk.Checkbutton(sf, text="☑ 深圳投资网 (wsbs.sz.gov.cn)", variable=self.var_sz).pack(anchor='w', pady=2)
        ttk.Checkbutton(sf, text="☑ 天眼查补全 (tianyancha.com)", variable=self.var_tianyancha).pack(anchor='w', pady=2)
        ttk.Checkbutton(sf, text="☑ 探迹补全 (tungee.com) [需账号]", variable=self.var_tungee).pack(anchor='w', pady=2)

        # 账号
        af = ttk.LabelFrame(left, text=" 探迹账号 ", padding=8)
        af.pack(fill='x', pady=(0, 8))
        ttk.Label(af, text="手机:").grid(row=0, column=0, sticky='w', pady=3)
        ttk.Entry(af, textvariable=self.cfg_phone, width=15).grid(row=0, column=1, padx=5, pady=3)
        ttk.Label(af, text="密码:").grid(row=1, column=0, sticky='w', pady=3)
        ttk.Entry(af, textvariable=self.cfg_password, show='*', width=15).grid(row=1, column=1, padx=5, pady=3)

        # 设置
        gf = ttk.LabelFrame(left, text=" 爬虫设置 ", padding=8)
        gf.pack(fill='x', pady=(0, 8))
        ttk.Label(gf, text="请求间隔(秒):").grid(row=0, column=0, sticky='w', pady=3)
        ttk.Entry(gf, textvariable=self.cfg_interval, width=8).grid(row=0, column=1, sticky='w', padx=5, pady=3)

        # 无头模式
        ttk.Label(gf, text="无头模式:").grid(row=1, column=0, sticky='w', pady=3)
        headless_frame = ttk.Frame(gf)
        headless_frame.grid(row=1, column=1, sticky='w', padx=5, pady=3)
        ttk.Radiobutton(headless_frame, text="关闭", variable=self.cfg_headless, value="0").pack(side='left', padx=(0, 5))
        ttk.Radiobutton(headless_frame, text="开启", variable=self.cfg_headless, value="1").pack(side='left')

        ttk.Label(gf, text="导出目录:").grid(row=2, column=0, sticky='nw', pady=3)
        dirrow = ttk.Frame(gf)
        dirrow.grid(row=2, column=1, sticky='w', padx=5, pady=3)
        ttk.Entry(dirrow, textvariable=self.cfg_export_dir, width=18).pack(side='left')
        ttk.Button(dirrow, text="…", command=self._browse_dir, width=2).pack(side='left', padx=2)

        # 关键词过滤
        kf = ttk.LabelFrame(left, text=" 关键词过滤（探迹） ", padding=8)
        kf.pack(fill='x', pady=(0, 8))
        ttk.Label(kf, text="仅抓取包含以下关键词的项目（空格分隔）:").pack(anchor='w')
        self.kw_entry = ttk.Entry(kf, textvariable=self.cfg_filter_keywords, width=30)
        self.kw_entry.pack(fill='x', pady=(5, 0))
        ttk.Label(kf, text="留空 = 不过滤，抓取全部项目", font=('', 8), foreground='gray').pack(anchor='w')

        # 控制按钮
        ctrl = ttk.Frame(left)
        ctrl.pack(fill='x')

        self.btn_start = ttk.Button(ctrl, text="▶ 开始爬取", command=self._on_start, width=12)
        self.btn_start.pack(fill='x', pady=(0, 4))

        self.btn_stop = ttk.Button(ctrl, text="⏹ 停止", command=self._on_stop, width=12, state=DISABLED)
        self.btn_stop.pack(fill='x', pady=(0, 4))

        ttk.Button(ctrl, text="💾 保存配置", command=self._save_cfg, width=12).pack(fill='x', pady=(0, 4))
        ttk.Button(ctrl, text="📂 导出Excel", command=self._on_export, width=12).pack(fill='x', pady=(0, 4))
        ttk.Button(ctrl, text="🔄 刷新统计", command=self._refresh_stats, width=12).pack(fill='x')

        # === 右侧面板 ===
        right = ttk.Frame(main)
        right.pack(side='left', fill='both', expand=True)

        # 统计
        stat = ttk.LabelFrame(right, text=" 数据统计 ", padding=8)
        stat.pack(fill='x', pady=(0, 8))

        self.lbl_stat = ttk.Label(stat, text="总记录: 0  |  深圳: 0  |  探迹: 0", font=('', 10))
        self.lbl_stat.pack(anchor='w')

        # 日志
        logf = ttk.LabelFrame(right, text=" 控制台 ", padding=8)
        logf.pack(fill='both', expand=True)

        self.log_txt = ScrolledText(logf, height=30, font=('Consolas', 9),
                                    state='disabled', wrap='word')
        self.log_txt.pack(fill='both', expand=True)

        menu = tk.Menu(self.log_txt, tearoff=0)
        menu.add_command(label="清空", command=self._clear_log)
        menu.add_command(label="复制全部", command=self._copy_log)
        self.log_txt.bind("<Button-3>", lambda e: menu.post(e.x_root, e.y_root))

        self._log("工具已就绪。请确保 Chrome 已开启调试模式（启动时加 --remote-debugging-port=9222）")
        self._log("深圳投资网: Selenium + API 混合模式")
        self._log("探迹: Selenium 自动登录模式")

    def _set_dates(self, days: int):
        end = datetime.now()
        start = end - timedelta(days=days)
        self.date_start.set(start.strftime('%Y-%m-%d'))
        self.date_end.set(end.strftime('%Y-%m-%d'))

    def _browse_dir(self):
        d = filedialog.askdirectory()
        if d:
            self.cfg_export_dir.set(d)

    def _save_cfg(self):
        save_config({
            "tungee_phone": self.cfg_phone.get(),
            "tungee_password": self.cfg_password.get(),
            "export_dir": self.cfg_export_dir.get(),
            "request_interval": self.cfg_interval.get(),
            "chrome_path": self.cfg_chrome_path.get(),
            "debug_port": self.cfg_debug_port.get(),
            "headless": self.cfg_headless.get(),
            "filter_keywords": self.cfg_filter_keywords.get(),
        })
        self._log("配置已保存")

    def _poll_log(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()
                self.log_txt.configure(state='normal')
                self.log_txt.insert(END, msg + '\n')
                self.log_txt.see(END)
                self.log_txt.configure(state='disabled')
        except queue.Empty:
            pass
        self.root.after(200, self._poll_log)

    def _log(self, msg: str):
        ts = datetime.now().strftime('%H:%M:%S')
        self.log_queue.put(f"[{ts}] {msg}")

    def _clear_log(self):
        self.log_txt.configure(state='normal')
        self.log_txt.delete('1.0', END)
        self.log_txt.configure(state='disabled')

    def _copy_log(self):
        txt = self.log_txt.get('1.0', END)
        self.root.clipboard_clear()
        self.root.clipboard_write(txt)

    def _refresh_stats(self):
        try:
            conn = sqlite3.connect(str(DB_FILE))
            c = conn.cursor()
            c.execute("SELECT COUNT(*) FROM companies"); total = c.fetchone()[0] or 0
            c.execute("SELECT COUNT(*) FROM companies WHERE source='深圳投资网'")
            sz = c.fetchone()[0] or 0
            c.execute("SELECT COUNT(*) FROM companies WHERE source='探迹'")
            tg = c.fetchone()[0] or 0
            conn.close()
            self.lbl_stat.config(text=f"总记录: {total}  |  深圳投资网: {sz}  |  探迹: {tg}")
        except Exception:
            pass

    def _on_start(self):
        if self.running:
            return
        start = self.date_start.get().strip()
        end = self.date_end.get().strip()
        if not start or not end:
            messagebox.showwarning("提示", "请设置日期范围")
            return
        if not self.var_sz.get() and not self.var_tungee.get():
            messagebox.showwarning("提示", "请至少选择一个数据源")
            return

        self.running = True
        self.btn_start.config(state=DISABLED)
        self.btn_stop.config(state=NORMAL)
        self.stop_event.clear()

        t = threading.Thread(target=self._run, args=(start, end), daemon=True)
        t.start()

    def _on_stop(self):
        self.stop_event.set()
        self.running = False
        self.btn_start.config(state=NORMAL)
        self.btn_stop.config(state=DISABLED)
        self._log("已发送停止信号")

    def _on_export(self):
        fp = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialdir=self.cfg_export_dir.get(),
            initialfile=f"新能源项目_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if not fp:
            return
        try:
            rows = get_db_records()
            if not rows:
                messagebox.showinfo("提示", "数据库暂无记录")
                return
            # 注意：DB 列顺序 = id, company_name, project_unit, contact_person, phone,
            # address, main_company, main_contact, main_phone, main_address, investor,
            # project_desc, project_status, location, credit_code, relation_graph,
            # remark, summary, source, record_date, created_at
            # DB列: id(0), company_name(1), ... record_date(19), created_at(20)
            keys = ['company_name','project_unit','contact_person','phone','address',
                    'main_company','main_contact','main_phone','main_address',
                    'investor','project_desc','project_status','location','credit_code',
                    'relation_graph','remark','summary','source','record_date']
            recs = [dict(zip(keys, r[1:-1])) for r in rows]  # 跳过id和created_at
            export_to_excel(recs, fp)
            messagebox.showinfo("成功", f"已导出 {len(recs)} 条到:\n{fp}")
            self._log(f"导出成功: {fp} ({len(recs)} 条)")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))
            self._log(f"导出失败: {e}")

    def _run(self, start_date: str, end_date: str):
        interval = int(self.cfg_interval.get() or '3')
        headless = self.cfg_headless.get() == '1'

        # 解析关键词（空格/逗号分隔）
        raw_kw = self.cfg_filter_keywords.get().strip()
        filter_keywords = None
        if raw_kw:
            filter_keywords = [k.strip() for k in re.split(r'[\s,，]+', raw_kw) if k.strip()]
            self._log(f"探迹关键词过滤: {' | '.join(filter_keywords)}")

        # Phase 1: 深圳招采网 → 按日期+光伏关键词 → 项目列表
        if self.var_sz.get() and not self.stop_event.is_set():
            self._log(f"=== Phase 1: 深圳招采网: {start_date} ~ {end_date} ===")
            log_id = log_start('深圳投资网', start_date)
            try:
                crawler = SZInvestCrawler(
                    self.log_queue, interval=interval,
                    chrome_path=self.cfg_chrome_path.get() or None,
                    debug_port=int(self.cfg_debug_port.get() or 9222),
                    headless=headless,
                    filter_keywords=filter_keywords,
                    stop_event=self.stop_event,
                )
                try:
                    records = crawler.crawl_by_date(start_date, end_date)
                    if records:
                        insert_records(records)
                    log_finish(log_id, len(records), 'success')
                    self._log(f"深圳招采网完成: +{len(records)} 条")
                finally:
                    crawler.quit()
            except Exception as e:
                log_finish(log_id, 0, 'failed', str(e))
                self._log(f"深圳招采网出错: {e}")
        else:
            self._log("深圳招采网未勾选，跳过 Phase 1")

        # Phase 2: 从 DB 读取所有待补全的公司名
        company_names = []
        if (self.var_tianyancha.get() or self.var_tungee.get()) \
                and not self.stop_event.is_set():
            try:
                conn = sqlite3.connect(str(DB_FILE))
                c = conn.cursor()
                c.execute("""
                    SELECT DISTINCT company_name FROM companies
                    WHERE company_name != '' AND record_date >= ? AND record_date <= ?
                    ORDER BY id DESC
                """, (start_date, end_date))
                company_names = [row[0] for row in c.fetchall() if row[0]]
                conn.close()
                self._log(f"Phase 2: 读取到 {len(company_names)} 个待补全公司 ({start_date}~{end_date})")
            except Exception as e:
                self._log(f"读取公司名失败: {e}")

        # Phase 3-4: 天眼查 + 探迹 并行补全
        tianyancha_records = []
        tungee_records = []
        tianyancha_log_id = None
        tungee_log_id = None

        def run_tianyancha():
            nonlocal tianyancha_records, tianyancha_log_id
            tianyancha_log_id = log_start('天眼查补全', start_date)
            try:
                enricher = TianyanchaEnricher(
                    self.log_queue,
                    interval=max(1, interval // 2),
                    chrome_path=self.cfg_chrome_path.get() or None,
                    debug_port=int(self.cfg_debug_port.get() or 9222),
                    stop_event=self.stop_event,
                )
                try:
                    tianyancha_records = enricher.enrich_batch(company_names)
                    log_finish(tianyancha_log_id, len(tianyancha_records), 'success')
                    self._log(f"天眼查补全完成: {len(tianyancha_records)} 条")
                finally:
                    enricher.quit()
            except Exception as e:
                log_finish(tianyancha_log_id, 0, 'failed', str(e))
                self._log(f"天眼查出错: {e}")

        def run_tungee():
            nonlocal tungee_records, tungee_log_id
            phone = self.cfg_phone.get().strip()
            password = self.cfg_password.get().strip()
            if not phone or not password:
                self._log("探迹账号未填写，跳过补全")
                return
            tungee_log_id = log_start('探迹_关键词', start_date)
            try:
                crawler = TungeeCrawler(
                    self.log_queue, phone=phone, password=password,
                    interval=max(1, interval // 2),
                    chrome_path=self.cfg_chrome_path.get() or None,
                    debug_port=int(self.cfg_debug_port.get() or 9222),
                    stop_event=self.stop_event,
                    headless=headless,
                )
                try:
                    tungee_records = crawler.crawl_by_keywords(company_names)
                    log_finish(tungee_log_id, len(tungee_records), 'success')
                    self._log(f"探迹补全完成: {len(tungee_records)} 条")
                finally:
                    crawler.quit()
            except Exception as e:
                log_finish(tungee_log_id, 0, 'failed', str(e))
                self._log(f"探迹出错: {e}")

        threads = []
        if self.var_tianyancha.get() and company_names and not self.stop_event.is_set():
            self._log(f"=== Phase 3: 天眼查补全: {len(company_names)} 个公司 ===")
            t = threading.Thread(target=run_tianyancha, daemon=True)
            threads.append(t)
        if self.var_tungee.get() and company_names and not self.stop_event.is_set():
            self._log(f"=== Phase 4: 探迹补全: {len(company_names)} 个公司 ===")
            t = threading.Thread(target=run_tungee, daemon=True)
            threads.append(t)

        for t in threads:
            t.start()
        for t in threads:
            t.join()

        # 串行更新 DB
        if tianyancha_records:
            self._update_records_by_name(tianyancha_records)
        if tungee_records:
            self._update_records_by_name(tungee_records)

        self._log("全部抓取完成")

        # 打印本次抓取统计摘要
        try:
            conn = sqlite3.connect(str(DB_FILE))
            c = conn.cursor()
            total = c.execute("SELECT COUNT(*) FROM companies").fetchone()[0] or 0
            self._log(f"总记录: {total} 条")
            # 各字段填充率
            fields = ['project_unit', 'contact_person', 'phone', 'address',
                      'main_company', 'investor', 'project_desc']
            for f in fields:
                cnt = c.execute(f"SELECT COUNT(*) FROM companies WHERE {f} IS NOT NULL AND {f} != ''").fetchone()[0] or 0
                pct = cnt / total * 100 if total > 0 else 0
                self._log(f"  {f}: {cnt}/{total} ({pct:.0f}%)")
            conn.close()
        except Exception as e:
            self._log(f"统计失败: {e}")

        self.running = False
        self.root.after(0, lambda: self.btn_start.config(state=NORMAL))
        self.root.after(0, lambda: self.btn_stop.config(state=DISABLED))
        self.root.after(0, self._refresh_stats)

    def _update_records_by_name(self, records: list[dict]):
        """按 company_name 更新已有记录（只更新探迹字段）"""
        if not records:
            return
        conn = sqlite3.connect(str(DB_FILE))
        c = conn.cursor()
        updated = 0
        for rec in records:
            company = rec.get('company_name', '')
            if not company:
                continue
            # COALESCE: 新值非空则覆盖，否则保留旧值
            c.execute("""
                UPDATE companies SET
                    contact_person = COALESCE(NULLIF(?, ''), contact_person),
                    phone = COALESCE(NULLIF(?, ''), phone),
                    address = COALESCE(NULLIF(?, ''), address),
                    main_company = COALESCE(NULLIF(?, ''), main_company),
                    main_contact = COALESCE(NULLIF(?, ''), main_contact),
                    main_phone = COALESCE(NULLIF(?, ''), main_phone),
                    main_address = COALESCE(NULLIF(?, ''), main_address),
                    investor = COALESCE(NULLIF(?, ''), investor),
                    credit_code = COALESCE(NULLIF(?, ''), credit_code),
                    relation_graph = COALESCE(NULLIF(?, ''), relation_graph),
                    remark = COALESCE(NULLIF(?, ''), remark),
                    summary = COALESCE(NULLIF(?, ''), summary),
                    project_desc = COALESCE(NULLIF(?, ''), project_desc),
                    project_status = COALESCE(NULLIF(?, ''), project_status)
                WHERE company_name = ?
            """, (
                rec.get('contact_person', ''),
                rec.get('phone', ''),
                rec.get('address', ''),
                rec.get('main_company', ''),
                rec.get('main_contact', ''),
                rec.get('main_phone', ''),
                rec.get('main_address', ''),
                rec.get('investor', ''),
                rec.get('credit_code', ''),
                rec.get('relation_graph', ''),
                rec.get('remark', ''),
                rec.get('summary', ''),
                rec.get('project_desc', ''),
                rec.get('project_status', ''),
                company,
            ))
            updated += c.rowcount
        conn.commit()
        conn.close()
        self._log(f"DB 更新完成: {updated} 条")


# ============================================================
# 入口
# ============================================================
def main():
    init_db()
    root = tk.Tk()
    App(root)
    root.mainloop()

if __name__ == '__main__':
    main()
